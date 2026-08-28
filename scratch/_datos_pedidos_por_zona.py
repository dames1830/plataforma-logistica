# -*- coding: utf-8 -*-
"""QUE PEDIDOS APUNTAN A QUE ZONA.

Daniel, 27-ago-2026: *"la ola se corre en el WMS cada treinta pedidos. Imaginate correr mil
pedidos en la zona buffer, o mil en el mezzanine uno: es practicamente hora muerta para un
asistente. Quiero saber exactamente que pedidos estan apuntando a cada zona"*.

La cuenta es la misma ola de Picking Hoy, pero sin juntar por marca: se guarda el NUMERO DE
ORDEN. Cada pedido termina con la lista de zonas que de verdad tiene que visitar, y cada
zona con la lista de pedidos que la mandan a buscar.

DE DONDE SALE CADA COSA -- y es la misma regla que usa el robot `armar_pendiente.py`, no
una nueva:

    Detalle Orden Pendientes.csv   el detalle linea por linea: orden, SKU, solicitada,
    (OneDrive, 27-ago 06:59)       asignada, tienda y destino
            |
            +-- se queda solo con estado 'Creada' o 'Parcialmente asignado'
            +-- se descartan las lineas repetidas (misma orden, SKU y destino)
            +-- SOLO LAS GUIAS QUE MANDO COMERCIAL: las que estan en los correos.
            |   *"pueden haber un millon de pedidos en el WMS, pero solamente vamos a tener
            |   lo que dice el correo de comercial"*
            +-- pendiente = solicitada - asignada
                    |
                    +-- se busca en las zonas, en el orden del almacen
"""
import csv, io, json, os, re, sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
try:
    import openpyxl
except ImportError:
    raise SystemExit('Falta openpyxl:  pip install openpyxl')

AQUI = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(os.path.expanduser('~'), 'OneDrive', 'danielames.bata', 'scraping Stock')
DETALLE = os.path.join(BASE, 'Detalle Orden', 'Detalle Orden Pendientes.csv')
CORREOS = os.path.join(BASE, 'Correos Picking')
# LA RUTA DE DESPACHO DE CADA TIENDA. Daniel, 27-ago-2026: *"el codigo de la tienda te sale
# en tres digitos; agregale 50 delante para que sea de cinco, y con eso lo verificas en el
# Excel de rutas"*. Comprobado sobre el correo del 27-ago: cruzan las 90 tiendas, 100%, y
# el nombre coincide en las 90.
RUTAS_XLSX = os.path.join(os.path.dirname(os.path.dirname(BASE)),
                          'danielames.bata', 'Proyecto web Logistico', 'RUTAS -  TURNOS.xlsx')
ESTADOS = ('Creada', 'Parcialmente asignado')
POR_OLA = 30          # cada cuantos pedidos corre una ola en el WMS

ORDEN = ['CDBUF_AB', 'CDBUF_C', 'CDBUF_X', 'SEL', 'MZN01', 'MZN02', 'MZN03', 'MZN04',
         'AND', 'PARED']
ETIQUETA = {'CDBUF_AB': 'Buffer A + B', 'CDBUF_C': 'Buffer C',
            'CDBUF_X': 'Buffer (otras letras)', 'SEL': 'Selectivo', 'MZN01': 'Mezzanine 1',
            'MZN02': 'Mezzanine 2', 'MZN03': 'Mezzanine 3', 'MZN04': 'Mezzanine 4',
            'AND': 'Andamios', 'PARED': 'Pared'}
RX_PREPACK = re.compile(r'^\d{7}-\d-(\d{5})$')


def limpio(v):
    """EL WMS EXPORTA ENVUELTO COMO FORMULA: ="8003594". El correo lo escribe pelado.

    Sin quitar la envoltura el cruce da 0% y parece que las guias no calzan. Esta anotado
    en `armar_pendiente.py` -le costo una vuelta entera a Daniel el 19-ago-2026- y aun asi
    volvi a caer: la primera corrida dio 0 pedidos y 69.581 lineas "fuera del correo".
    """
    return re.sub(r'^="?|"?$', '', str(v if v is not None else '').strip()).strip()


def pares_por_caja(sku):
    """Los dos primeros digitos del sufijo son los pares que trae la caja."""
    m = RX_PREPACK.match(str(sku))
    if not m:
        return 1
    try:
        n = int(m.group(1)[:2])
    except Exception:
        return 1
    return n if n > 0 else 1


def zona_de(area, ubicacion):
    """El buffer se parte por la letra de su ubicacion; el resto es su area."""
    if area != 'CDBUFFER':
        return area
    p = str(ubicacion).strip().upper().split('-')
    letra = p[1] if len(p) > 1 else ''
    return 'CDBUF_AB' if letra in ('A', 'B') else 'CDBUF_C' if letra == 'C' else 'CDBUF_X'


# ── LAS GUIAS QUE MANDO COMERCIAL ───────────────────────────────────────────
# Una guia puede venir en dos correos y manda la PRIMERA. La hoja buena no es siempre la
# primera del libro: se busca la que tenga la columna GUIA en su cabecera.
def leer_guias():
    """Las guias del correo, con su ficha: cadena, tienda, prioridad y fecha.

    Una guia puede venir en dos correos y manda la PRIMERA vez: asi conserva la prioridad
    y la fecha de cuando de verdad la mandaron.
    """
    guias = {}
    archivos = sorted(n for n in os.listdir(CORREOS) if n.lower().endswith(('.xlsx', '.xls')))
    leidos = 0
    for n in archivos:
        try:
            wb = openpyxl.load_workbook(os.path.join(CORREOS, n), read_only=True, data_only=True)
        except Exception:
            continue
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            try:
                cab = [str(c).strip() if c is not None else '' for c in next(it)]
            except StopIteration:
                continue
            if 'GUIA' not in cab:
                continue
            ig = cab.index('GUIA')
            col = {n: cab.index(n) for n in ('Cadena', 'TIEND', 'NOMBR', 'Prioridad', 'FECHA')
                   if n in cab}
            for fila in it:
                if ig >= len(fila) or fila[ig] is None:
                    continue
                g = limpio(fila[ig])
                if not g or g in guias:       # la primera vez manda
                    continue
                dame = lambda k: (str(fila[col[k]]).strip()
                                  if k in col and col[k] < len(fila) and fila[col[k]] is not None
                                  else '')
                # EL CODIGO DE 3 DIGITOS PASA A 5 CON EL '50' DELANTE, y con ese se busca
                # la ruta. Es la regla de Daniel y cruza el 100% de las tiendas.
                cod3 = dame('TIEND')
                cdg = ('50' + cod3.zfill(3)) if cod3 else ''
                r = RUTAS.get(cdg) or {}
                guias[g] = {'cadena': dame('Cadena'), 'tienda': cod3, 'cdg': cdg,
                            'nombre': dame('NOMBR'), 'prioridad': dame('Prioridad'),
                            'fecha': dame('FECHA')[:10], 'ruta': r.get('ruta', ''),
                            'zonaRuta': r.get('zona', ''), 'turno': r.get('turno', ''),
                            'diaRuta': r.get('dia', '')}
            leidos += 1
            break
        wb.close()
    return guias, leidos, len(archivos)


def leer_rutas():
    """CDG de 5 digitos -> ruta, zona, turno y dia de despacho."""
    if not os.path.isfile(RUTAS_XLSX):
        print('   AVISO: no se encontro el Excel de rutas, los pedidos van sin ruta')
        return {}
    ws = openpyxl.load_workbook(RUTAS_XLSX, read_only=True, data_only=True)['Hoja1']
    it = ws.iter_rows(values_only=True)
    next(it, None)
    r = {}
    for f in it:
        if f and f[0] is not None:
            r[str(f[0]).strip()] = {'tienda': str(f[1] or '').strip(),
                                    'zona': str(f[2] or '').strip(),
                                    'turno': str(f[5] or '').strip(),
                                    'ruta': str(f[6] or '').strip(),
                                    'dia': str(f[7] or '').strip()}
    return r


RUTAS = leer_rutas()
print('Rutas de despacho: %s tiendas · %s rutas distintas'
      % (format(len(RUTAS), ','), len(set(v['ruta'] for v in RUTAS.values() if v['ruta']))))
print('Leyendo los correos de comercial...')
GUIAS, leidos, total = leer_guias()
print('   %s correos leidos de %s · %s guias distintas' % (leidos, total, format(len(GUIAS), ',')))

# ── DONDE VIVE CADA SKU ─────────────────────────────────────────────────────
act = json.load(io.open(os.path.join(AQUI, '_activo_ancla_manana.json'), encoding='utf-8'))
donde = defaultdict(lambda: defaultdict(float))
for f in act['data']:
    a = str(f.get('Área') or '').strip().upper()
    z = zona_de(a, f.get('Ubicación'))
    if z not in ORDEN:
        continue
    sku = str(f.get('Artículo') or '').strip()
    try:
        q = float(str(f.get('Cantidad actual') or 0).replace(',', '') or 0)
    except Exception:
        continue
    if sku and q > 0:
        donde[sku][z] += q

# ── EL DETALLE, LINEA POR LINEA ─────────────────────────────────────────────
csv.field_size_limit(10 ** 7)
pedidos = defaultdict(lambda: {'zonas': defaultdict(float), 'pend': 0.0, 'sinUbicar': 0.0,
                               'tienda': '', 'cadena': '', 'prioridad': '', 'fecha': '',
                               'cdg': '', 'ruta': '', 'zonaRuta': '', 'turno': '', 'diaRuta': '',
                               'lineas': 0, 'skus': set()})
porSku = defaultdict(lambda: [0.0, 0.0])
vistas = set()
leidas = fuera = repetidas = 0
with io.open(DETALLE, encoding='utf-8-sig', newline='') as fh:
    r = csv.reader(fh, delimiter=';')
    next(r, None)
    for row in r:
        if len(row) < 14 or row[4].strip() not in ESTADOS:
            continue
        orden, sku, dest = limpio(row[1]), limpio(row[5]), limpio(row[13])
        if (orden, sku, dest) in vistas:
            repetidas += 1
            continue
        vistas.add((orden, sku, dest))
        if orden not in GUIAS:
            fuera += 1
            continue
        try:
            sol, asig = float(row[6] or 0), float(row[9] or 0)
        except Exception:
            continue
        # EL MISMO DETALLE ALIMENTA LAS DOS PANTALLAS. Antes Picking Hoy leia el area
        # `buffer` del servidor -armada a las 19:18 de anoche- y esta seccion leia el
        # Detalle Orden de las 06:59. Dos totales distintos en una sola pantalla se leen
        # como un error, y con razon. Ahora sale todo de aca, del mismo momento.
        #
        # SE SUMA ANTES DE DESCARTAR LAS LINEAS YA SERVIDAS, igual que `armar_pendiente.py`.
        # Poniendolo despues, lo asignado daba 443 en vez de decenas de miles: solo contaba
        # la asignacion de las lineas que ademas tenian pendiente.
        porSku[sku][0] += sol
        porSku[sku][1] += asig
        pend = sol - asig
        if pend <= 0:
            continue
        leidas += 1
        p = pedidos[orden]
        p['pend'] += pend
        p['lineas'] += 1
        p['skus'].add(sku)
        if not p['tienda']:
            g = GUIAS[orden]
            p['tienda'] = (g['nombre'] or g['tienda'] or dest)
            p['cadena'] = g['cadena']
            p['prioridad'] = g['prioridad']
            p['fecha'] = g['fecha']
            p['cdg'] = g['cdg']
            p['ruta'] = g['ruta']
            p['zonaRuta'] = g['zonaRuta']
            p['turno'] = g['turno']
            p['diaRuta'] = g['diaRuta']
        # LA OLA: se busca en el orden del almacen y se toma lo que cada zona tenga.
        queda = pend
        for z in ORDEN:
            if queda <= 0:
                break
            hay = (donde.get(sku) or {}).get(z, 0)
            if hay <= 0:
                continue
            toma = min(queda, hay)
            p['zonas'][z] += toma * pares_por_caja(sku)
            queda -= toma
        p['sinUbicar'] += queda

# ── SE ARMA LA SALIDA ───────────────────────────────────────────────────────
porZona = defaultdict(lambda: {'pedidos': [], 'pares': 0.0})
filas = []
for orden, p in pedidos.items():
    zs = sorted(((z, v) for z, v in p['zonas'].items() if v > 0),
                key=lambda x: ORDEN.index(x[0]))
    filas.append({'orden': orden, 'tienda': p['tienda'], 'cadena': p['cadena'],
                  'nombre': GUIAS.get(orden, {}).get('nombre', ''),
                  'cdg': p['cdg'], 'ruta': p['ruta'], 'zonaRuta': p['zonaRuta'],
                  'turno': p['turno'], 'diaRuta': p['diaRuta'],
                  'prioridad': p['prioridad'], 'fecha': p['fecha'], 'lineas': p['lineas'],
                  'skus': len(p['skus']), 'pend': round(p['pend']),
                  'sinUbicar': round(p['sinUbicar']),
                  'zonas': [{'z': z, 'etiqueta': ETIQUETA[z], 'pares': round(v)} for z, v in zs],
                  'nZonas': len(zs)})
    for z, v in zs:
        porZona[z]['pedidos'].append(orden)
        porZona[z]['pares'] += v

filas.sort(key=lambda f: -f['pend'])
zonas = [{'z': z, 'etiqueta': ETIQUETA[z], 'pedidos': len(porZona[z]['pedidos']),
          'pares': round(porZona[z]['pares']),
          'olas': -(-len(porZona[z]['pedidos']) // POR_OLA),
          'top': sorted(({'orden': o} for o in porZona[z]['pedidos'][:0]), key=lambda x: x['orden'])}
         for z in ORDEN if porZona.get(z)]

# Cuantas zonas visita un pedido: es lo que decide si la ola vale la pena
reparto = defaultdict(int)
for f in filas:
    reparto[f['nZonas']] += 1

# ── POR RUTA DE DESPACHO ────────────────────────────────────────────────────
# Una ruta despacha a varias tiendas y sale junta. Agrupar la ola POR RUTA en vez de por
# zona sale mas barato -medido sobre hoy: 88 olas contra 97- y ademas cada ruta se despacha
# completa, sin dejar pedidos a medias.
porRuta = defaultdict(lambda: {'ped': 0, 'pares': 0, 'tiendas': set(),
                               'zonas': defaultdict(int), 'zona': '', 'turno': '', 'dia': ''})
for f in filas:
    d = porRuta[f['ruta'] or '(sin ruta)']
    d['ped'] += 1
    d['pares'] += f['pend']
    d['tiendas'].add(f['cdg'])
    d['zona'] = d['zona'] or f['zonaRuta']
    d['turno'] = d['turno'] or f['turno']
    d['dia'] = d['dia'] or f['diaRuta']
    for z in f['zonas']:
        d['zonas'][z['z']] += 1
rutas = sorted(({'ruta': k, 'pedidos': v['ped'], 'pares': v['pares'],
                 'tiendas': len(v['tiendas']), 'zona': v['zona'], 'turno': v['turno'],
                 'dia': v['dia'], 'olas': -(-v['ped'] // POR_OLA),
                 'zonas': sorted(({'z': z, 'etiqueta': ETIQUETA[z], 'pedidos': n}
                                  for z, n in v['zonas'].items()), key=lambda x: -x['pedidos'])}
                for k, v in porRuta.items()), key=lambda r_: -r_['pedidos'])

datos = {
    'fechaDetalle': '2026-08-27 06:59',
    'porOla': POR_OLA,
    'totales': {'pedidos': len(filas), 'pares': round(sum(f['pend'] for f in filas)),
                'lineas': leidas, 'repetidas': repetidas, 'fueraDelCorreo': fuera,
                'guias': len(GUIAS), 'correos': leidos},
    'zonas': zonas,
    'rutas': rutas,
    'olasPorRuta': sum(r_['olas'] for r_ in rutas),
    'olasPorZona': sum(z['olas'] for z in zonas),
    'reparto': [{'zonas': k, 'pedidos': reparto[k]} for k in sorted(reparto)],
    'pedidos': filas,
}
io.open(os.path.join(AQUI, '_pedidos_por_zona.json'), 'w', encoding='utf-8', newline='').write(
    json.dumps(datos, ensure_ascii=False))

# El pedido por SKU, en el formato del area `buffer`, para que Picking Hoy lea lo mismo.
io.open(os.path.join(AQUI, '_pedido_del_detalle.json'), 'w', encoding='utf-8', newline='').write(
    json.dumps({'area': 'buffer', 'updated_at': datos['fechaDetalle'],
                'origen': 'Detalle Orden Pendientes.csv',
                'data': [{'Código de artículo': k, 'Cantidad solicitada': v[0],
                          'Cantidad asignada': v[1]} for k, v in porSku.items()]},
               ensure_ascii=False))

F = lambda x: format(int(x), ',')
t = datos['totales']
print('\nEL DETALLE DE ORDEN de %s' % datos['fechaDetalle'])
print('   lineas que cuentan     %10s' % F(t['lineas']))
print('   repetidas descartadas  %10s' % F(t['repetidas']))
print('   fuera del correo       %10s   <- el WMS las tiene, comercial no las mando' % F(t['fueraDelCorreo']))
print('   PEDIDOS                %10s   ·  %s pares' % (F(t['pedidos']), F(t['pares'])))
print()
print('%-24s %9s %10s %8s   %s' % ('ZONA', 'PEDIDOS', 'PARES', 'OLAS', 'de %d' % POR_OLA))
for z in zonas:
    print('%-24s %9s %10s %8s' % (z['etiqueta'], F(z['pedidos']), F(z['pares']), F(z['olas'])))
print('%-24s %9s' % ('(un pedido puede estar en varias)', F(sum(z['pedidos'] for z in zonas))))
print()
print('LAS RUTAS: %d distintas   ·   olas por ruta %d   ·   olas por zona %d'
      % (len(rutas), datos['olasPorRuta'], datos['olasPorZona']))
print()
print('CUANTAS ZONAS VISITA UN PEDIDO')
for r_ in datos['reparto']:
    print('   %d zona%s  %6s pedidos' % (r_['zonas'], 's' if r_['zonas'] != 1 else ' ',
          F(r_['pedidos'])))
