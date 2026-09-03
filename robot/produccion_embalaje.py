# -*- coding: utf-8 -*-
"""
ROBOT: el cuadro de EMBALAJE POR DIA (persona x hora, canal y efectividad)

Va ENGANCHADO DETRAS del robot que ya baja el OBLPN de embalaje del WMS cada 2 horas
-`ejecutar_oblpn_hora.bat`-, asi que NO entra al WMS ni descarga nada: lee el archivo
que ese acaba de dejar y publica el cuadro. Daniel, 02-sep-2026: *"ese picking por
hora es el que tienes que agarrar para el modulo de picking dia"*.

Por eso no pide turno al candado del WMS y no puede chocar con ninguna otra
corrida: lo unico que hace es leer un CSV de OneDrive y mandar un JSON.

SE LLAMA `produccion_embalaje.py` Y NO `oblpn_embalaje.py` A PROPOSITO: en el servidor ya
existe un `oblpn_embalaje.py` -el que baja del WMS-, y dos archivos con el
mismo nombre en la carpeta del robot se pisan. El AREA de la plataforma si se llama
`embalaje_por_hora`, que es lo que leen las pantallas.

Publica en el area `embalaje_por_hora`, fechada con el DIA DEL ARCHIVO -no con la
fecha de hoy-: el ultimo pase del dia es a las 20:40 y con la hora de la maquina
quedaria estampado el dia siguiente.

PERSONA x HORA DEL EMBALAJE (OBLPN): volumen, efectividad y canal.

La misma pantalla que Picking por dia, pero sobre `OBLPN Embalaje\\OBLPN DD-MM.csv`.
Daniel, 01-sep-2026: *"igualito, tal cual figura aca, no cambia nada"*.

QUE CUENTA COMO EMBALADO: la linea cuyo **`Registro de hora de empaquetado`** cae
en el dia Y CUYO NUMERO DE LPN NO EMPIEZA CON "PRE".

EL TRUCO DEL LPN, que cazó Daniel el 01-sep-2026: *"el estado dice empaquetado
pero el LPN inicia como PRE; eso quiere decir que todavia no esta embalado"*. El
`PRE...` es la PRE-ETIQUETA, no la caja cerrada; la caja de verdad tiene numero
puro. El archivo lo confirma solo:

    PREFIJO        Cancelado  Empaquetado    Enviado
    (sin letras)          61        6.748     12.592
    PRE               11.553        5.349        862

11.553 de los 11.629 cancelados son PRE. Sin esta regla el reporte contaba 11.728
pares de mas -el 29%- y 20 personas que en realidad no embalaron nada.

Tambien se saca `WPRE`, que es la misma familia (35 lineas).

No se filtra por `Estado de LPN`: tienen que venir Empaquetado, Cargado y Enviado
juntos, que es la gracia de comparar los escalones.

EL EMBALADOR ES `Usuario de modificacion de ASIGNACION`, y no se eligio a dedo:
se probaron LAS SIETE columnas de usuario del OBLPN contra el web report del WMS
-`PRODUCCION EMBALAJE ALDEAS X HORA acc calz`, que Daniel da por bueno- y se midio
cual reproduce su reparto por persona. El 31-08, en pares de retail:

    Usuario de modificacion de asignacion      264   <- gana
    Usuario de modificacion de LPN          20.530
    Usuario de carga                        20.542
    Usuario de seleccion                    39.562
    Usuario de paquete                      42.108
    Detail Pick User                        48.626

Ocho de doce personas dan EXACTO y el error total es el 1%.

Daniel habia dicho *"el embalador es el que modifica"* y tenia razon; lo que yo
tome mal fue CUAL de las dos columnas de modificacion. La del LPN se equivoca en
20.530 pares.

NO SIRVE `Usuario de paquete`: viene VACIO en el 58% de las lineas y donde SI
viene coincide al 100% con `Usuario de seleccion`, o sea que ahi el WMS pone al
que pickeo.

Ojo: el archivo tiene **`Usuario de seleccion` DOS VECES**, asi que se lee por
INDICE de columna y no con DictReader, que se queda con la ultima.

EL RESTO DE LAS REGLAS son las mismas del picking: el prepack cuenta CAJAS y sus
pares salen de la curva del SKU; el calzado lo corta `G. Gender` del Maestro; y
el canal sale del maestro de rutas + `Tipo de orden`.
"""
import csv
import io
import json
import os
import re
import shutil
import sys
from collections import defaultdict

import openpyxl

csv.field_size_limit(10 ** 7)

FORMA_PREPACK = re.compile(r'^\d{7}-\d-\d{5}$')
# LAS VEINTICUATRO HORAS, no solo el turno.
#
# Estaba en 8..19 y lo que se movia fuera de ahi se contaba en el total pero NO
# tenia columna donde aparecer: el 28-ago quedaron 275 lineas de picking y 217 de
# embalaje sin fila que las mostrara, y la suma de las horas no daba el total.
#
# Daniel, 02-sep-2026: *"el noventa y cinco por ciento se mueve entre ocho de la
# manana y las siete de la noche, pero hay un minimo que se mueve en la madrugada.
# Necesito las veinticuatro horas"*.
#
# La pantalla no dibuja las 24 siempre: muestra el turno completo y agrega solo
# las horas de afuera que ese dia tuvieron movimiento.
HORAS = list(range(0, 24))
CL = ('cal_suelto', 'cal_prepack', 'no_cal', 'materiales', 'sin_tipo')
TODOS = 'TODOS'
ORDEN_CANAL = ['RETAIL', 'MAYORISTA', 'CATALOGO', 'ECOMMERCE', 'INDUSTRIAL',
               'OTROS', 'SIN CANAL']

LINEAS_MIN_CELDA = 8
LINEAS_MIN_DIA = 20
SEG_MIN_CELDA = 5 * 60
SEG_MIN_DIA = 15 * 60
SEG_LINEA_MIN = 5
SEG_MUESTRA_CORTA = 60 * 60
# DOS MARCAS PEGADAS SON LA MISMA TANDA. En embalaje el WMS estampa un solo
# instante por caja, asi que sin este puente cada tarea duraria cero y el
# ritmo se iria al cielo. Un hueco mayor que esto si es tiempo parado.
PUENTE_SEG = 15 * 60


def base_onedrive():
    for c in (os.environ.get('OneDrive'),
              'C:' + os.sep + os.path.join('Users', 'Administrator', 'OneDrive'),
              'C:' + os.sep + os.path.join('Users', 'dames', 'OneDrive')):
        if c:
            r = os.path.join(c, 'danielames.bata', 'scraping Stock')
            if os.path.isdir(r):
                return r
    raise SystemExit('no encuentro OneDrive')


def dia_pedido():
    """El dia que vino por `--dia AAAA-MM-DD`, o None.

    EN EL RELLENO EL DIA LO MANDA QUIEN LLAMA. Los archivos viejos del OBLPN
    mezclan hasta doce fechas —el WMS mete lineas empaquetadas antes— y la
    mayoria no siempre es la del archivo: el `OBLPN 01-08.csv` daba 30 de julio.
    El que rellena si sabe de que dia es cada archivo, porque lo dice el nombre.

    En la corrida normal de cada 2 horas no se pasa nada y sigue mandando la
    mayoria de las filas, que ahi es lo correcto.
    """
    a = sys.argv[1:]
    if '--dia' in a:
        i = a.index('--dia')
        if len(a) > i + 1 and re.match(r'^\d{4}-\d{2}-\d{2}$', a[i + 1]):
            return a[i + 1]
    return None


def dia_mayoritario(ruta, columna):
    """EL DIA DEL ARCHIVO ES EL DE LA MAYORIA DE SUS FILAS, no el de la primera.

    Salia de la primera fila con hora valida, y eso es fragil: el OBLPN del
    01-09 traia arriba una linea empaquetada el 31-08 y el cuadro entero quedo
    fechado el 31, tirando las 39.000 filas del dia bueno a `sin hora`. Se vio el
    02-sep-2026 en la primera corrida del robot.

    Se cuenta con una pasada liviana -solo esa columna, sin armar diccionarios-
    y gana la fecha que mas veces aparece. Devuelve AAAA-MM-DD, o None si el
    archivo no tiene ni una fecha legible.
    """
    from collections import Counter
    try:
        f = io.open(ruta, encoding='utf-8-sig', newline='', errors='replace')
        cabeza = f.read(4000)
        f.seek(0)
        r = csv.reader(f, delimiter=';' if cabeza.count(';') > cabeza.count(',') else ',')
        cab = [c.strip() for c in next(r)]
        try:
            i = cab.index(columna)
        except ValueError:
            f.close()
            return None
        cuenta = Counter()
        for x in r:
            if i < len(x):
                mm = re.match(r'^(\d{2})/(\d{2})/(\d{4})\s', str(x[i]).strip().strip('"'))
                if mm:
                    cuenta['%s-%s-%s' % (mm.group(3), mm.group(2), mm.group(1))] += 1
        f.close()
        if not cuenta:
            return None
        gana, veces = cuenta.most_common(1)[0]
        if len(cuenta) > 1:
            print('[AVISO] el archivo mezcla %d fechas; gana %s con %d filas de %d'
                  % (len(cuenta), gana, veces, sum(cuenta.values())))
        return gana
    except Exception as e:
        print('[AVISO] no se pudo mirar la fecha del archivo: %s' % e)
        return None


def elegir_archivo(carpetas, plantillas):
    """QUE ARCHIVO LE TOCA A ESTA CORRIDA.

    Se puede pasar el nombre a mano -util para rehacer un dia viejo-. Sin eso,
    se busca el de HOY, que es el que el robot de la hora acaba de dejar.

    EL NOMBRE DEL DIA NO ES UNO SOLO: el picking lo escribe sin cero adelante
    ("Picking 3-9.csv") y el OBLPN con cero ("OBLPN 03-09.csv"). Se prueban las
    dos formas antes de rendirse.

    Si el de hoy no esta -el WMS no contesto, o es de madrugada y todavia no
    corrio ningun pase- se toma EL MAS NUEVO de la carpeta y se avisa. Es mejor
    republicar el cuadro de ayer que dejar la pantalla sin nada.
    """
    if isinstance(carpetas, str):
        carpetas = [carpetas]
    if len(sys.argv) > 1 and not sys.argv[1].startswith('-'):
        for c in carpetas:
            r = os.path.join(c, sys.argv[1])
            if os.path.isfile(r):
                return r
        return os.path.join(carpetas[0], sys.argv[1])
    hoy = __import__('datetime').datetime.now()
    for c in carpetas:
        for pl in plantillas:
            r = os.path.join(c, pl % (hoy.day, hoy.month))
            if os.path.isfile(r):
                return r
    cand = []
    for c in carpetas:
        try:
            cand += [os.path.join(c, n) for n in os.listdir(c)
                     if n.lower().endswith('.csv')]
        except Exception:
            pass
    if cand:
        nuevo = max(cand, key=os.path.getmtime)
        print('[AVISO] no hay archivo de hoy; se usa el mas nuevo: %s'
              % os.path.basename(nuevo))
        return nuevo
    return os.path.join(carpetas[0], plantillas[0] % (hoy.day, hoy.month))

BASE = base_onedrive()
# EL OBLPN SI LO DEJA EL ROBOT DE LA HORA EN ONEDRIVE, y con el dia en el nombre:
# `oblpn_embalaje.py --hoy` lo pisa en cada pase, asi que ahi esta siempre el
# ultimo estado del dia. Una sola carpeta, sin respaldo que buscar.
ARCHIVO = elegir_archivo(os.path.join(BASE, 'OBLPN Embalaje'),
                         ['OBLPN %02d-%02d.csv', 'OBLPN %d-%d.csv'])
CARPETA_ORD = os.path.join(BASE, 'Detalle Orden')
MAESTROS = [os.path.join(os.path.dirname(BASE), 'Maestro_Articulos.xlsx'),
            os.path.join(BASE, 'Archivos', 'Maestro_Articulos.xlsx')]
RUTAS_CAND = [os.path.join(os.path.dirname(BASE), 'Proyecto web Logistico',
                           'RUTAS -  TURNOS.xlsx'),
              os.path.join('C:' + os.sep, 'wms_scraping', '_rutas.xlsx')]
# EL AREA Y EL ARCHIVO SON LO MISMO, escrito una sola vez. Estuvieron sueltos y
# el robot de embalaje quedo escribiendo en el JSON del picking: el cruce comparo
# el web report de picking contra los numeros del embalaje y dio 3.480 de
# diferencia sin que nada pareciera roto.
AREA = 'embalaje_por_hora'
SALIDA = os.path.join('C:' + os.sep, 'wms_scraping', 'logs', AREA + '.json')


def limpio(v):
    s = str(v if v is not None else '').strip()
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1].strip()
    return s.strip('"').strip()


def entero(v):
    try:
        return int(float(str(v).replace(',', '.')))
    except (TypeError, ValueError):
        return 0


def es_prepack(sku):
    return bool(FORMA_PREPACK.match(str(sku or '').strip()))


# ── LOS TIPOS QUE PIDIO DANIEL ───────────────────────────────────────────────
#
# Daniel, 02-sep-2026: *"solamente calzado y no calzado. En no calzado entra todo
# lo que son bolsas, etiquetas, etcetera"*. Y aparte, para lo que el Maestro no
# conoce: *"ponle materiales porque si son solamente cinco digitos, es material,
# si no me equivoco. Revisa la descripcion que es"*.
#
# SE REVISO ANTES DE CREER LA CORAZONADA, leyendo la descripcion que trae el
# propio archivo del WMS. Los 25 codigos que no estaban en el Maestro son TODOS
# de cinco digitos y TODOS material:
#
#     70104   74.768 u   TISSUE PAPER BATA N 4
#     69050   18.695 u   HANG TAG ORTHOLITE BATA ROJO
#     70103   16.795 u   TISSUE PAPER BATA N 3
#     88424    2.575 u   CAJA MICROC. KRAFT BATA N.24
#     26036      540 u   PLANT LAURA C/GEL ROJO N. 36
#
# Papel de seda, etiquetas colgantes, cajas de carton y plantillas. Tenia razon.
#
# LA REGLA EXIGE LAS DOS COSAS -cinco digitos Y no estar en el Maestro- y no solo
# la primera. Si manana falta en el Maestro un articulo de calzado de verdad, cae
# en "sin tipo" y se ve; llamarlo "materiales" sin conocerlo seria inventar.
# `sin_tipo` es la respuesta honesta de "no se que es esto", y ademas es la lista
# de lo que hay que agregarle al Maestro.
CINCO_DIGITOS = re.compile(r'^\d{5}$')


def tipo_de(sku, gender, esta_en_el_maestro):
    """calzado suelto / calzado prepack / no calzado / materiales / sin tipo."""
    if not esta_en_el_maestro:
        return 'materiales' if CINCO_DIGITOS.match((sku or '')[:7]) else 'sin_tipo'
    if not gender or gender == 'Sin dato':
        return 'sin_tipo'
    if gender != 'Footwear':
        return 'no_cal'
    return 'cal_prepack' if es_prepack(sku) else 'cal_suelto'


def pares_de_la_caja(sku):
    s = str(sku or '').strip()
    if not FORMA_PREPACK.match(s):
        return 1
    try:
        n = int(s[-5:][:2])
    except ValueError:
        return 1
    return n if 0 < n <= 24 else 1


def abrir(ruta):
    f = io.open(ruta, encoding='utf-8-sig', newline='', errors='replace')
    cabeza = f.read(4000)
    f.seek(0)
    sep = ';' if cabeza.count(';') > cabeza.count(',') else ','
    return f, csv.DictReader(f, delimiter=sep)


# ── el Maestro de articulos ─────────────────────────────────────────────
ruta_m = next((r for r in MAESTROS if os.path.isfile(r)), None)
wb = openpyxl.load_workbook(ruta_m, read_only=True, data_only=True)
it = wb.worksheets[0].iter_rows(values_only=True)
cab = [str(c).strip() if c is not None else '' for c in next(it)]


def colm(*nombres):
    bajos = [n.lower() for n in nombres]
    for i, c in enumerate(cab):
        if c.lower() in bajos:
            return i
    return -1


iS = colm('CodArticulo', 'CodigoArticulo')
iG, iM, iC = colm('G. Gender', 'G Gender'), colm('Marcas', 'MarcaStd'), colm('Coleccion PO')
maestro = {}
for f in it:
    if iS < 0 or iS >= len(f) or f[iS] is None:
        continue
    k = limpio(f[iS])[:7]
    if k and k not in maestro:
        def dd(i):
            v = limpio(f[i]) if 0 <= i < len(f) else ''
            return v if v and v != '(en blanco)' else 'Sin dato'
        maestro[k] = (dd(iG), dd(iM), dd(iC))
wb.close()
print('maestro de articulos: %d codigos' % len(maestro))

# ── el maestro de RUTAS ─────────────────────────────────────────────────
ruta_r = next((r for r in RUTAS_CAND if os.path.isfile(r)), None)
copia = os.path.join('C:' + os.sep, 'wms_scraping', 'logs', '_rutas_emb.xlsx')
try:
    shutil.copyfile(ruta_r, copia)
except Exception:
    copia = ruta_r
wb = openpyxl.load_workbook(copia, read_only=True, data_only=True)
it = wb.worksheets[0].iter_rows(values_only=True)
cr = [str(c).strip() if c is not None else '' for c in next(it)]
kC = cr.index('CDG')
tiendas = {str(f[kC]).strip() for f in it if kC < len(f) and f[kC] is not None}
wb.close()
print('maestro de rutas: %d tiendas' % len(tiendas))

# ── el canal fino ───────────────────────────────────────────────────────
tipo_orden = {}


def tragar(ruta2):
    if not os.path.isfile(ruta2):
        return
    try:
        f3, r3 = abrir(ruta2)
    except OSError:
        return
    for x in r3:
        o = limpio(x.get('Número de orden'))
        if o and o not in tipo_orden:
            tipo_orden[o] = limpio(x.get('Tipo de orden'))
    f3.close()


tragar(os.path.join(CARPETA_ORD, 'Detalle Orden Pendientes.csv'))
tragar(os.path.join(CARPETA_ORD, 'Detalle Orden Despachados.csv'))
for n in sorted((n for n in os.listdir(CARPETA_ORD)
                 if re.match(r'^Detalle Orden \d{2}-\d{2}\.csv$', n)), reverse=True):
    tragar(os.path.join(CARPETA_ORD, n))
print('tipo de orden conocido para %s ordenes' % '{:,}'.format(len(tipo_orden)))


def canal_de(destino, orden):
    if destino in tiendas:
        return 'RETAIL'
    t = (tipo_orden.get(orden) or '').upper()
    if not t:
        return 'SIN CANAL'
    if 'MAYOR' in t:
        return 'MAYORISTA'
    if 'CATALOGO' in t:
        return 'CATALOGO'
    if 'ECOMMERCE' in t or 'VIRTUAL' in t:
        return 'ECOMMERCE'
    if 'INDUSTRIAL' in t:
        return 'INDUSTRIAL'
    return 'OTROS'


# ── el archivo de OBLPN ─────────────────────────────────────────────────
# POR INDICE, NO CON DictReader: "Usuario de seleccion" sale DOS VECES y el
# diccionario se queda con la ultima, perdiendo la primera.
def _abrir(ruta):
    f = io.open(ruta, encoding='utf-8-sig', newline='', errors='replace')
    cabeza = f.read(4000)
    f.seek(0)
    r = csv.reader(f, delimiter=';' if cabeza.count(';') > cabeza.count(',') else ',')
    cab = [c.strip() for c in next(r)]
    filas = [x for x in r if len(x) >= len(cab) - 2]
    f.close()
    return cab, filas


def juntar_todos(carpeta, dia):
    """TODAS las filas de ese dia, de TODOS los archivos, sin repetir.

    UN ARCHIVO DEL OBLPN NO ES UN DIA. Medido el 02-sep-2026 sobre los 28 que hay:
    `OBLPN 31-08.csv` trae 12.497 lineas del 31 pero tambien 3.536 del 27, 1.525
    del 28 y 1.458 del 26. Y al reves: las lineas de un dia quedan repartidas
    entre los archivos de los dias que siguen.

    Por eso tomar un archivo como si fuera un dia deja el dia corto. Comprobado
    contra lo que se habia publicado: al 27-08 le faltaba el 60%, al 20-08 el 92%.

    La huella de una linea es LPN + articulo + hora de empaquetado. La misma linea
    aparece en varios archivos y hay que contarla una sola vez.

    El picking NO tiene este problema -0 de 32 archivos mezclan dias-, por eso
    esto vive solo aca.
    """
    cab = None
    filas = []
    vistas = set()
    ddmmaaaa = '%s/%s/%s' % (dia[8:], dia[5:7], dia[:4])
    nombres = sorted(n for n in os.listdir(carpeta)
                     if re.match(r'^OBLPN \d{1,2}-\d{1,2}\.csv$', n, re.I))
    for n in nombres:
        c, fs = _abrir(os.path.join(carpeta, n))
        if cab is None:
            cab = c
            iH = c.index('Registro de hora de empaquetado') if 'Registro de hora de empaquetado' in c else -1
            iL = c.index('Número de LPN') if 'Número de LPN' in c else -1
            iS = c.index('Código de artículo') if 'Código de artículo' in c else -1
        if c != cab:
            print('[AVISO] %s tiene otras columnas; se saltea' % n)
            continue
        for x in fs:
            if iH < 0 or iH >= len(x):
                continue
            if not limpio(x[iH]).startswith(ddmmaaaa):
                continue
            huella = (limpio(x[iL]) if 0 <= iL < len(x) else '',
                      limpio(x[iS]) if 0 <= iS < len(x) else '',
                      limpio(x[iH]))
            if huella in vistas:
                continue
            vistas.add(huella)
            filas.append(x)
    print('[JUNTANDO] %s: %s lineas unicas de %d archivos'
          % (dia, '{:,}'.format(len(filas)), len(nombres)))
    return cab, filas


_dia_pedido = dia_pedido()
if '--juntando' in sys.argv and _dia_pedido:
    cabo, crudas = juntar_todos(os.path.dirname(ARCHIVO), _dia_pedido)
else:
    cabo, crudas = _abrir(ARCHIVO)

pos = {}
for i, c in enumerate(cabo):
    pos.setdefault(c, []).append(i)


def col(nombre, n=0):
    v = pos.get(nombre)
    return v[n] if v and n < len(v) else -1


def dt(x, i):
    return limpio(x[i]) if 0 <= i < len(x) else ''


iPk = col('Registro de hora de empaquetado')
iUp, iUs = col('Usuario de paquete'), col('Usuario de selección')
iMod = col('Usuario de modificación de asignación')
iQ, iSku = col('Cantidad empaquetada'), col('Código de artículo')
iDest, iOrd = col('Instalación de destino'), col('Número de orden')
iUbi, iEst = col('Ubicación de selección'), col('Estado de LPN')
iLpn = col('Número de LPN')
iTarea = col('Número de tarea')
# la pre-etiqueta, no la caja: PRE... y su variante WPRE...
ES_PRE = re.compile(r'^W?PRE', re.I)

cel = defaultdict(lambda: defaultdict(float))
# (canal, persona, hora, clase) -> {tarea: [segundos]}. Hace falta la
# TAREA para poder armar los tramos; una lista suelta no distingue el
# rato trabajado del rato parado.
sellos = defaultdict(lambda: defaultdict(list))

# EL TIEMPO PARA MEDIR PRODUCTIVIDAD. La misma regla que en picking, ver el
# comentario largo en `produccion_picking.py`: cada tarea aporta (ultimo pick -
# primer pick) y se SUMAN todas, sin puente y sin descontar solapes, y solo
# cuentan las lineas con `Numero de tarea` de verdad -no el apaño del LPN-.
sellos_tarea = defaultdict(lambda: defaultdict(list))
pares_tarea = defaultdict(float)
lineas_ph = defaultdict(lambda: defaultdict(float))
marcas = defaultdict(lambda: defaultdict(float))
colec = defaultdict(lambda: defaultdict(float))
zonas = defaultdict(lambda: defaultdict(float))
zonas_ubi = defaultdict(set)
totales = defaultdict(lambda: defaultdict(float))
personas = defaultdict(set)
estados = defaultdict(float)
tipos_vistos = defaultdict(lambda: defaultdict(float))
destinos = defaultdict(lambda: defaultdict(float))
dia = dia_pedido() or dia_mayoritario(ARCHIVO, 'Registro de hora de empaquetado')
leidas = sin_hora = rellenados = sin_usuario = distintos = 0
pre_fuera = 0
pre_pares = 0.0
emb_vs_pick = defaultdict(int)
und_wms = 0

for x in crudas:
    hs = dt(x, iPk)
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})\s+(\d{1,2}):(\d{2}):(\d{2})$', hs)
    if not m:
        sin_hora += 1
        continue
    if dia is None:
        dia = '%s-%s-%s' % (m.group(3), m.group(2), m.group(1))
    elif not hs.startswith('%s/%s/%s' % (dia[8:], dia[5:7], dia[:4])):
        sin_hora += 1
        continue
    # EL LPN QUE EMPIEZA CON PRE NO ESTA EMBALADO, diga lo que diga el estado.
    if ES_PRE.match(dt(x, iLpn)):
        pre_fuera += 1
        pre_pares += entero(dt(x, iQ)) * pares_de_la_caja(dt(x, iSku))
        continue
    leidas += 1
    h = int(m.group(4))
    seg = h * 3600 + int(m.group(5)) * 60 + int(m.group(6))

    # EL EMBALADOR ES EL QUE MODIFICA LA ASIGNACION. Viene lleno en el 100% de
    # las filas, y es la unica columna que reproduce el web report del WMS.
    usr = dt(x, iMod)
    if not usr:
        usr = '(sin usuario)'
        sin_usuario += 1
    # se guarda el otro para poder comparar los dos en pantalla
    otro = dt(x, iUp) or dt(x, iUs)
    if otro and otro != usr:
        distintos += 1
    if otro:
        emb_vs_pick[(usr, otro)] += 1

    sku = dt(x, iSku)
    ubi = dt(x, iUbi) or '?'
    cant = entero(dt(x, iQ))
    und_wms += cant
    pares = cant * pares_de_la_caja(sku)
    estados[dt(x, iEst) or '(vacio)'] += pares

    g, mar, cl = maestro.get(sku[:7], ('Sin dato', 'Sin dato', 'Sin dato'))
    clase = tipo_de(sku, g, sku[:7] in maestro)
    can = canal_de(dt(x, iDest), dt(x, iOrd))
    z = ubi.split('-')[0]
    tipos_vistos[can][tipo_orden.get(dt(x, iOrd)) or '(sin dato)'] += 1
    destinos[can][dt(x, iDest)] += 1

    # LA TAREA: la del WMS, y si falta el LPN —la caja que se esta cerrando—.
    tarea = dt(x, iTarea) or ('L:' + dt(x, iLpn))
    tarea_real = dt(x, iTarea)      # sin el apaño del LPN
    for k in (can, TODOS):
        personas[k].add(usr)
        cel[(k, usr, h, clase)]['pares'] += pares
        cel[(k, usr, h, clase)]['lineas'] += 1
        sellos[(k, usr, h, clase)][tarea].append(seg)
        sellos[(k, usr, None, clase)][tarea].append(seg)
        if tarea_real:
            sellos_tarea[(k, usr, h, clase)][tarea_real].append(seg)
            sellos_tarea[(k, usr, None, clase)][tarea_real].append(seg)
            pares_tarea[(k, usr, h, clase)] += pares
            pares_tarea[(k, usr, None, clase)] += pares
        sellos[(k, usr, h, 'total')][tarea].append(seg)
        sellos[(k, usr, None, 'total')][tarea].append(seg)
        lineas_ph[(k, h)][clase] += pares
        lineas_ph[(k, h)]['lineas'] += 1
        marcas[(k, mar)][clase] += pares
        marcas[(k, mar)]['lineas'] += 1
        colec[(k, cl)][clase] += pares
        colec[(k, cl)]['lineas'] += 1
        zonas[(k, z)][clase] += pares
        zonas[(k, z)]['lineas'] += 1
        zonas_ubi[(k, z)].add(ubi)
        totales[k][clase] += pares
        totales[k]['lineas'] += 1


def vol(d):
    o = {c: int(round(d.get(c, 0))) for c in CL}
    o['lineas'] = int(d.get('lineas', 0))
    # EL TOTAL SE SUMA SOBRE `CL`, no sobre tres nombres escritos a mano. Al
    # agregar `materiales` y `sin_tipo` la suma vieja los dejaba afuera y el
    # cuadro no cuadraba, sin una sola queja.
    o['total'] = sum(o[c] for c in CL)
    return o


def ritmo(mt, lineas, minimo, span_min):
    if not mt or len(mt) < 2:
        return None, None, None, False
    span = max(mt) - min(mt)
    mins = round(span / 60.0, 1)
    if lineas < minimo or span < span_min:
        return None, None, mins, False
    sl = span / float(lineas - 1)
    if sl < SEG_LINEA_MIN:
        return None, None, mins, False
    return (int(round(lineas / (span / 3600.0))), int(round(sl)), mins,
            span < SEG_MUESTRA_CORTA)


def minutos_sumados(por_tarea):
    """(ultimo pick - primer pick) de cada tarea, SUMADO. Ver produccion_picking."""
    if not por_tarea:
        return 0
    return sum(max(v) - min(v) for v in por_tarea.values() if len(v) > 1)


def tramos(por_tarea):
    """Los ratos en que esa persona estuvo trabajando, ya fusionados.

    UNIR, NO SUMAR. Las tareas se solapan -varios contenedores en un mismo
    recorrido-, asi que sumar sus duraciones cuenta el mismo minuto dos veces y
    da mas horas que las del dia. Fusionando, lo que queda entre tramo y tramo
    es tiempo parado: el refrigerio sale solo, sin descontarlo a mano.
    """
    if not por_tarea:
        return []
    ts = sorted((min(v), max(v)) for v in por_tarea.values() if v)
    if not ts:
        return []
    fus = [list(ts[0])]
    for a, b in ts[1:]:
        # se pisan, o estan lo bastante pegados como para ser la misma tanda
        if a <= fus[-1][1] + PUENTE_SEG:
            fus[-1][1] = max(fus[-1][1], b)
        else:
            fus.append([a, b])
    return fus


def celda(can, usr, h):
    """Una celda: el volumen de las tres clases y EL PRIMER Y ULTIMO PICK.

    NO SE PUBLICA EL RITMO YA CALCULADO. La pantalla deja elegir varios canales a
    la vez, y los ritmos no se suman: hay que rehacerlos sobre el conjunto. Con
    el minimo de los minimos y el maximo de los maximos, mas las lineas sumadas,
    el guion saca exactamente el mismo numero que sacaria aca.
    """
    o = {}
    tot_l = 0
    for c in CL + ('total',):
        if c == 'total':
            o['total'] = sum(o[c] for c in CL)
            o['lineas'] = tot_l
            o['total_l'] = tot_l
        else:
            if h is None:
                d = {'pares': sum(cel.get((can, usr, y, c), {}).get('pares', 0)
                                  for y in HORAS),
                     'lineas': sum(cel.get((can, usr, y, c), {}).get('lineas', 0)
                                   for y in HORAS)}
            else:
                d = cel.get((can, usr, h, c), {})
            n = int(d.get('lineas', 0))
            tot_l += n
            o[c] = int(round(d.get('pares', 0)))
            o[c + '_l'] = n
        o[c + '_iv'] = tramos(sellos.get((can, usr, h, c)))
        if c != 'total':
            o[c + '_s'] = minutos_sumados(sellos_tarea.get((can, usr, h, c)))
            o[c + '_q'] = int(round(pares_tarea.get((can, usr, h, c), 0)))
    return o


def vista(can):
    ph = {}
    for h in HORAS:
        v = vol(lineas_ph.get((can, h), {}))
        v['personas'] = sum(1 for u in personas[can]
                            if any((can, u, h, c) in cel for c in CL))
        ph[str(h)] = v
    return {
        'totales': vol(totales[can]),
        'por_hora': ph,
        'gente': sorted([{'usuario': u, 'total': celda(can, u, None),
                          # SOLO LAS HORAS EN QUE ESA PERSONA MOVIO ALGO.
                          # Con las 24 completas el archivo del dia pasaba de 368
                          # a 585 KB —un mes serian 35 MB y bajar un rango de
                          # treinta dias, 17 MB al navegador— y veinte de esas
                          # veinticuatro celdas venian en cero. La pantalla trata
                          # la hora que falta como vacia.
                          'horas': {str(h): c for h, c in
                                    ((h, celda(can, u, h)) for h in HORAS)
                                    if c.get('total')}}
                         for u in personas[can]],
                        key=lambda x: -x['total']['total']),
        'marcas': sorted([dict(nom=k[1], **vol(v)) for k, v in marcas.items()
                          if k[0] == can], key=lambda x: -x['total'])[:14],
        'coleccion': sorted([dict(nom=k[1], **vol(v)) for k, v in colec.items()
                             if k[0] == can], key=lambda x: -x['total'])[:14],
        'zonas': sorted([dict(nom=k[1], ubicaciones=len(zonas_ubi[k]),
                              ubis=sorted(zonas_ubi[k]), **vol(v))
                         for k, v in zonas.items() if k[0] == can],
                        key=lambda x: -x['total']),
    }


con_datos = [c for c in ORDEN_CANAL if totales.get(c, {}).get('lineas')]
salida = {
    'dia': dia,
    'archivo': os.path.basename(ARCHIVO),
    'lineas_buenas': leidas,
    'lineas_descartadas': sin_hora,
    'horas': HORAS,
    'cortes': {'lineasCelda': LINEAS_MIN_CELDA, 'lineasDia': LINEAS_MIN_DIA,
               'minutosCelda': SEG_MIN_CELDA // 60, 'minutosDia': SEG_MIN_DIA // 60,
               'segLineaMin': SEG_LINEA_MIN, 'puenteMin': PUENTE_SEG // 60, 'muestraCortaMin': SEG_MUESTRA_CORTA // 60},
    'canales': [TODOS] + con_datos,
    'gentePorCanal': {c: len(personas[c]) for c in [TODOS] + con_datos},
    'preFuera': {'lineas': pre_fuera, 'pares': int(round(pre_pares))},
    'usuario': {'columna': 'Usuario de modificación de asignación',
                'sinUsuario': sin_usuario, 'distintoDelPicker': distintos},
    'estados': sorted([[k, int(v)] for k, v in estados.items()], key=lambda x: -x[1]),
    'unidadesWms': und_wms,
    'tiposPorCanal': {c: sorted([[k, int(v)] for k, v in tipos_vistos[c].items()],
                                key=lambda x: -x[1])[:6] for c in con_datos},
    'destinosPorCanal': {c: sorted([[k, int(v)] for k, v in destinos[c].items()],
                                   key=lambda x: -x[1])[:6] for c in con_datos},
    'vistas': {c: vista(c) for c in [TODOS] + con_datos},
}

os.makedirs(os.path.dirname(SALIDA), exist_ok=True)
io.open(SALIDA, 'w', encoding='utf-8').write(json.dumps(salida, ensure_ascii=False))


def mm(n):
    return '{:,}'.format(int(n))


T = salida['vistas'][TODOS]['totales']
print('')
print('DIA %s  -  %s lineas EMBALADAS  -  %s sin hora del dia  -  %s con LPN "PRE" (%s pares)'
      % (dia, mm(leidas), mm(sin_hora), mm(pre_fuera), mm(pre_pares)))
print('TOTAL  suelto %s  -  prepack %s  -  no calzado %s  =  %s pares'
      % (mm(T['cal_suelto']), mm(T['cal_prepack']), mm(T['no_cal']), mm(T['total'])))
print('el WMS dice %s unidades; con la curva del prepack son %s pares'
      % (mm(und_wms), mm(T['total'])))
print('embalador = "Usuario de modificacion de asignacion"; %s sin dueno' % mm(sin_usuario))
print('en %s de %s lineas el embalador es DISTINTO del que pickeo (%.1f%%)'
      % (mm(distintos), mm(leidas), 100.0 * distintos / (leidas or 1)))
print('')
print('LOS 10 QUE MAS EMBALARON, y quien habia pickeado eso')
pe = defaultdict(float)
for (e, pk), v in emb_vs_pick.items():
    pe[e] += v
for e, v in sorted(pe.items(), key=lambda y: -y[1])[:10]:
    suyos = sorted([(pk, n2) for (e2, pk), n2 in emb_vs_pick.items() if e2 == e],
                   key=lambda y: -y[1])[:3]
    print('  %-16s %8s lineas   pickearon: %s'
          % (e[:16], mm(v), ', '.join('%s (%s)' % (a, mm(b)) for a, b in suyos)))
print('')
print('ESTADO DEL LPN (en pares)')
for k, v in salida['estados']:
    print('  %-22s %10s' % (k[:22], mm(v)))
print('')
print('POR CANAL')
print('  %-12s %9s %9s %10s %10s %9s %7s'
      % ('CANAL', 'LINEAS', 'PARES', 'SUELTO', 'PREPACK', 'NO CALZ', 'GENTE'))
suma_l = 0
for c in con_datos:
    v = salida['vistas'][c]['totales']
    suma_l += v['lineas']
    print('  %-12s %9s %9s %10s %10s %9s %7d'
          % (c, mm(v['lineas']), mm(v['total']), mm(v['cal_suelto']),
             mm(v['cal_prepack']), mm(v['no_cal']), len(personas[c])))
print('  %-12s %9s %9s %10s %10s %9s %7d'
      % ('TODOS', mm(T['lineas']), mm(T['total']), mm(T['cal_suelto']),
         mm(T['cal_prepack']), mm(T['no_cal']), len(personas[TODOS])))
print('  los canales suman %s lineas y el total dice %s  ->  %s'
      % (mm(suma_l), mm(T['lineas']), 'CUADRA' if suma_l == T['lineas'] else 'NO CUADRA'))
print('')
print('json en %s  (%.0f KB)' % (SALIDA, os.path.getsize(SALIDA) / 1024.0))


# ── SE PUBLICA LO QUE SE ACABA DE CALCULAR ──────────────────────────────
# Va al final y no antes: si el calculo se cae, no se pisa el cuadro bueno que
# quedo del pase anterior.
try:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from publicar_area import publicar

    def _log(t, nivel='INFO'):
        print('[%s] %s' % (nivel, t))

    # MODO HISTORICO: se calcula y se guarda aparte, SIN publicar.
    #
    # Sirve para rellenar agosto entero de una sola vez. No se publica en el
    # momento porque el servidor hoy guarda solo 2 dias por area: subir treinta
    # dias haria que se pisaran entre ellos y quedarian los dos ultimos. Se dejan
    # calculados y se suben todos juntos cuando el tope este arriba.
    if '--historico' in sys.argv:
        _dir = os.path.join(os.path.dirname(SALIDA), 'historico')
        os.makedirs(_dir, exist_ok=True)
        _f = os.path.join(_dir, '%s_%s.json' % (AREA, dia))
        io.open(_f, 'w', encoding='utf-8').write(json.dumps(salida, ensure_ascii=False))
        print('[HISTORICO] guardado %s (%.0f KB), sin publicar'
              % (os.path.basename(_f), os.path.getsize(_f) / 1024.0))
        raise SystemExit(0)

    # UN CUADRO VACIO NO SE PUBLICA NUNCA.
    #
    # El 02-sep-2026 una corrida leyo el archivo equivocado, saco cero lineas y
    # sin fecha, y piso en el servidor el cuadro bueno del dia anterior. Que el
    # calculo salga mal no puede borrar lo que ya estaba: si no hay dia o no hay
    # ni una linea, se avisa y no se manda nada.
    _T = salida['vistas'][TODOS]['totales']
    if not dia or not _T.get('lineas'):
        print('[AVISO] el cuadro salio vacio (dia=%s, lineas=%s): NO se publica, '
              'se deja el que ya estaba' % (dia, _T.get('lineas')))
        raise SystemExit(1)

    publicar(AREA, salida, dia, _log)
except Exception as _e:
    print('[ERROR] no se pudo publicar: %s: %s' % (type(_e).__name__, _e))
