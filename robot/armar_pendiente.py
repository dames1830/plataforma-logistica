# -*- coding: utf-8 -*-
"""EL PENDIENTE DE DESPACHO: lo que comercial mando y el CD todavia no atendio.

Lo pidio Daniel el 20-ago-2026, como fase 2 de `correo_guias.py`. Aquel baja el
correo; este lo convierte en la lista de trabajo.

LA REGLA, DICHA POR EL:

    "Voy a separar el SKU X, pero que venga de reserva y sea el mismo SKU X que
     me pide el analisis separar. Ese es tu filtro."
    "Ponte que del WMS saques veinte mil, pero de esos comercial solo mando diez
     mil."

    pendiente = lineas del "Detalle Orden Pendientes.csv" del WMS
                cuyo NUMERO DE ORDEN aparezca en algun correo de comercial
    lo que falta de cada linea = Cantidad solicitada - Cantidad asignada

NO SE ACUMULA NADA, SE RECALCULA ENTERO. Daniel lo hacia a mano juntando su
pendiente de ayer con el correo del dia, mandando la lista de ordenes al WMS
para preguntar estados, filtrando "Creada" y "Parcialmente asignado" y volviendo
a bajar el detalle. Ese ida y vuelta no hace falta:

  - El archivo de Pendientes del WMS YA viene con solo esos dos estados -lo
    filtra `picking_y_orden.py`, un año hacia atras, porque en los demas estados
    solicitada y asignada son iguales-. **Esa ES la validacion en el WMS.**
  - Lo que se cerro desaparecio solo del archivo: no hace falta el de ayer.
  - Lo picado ya bajo de la columna asignada: no hace falta descontarlo.

POR QUE EL SEGUNDO FILTRO NO ES UN ADORNO. Medido el 20-ago-2026: de 2.302
ordenes abiertas en el WMS, comercial mando 1.594 y **nunca libero 708, con
165.580 unidades** -mas de lo que si mando, y son cajas, papel tissue y bolsas-.
Sin el cruce, todo eso entraria al buffer como si fuera trabajo del CD.

QUE DEJA:

  1. El area `pendiente_despacho?date=AAAA-MM-DD` con los totales y los siete
     cortes que dibuja Zona Buffer -> Pendiente. Son ~300 filas de resumen, no el
     detalle: el detalle pesa 13 MB y para eso esta el Excel.
  2. La tarjeta **PEDIDOS** de Zona Buffer -> Archivo, con los SKU que tienen
     pendiente y sus cantidades solicitada y asignada. Es la misma area que antes
     se llenaba a mano, asi que Daniel puede seguir quitandola o cambiandola: la
     quita cuando quiere correr solo Replenishment u Otras solicitudes.
  3. `Pendiente DD-MM-AA.xlsx` en el modulo Descargas, con dos hojas:
       Detalle  - la CARA del correo de comercial, con las mismas columnas, pero
                  la cantidad es lo que falta por atender. Decision de Daniel:
                  *"debe ser tal cual el archivo que manda comercial, solo que
                  las cantidades deberian variar"*.
       Resumen  - codigo, solicitada, asignada, pendiente. Lo que come el buffer.

CUANDO CORRE. Lo dispara `correo_guias.py` en cuanto guarda un Excel nuevo, y no
una hora fija: **la foto del WMS tiene que ser posterior al correo**. El 20-ago se
midio que el robot baja el pendiente a las 06:57 y el correo llega a las 19:00;
con la foto de la mañana faltaban 277 ordenes -las nacidas durante el dia, entre
ellas las del correo de esa misma tarde- y sobraban 492 ya cerradas.

ANTES DE CRUZAR SE BAJA UNA FOTO NUEVA DEL WMS. El 21-ago-2026 esto no estaba y el
pendiente salio con la foto de las 06:57 contra un correo de las 18:32: publico
**31.246 unidades cuando lo real eran 116.467**. Las ordenes coincidian -1.608
contra 1.583, porque las define el correo-; lo que faltaba eran las lineas nacidas
durante el dia, el 87% del pendiente. Ahora corre `picking_y_orden.py
--solo-pendientes` y **no publica nada si la foto no queda posterior al correo**:
vale mas el pendiente de ayer que uno corto encima del bueno.

    python armar_pendiente.py              baja la foto, arma y publica el de hoy
    python armar_pendiente.py --probar     calcula y muestra, sin bajar ni publicar
    python armar_pendiente.py --sin-bajar  usa la foto que ya esta en disco
    python armar_pendiente.py --fecha 2026-08-20
"""

import collections
import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import traceback
import urllib.parse
import urllib.request
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    openpyxl = None

WEB_DATOS_API = "https://logistics-backend-wv0x.onrender.com/api/logistics"
# EL TOKEN DEL ROBOT. Desde v29.0415 el servidor puede EXIGIR credencial para
# escribir datos (ver EXIGIR_TOKEN_ESCRITURA en backend/main.py). El robot no tiene
# sesion, asi que lleva su propio token, leido del entorno del Contabo -NUNCA escrito
# aca, o estaria publico en el repo-. Si la variable no esta, se manda vacio y el
# servidor, mientras el candado siga apagado, lo deja pasar igual.
ROBOT_TOKEN = os.environ.get('ROBOT_TOKEN', '')

WEB_ARCHIVOS_API = "https://logistics-backend-wv0x.onrender.com/api/archivos"
AREA = "pendiente_despacho"
# La tarjeta PEDIDOS de Zona Buffer -> Archivo. Es un area COMPARTIDA que ya
# existia: hasta hoy la llenaba Daniel subiendo un CSV a mano y todas las PC lo
# leian de ahi. El robot escribe en el mismo lugar y con el mismo formato, asi que
# los botones de cambiar archivo y quitarlo siguen funcionando igual.
AREA_PEDIDOS = "buffer"

csv.field_size_limit(10 ** 7)


def _base_onedrive():
    """La carpeta de OneDrive. SE BUSCA, NO SE ESCRIBE A MANO.

    En la laptop el usuario de Windows es 'dames' y en el servidor
    'Administrator'. Una ruta fija sirve en una maquina y revienta en la otra.
    Misma funcion que `correo_guias.py` y `generar_slotting.py`.
    """
    for c in (os.environ.get('OneDrive'), os.environ.get('OneDriveCommercial'),
              os.path.join(os.path.expanduser('~'), 'OneDrive'),
              r'C:\Users\Administrator\OneDrive', r'C:\Users\dames\OneDrive'):
        if not c:
            continue
        ruta = os.path.join(c, 'danielames.bata', 'scraping Stock')
        if os.path.isdir(ruta):
            return ruta
    return os.path.join(os.path.expanduser('~'), 'OneDrive', 'danielames.bata',
                        'scraping Stock')


BASE = _base_onedrive()
CORREOS = os.path.join(BASE, 'Correos Picking')
PENDIENTES = os.path.join(BASE, 'Detalle Orden', 'Detalle Orden Pendientes.csv')
MAESTRO_CANDIDATOS = [
    os.path.join(os.path.dirname(BASE), 'Maestro_Articulos.xlsx'),
    os.path.join(BASE, 'Archivos', 'Maestro_Articulos.xlsx'),
    os.path.join(os.path.dirname(BASE), 'Pruebas Sistema', 'Maestro_Articulos.xlsx'),
]
# EL MAESTRO DE RUTAS. Dice de cada tienda si es LIMA o PROVINCIA, que dia sale y
# con que transportista. OJO: el nombre lleva DOS ESPACIOS entre RUTAS y TURNOS.
RUTAS_CANDIDATOS = [
    os.path.join(os.path.dirname(BASE), 'Proyecto web Logistico',
                 'RUTAS -  TURNOS.xlsx'),
    os.path.join(os.path.dirname(BASE), 'RUTAS -  TURNOS.xlsx'),
]
AQUI = os.path.dirname(os.path.abspath(__file__))
LOG = os.path.join(AQUI, 'logs', 'armar_pendiente.log')

SELLO = os.path.join(AQUI, 'logs', 'pendiente_armado.txt')

ESTADOS = ('Creada', 'Parcialmente asignado')
MINIMO_CRUCE = 0.30      # si cruza menos que esto, algo se rompio: no se publica
# Cuanto se le da a la bajada del WMS. Son 365 dias -unas 65.000 lineas- y tarda
# unos 8 minutos, pero puede pasarse 20 esperando al robot del stock y reintentar.
ESPERA_BAJADA = 45 * 60


def log(t, nivel='INFO'):
    linea = '[%s] [%-5s] %s' % (datetime.now().strftime('%H:%M:%S'), nivel, t)
    print(linea)
    try:
        os.makedirs(os.path.dirname(LOG), exist_ok=True)
        with io.open(LOG, 'a', encoding='utf-8') as fh:
            fh.write(datetime.now().strftime('%Y-%m-%d ') + linea + '\n')
    except Exception:
        pass


def arg(nombre, por_defecto=None):
    for i, a in enumerate(sys.argv):
        if a == nombre and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return por_defecto


def correos_excluidos():
    """Las fechas que se pidieron dejar fuera, como {(mes, dia)}.

    Se acepta `--sin-correo 24.08`, `24-08` y repetido varias veces. Se lee con el mismo
    `fecha_del_nombre` que nombra los archivos, para que no haya dos formas de entender
    una fecha: si el nombre del archivo se lee de una manera, el filtro tambien.
    """
    fuera = set()
    for i, a in enumerate(sys.argv):
        if a == '--sin-correo' and i + 1 < len(sys.argv):
            f = fecha_del_nombre(sys.argv[i + 1])
            if f:
                fuera.add(f)
            else:
                raise SystemExit('No entiendo la fecha "%s". Va como 24.08 o 24-08.'
                                 % sys.argv[i + 1])
    return fuera


def limpio(v):
    """El WMS exporta envuelto como formula: ="7997215". El correo lo escribe pelado.

    Sin quitar la envoltura el cruce da 0% y parece que no calzan. Costo una vuelta
    entera el 19-ago-2026.
    """
    return re.sub(r'^="?|"?$', '', str(v if v is not None else '').strip()).strip()


def num(v):
    try:
        return float(limpio(v).replace(',', ''))
    except Exception:
        return 0.0


# ══════════════════════════════════════════════════════════════════════════════
#  1. LOS CORREOS DE COMERCIAL
# ══════════════════════════════════════════════════════════════════════════════

def fecha_del_nombre(nombre):
    """`Guías 15.07.xlsx` y `Guías 15-06.xlsx` son el mismo formato con distinto separador.

    Julio y agosto llegaron con punto y junio con guion; leyendo solo el punto,
    junio entero entraba con fecha invalida. El `(?!\\d)` evita comerse un tercer
    numero.
    """
    m = re.search(r'(\d{2})[.\-](\d{2})(?!\d)', nombre)
    if not m:
        return None
    dia, mes = int(m.group(1)), int(m.group(2))
    if not (1 <= dia <= 31 and 1 <= mes <= 12):
        return None
    return (mes, dia)


def leer_correos():
    """Todas las guias que comercial mando alguna vez -> {guia: (fila, mes, dia)}.

    UNA MISMA GUIA PUEDE VENIR EN DOS CORREOS y manda la PRIMERA vez: asi conserva
    la prioridad y la fecha de cuando de verdad la mandaron.

    LA HOJA BUENA NO ES SIEMPRE LA PRIMERA. `Guías 07.07.xlsx` trae los datos en la
    segunda y un dia entero se perdio sin avisar -15.276 guias en vez de 15.623-.
    Se busca la hoja cuya cabecera traiga la columna GUIA; la mas grande no sirve
    como criterio, porque las del 11/12/13-ago traen una segunda hoja "Tiendas".
    """
    if openpyxl is None:
        raise SystemExit('Falta openpyxl. Instalalo con:  pip install openpyxl')
    if not os.path.isdir(CORREOS):
        raise SystemExit('No existe la carpeta de correos: %s' % CORREOS)

    archivos = []
    for n in os.listdir(CORREOS):
        if not n.lower().endswith(('.xlsx', '.xls')):
            continue
        f = fecha_del_nombre(n)
        if f:
            archivos.append((f, n))
    archivos.sort()

    # LOS QUE SE PIDIERON DEJAR FUERA. Se descuentan del total ANTES de contar, para que
    # el aviso de "se reconocieron X de Y" siga cazando un formato de nombre nuevo.
    fuera = correos_excluidos()
    if fuera:
        antes = len(archivos)
        archivos = [(f, n) for (f, n) in archivos if f not in fuera]
        log('SE DEJAN FUERA %d correo(s): %s'
            % (antes - len(archivos),
               ', '.join('%02d.%02d' % (d, m) for (m, d) in sorted(fuera))), 'AVISO')

    guias, cabecera, iq_out, ig_out = {}, None, None, None
    leidos = 0
    for (mes, dia), nombre in archivos:
        ruta = os.path.join(CORREOS, nombre)
        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        except Exception as e:
            log('No se pudo abrir %s (%s)' % (nombre, type(e).__name__), 'AVISO')
            continue
        hallado = False
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            try:
                cab = [str(c).strip() if c is not None else '' for c in next(it)]
            except StopIteration:
                continue
            if 'GUIA' not in cab:
                continue
            ig = cab.index('GUIA')
            iq = next((i for i, c in enumerate(cab) if 'CANTI' in c.upper()), None)
            if iq is None:
                continue
            if cabecera is None:
                cabecera, iq_out, ig_out = list(cab), iq, ig
                cabecera[iq] = 'CANTIDAD PENDIENTE'
            for r in it:
                g = limpio(r[ig])
                if g and g not in guias:
                    guias[g] = (list(r), mes, dia)
            hallado = True
            break
        if hallado:
            leidos += 1
        try:
            wb.close()
        except Exception:
            pass

    # ANTES DE DAR POR BUENA UNA CORRIDA, que la cuenta de dias reconocidos sea
    # igual a la de archivos de la carpeta. Es como se caza un formato de nombre
    # nuevo que este entrando con fecha invalida.
    if leidos < len(archivos):
        log('Se reconocieron %d de %d archivos de correo. Revisar los nombres.'
            % (leidos, len(archivos)), 'AVISO')
    log('Correos leidos: %d archivos, %s guias' % (leidos, format(len(guias), ',d')))
    return guias, cabecera, iq_out, ig_out


# ══════════════════════════════════════════════════════════════════════════════
#  2. EL MAESTRO — SIEMPRE POR NOMBRE DE COLUMNA
# ══════════════════════════════════════════════════════════════════════════════

def leer_maestro():
    """Gender RIMS, G. Gender y Coleccion PO de cada articulo.

    SE LEE POR NOMBRE DE COLUMNA, NUNCA POR POSICION. La tabla se corrio un lugar
    -aparecio un `CodCanal` adelante- y leyendo por indice fijo el cruce daba **1
    codigo de 29.465 y no avisaba nada**.

    Y OJO CON LA TEMPORADA: hay dos campos y se confunden. Lo que Daniel llama
    "la coleccion" es **Coleccion PO** (el `2026-Q4`). La columna que *se llama*
    Temporada es la franja del mezzanine -actual o anterior- y NO es esto.
    """
    ruta = next((r for r in MAESTRO_CANDIDATOS if os.path.isfile(r)), None)
    if not ruta:
        log('No se encontro el Maestro de articulos. Los cortes por Gender RIMS '
            'y por coleccion van a salir vacios.', 'AVISO')
        return {}, {}, {}
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else '' for c in next(it)]

    def col(*nombres):
        for i, c in enumerate(cab):
            if c.lower().replace(' ', '').replace('.', '') in nombres:
                return i
        return None

    iC = col('codarticulo', 'codigoarticulo')
    iG = col('ggender', 'gender')
    iR = col('genderrims')
    iK = col('coleccionpo')
    if iC is None:
        log('El Maestro no trae columna CodArticulo. Se ignora.', 'AVISO')
        return {}, {}, {}

    gen, rims, colec = {}, {}, {}
    for r in it:
        c = limpio(r[iC]) if iC < len(r) else ''
        if not c:
            continue
        if iG is not None and iG < len(r):
            gen[c] = str(r[iG] or '').strip()
        if iR is not None and iR < len(r):
            rims[c] = str(r[iR] or '').strip()
        if iK is not None and iK < len(r):
            colec[c] = str(r[iK] or '').strip()
    try:
        wb.close()
    except Exception:
        pass
    log('Maestro: %s articulos (%s)' % (format(len(gen), ',d'), os.path.basename(ruta)))
    return gen, rims, colec


def leer_rutas():
    """Cada tienda -> zona, transportista, turno y dia de despacho.

    SE COPIA ANTES DE ABRIRLO. El archivo esta *solo en la nube* en OneDrive y
    openpyxl lo ve como un zip roto -"File is not a zip file"-. Copiarlo lo baja,
    asi que siempre se lee de la copia.

    EL CODIGO DE TIENDA DEL CORREO LLEVA 50 DELANTE. El correo dice 238 y el
    maestro 50238. Regla de Daniel, 01-sep-2026.
    """
    ruta = next((r for r in RUTAS_CANDIDATOS if os.path.isfile(r)), None)
    if not ruta:
        log('No se encontro el maestro de rutas. El corte por ruta va a salir '
            'vacio.', 'AVISO')
        return {}
    copia = os.path.join(AQUI, 'logs', '_rutas_copia.xlsx')
    try:
        os.makedirs(os.path.dirname(copia), exist_ok=True)
        shutil.copyfile(ruta, copia)
    except Exception as e:
        log('No se pudo copiar el maestro de rutas (%s). Se intenta el original.'
            % type(e).__name__, 'AVISO')
        copia = ruta
    try:
        wb = openpyxl.load_workbook(copia, read_only=True, data_only=True)
    except Exception as e:
        log('No se pudo abrir el maestro de rutas (%s).' % type(e).__name__, 'AVISO')
        return {}
    it = wb.worksheets[0].iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else '' for c in next(it)]

    def col(nombre):
        for i, c in enumerate(cab):
            if c.strip().upper() == nombre:
                return i
        return None

    iC, iZ = col('CDG'), col('ZONA')
    iP, iT, iD = col('PROVEEDOR'), col('TURNO'), col('DIA')
    if iC is None or iZ is None:
        log('El maestro de rutas no trae CDG o ZONA. Se ignora.', 'AVISO')
        return {}

    def txt(r, i):
        return (str(r[i]).strip().upper()
                if i is not None and i < len(r) and r[i] is not None else '')

    rutas = {}
    for r in it:
        if iC >= len(r) or r[iC] is None:
            continue
        rutas[str(r[iC]).strip()] = (txt(r, iZ), txt(r, iP), txt(r, iT), txt(r, iD))
    wb.close()
    log('Maestro de rutas: %d tiendas' % len(rutas))
    return rutas


# ══════════════════════════════════════════════════════════════════════════════
#  3. EL CRUCE
# ══════════════════════════════════════════════════════════════════════════════

# --------- LA FOTO DEL WMS TIENE QUE SER POSTERIOR AL CORREO ---------

def _reloj(t):
    return datetime.fromtimestamp(t).strftime('%d-%m %H:%M') if t else '(no hay)'


def hora_foto():
    """Cuando se bajo el 'Detalle Orden Pendientes.csv' que hay en disco."""
    try:
        return os.path.getmtime(PENDIENTES)
    except OSError:
        return 0.0


def hora_correo():
    """Cuando llego el ultimo correo de comercial. Es la vara contra la que se mide
    la foto: el pendiente se arma cruzando los dos, y cruzar un correo de las 19:00
    contra una foto de las 06:57 deja fuera todo lo que nacio durante el dia."""
    ultimo = 0.0
    try:
        for n in os.listdir(CORREOS):
            if n.startswith('~$') or not n.lower().endswith(('.xlsx', '.xls')):
                continue
            m = os.path.getmtime(os.path.join(CORREOS, n))
            if m > ultimo:
                ultimo = m
    except Exception:
        pass
    return ultimo


def refrescar_pendientes():
    """BAJA DEL WMS UNA FOTO DE HOY ANTES DE CRUZAR. Devuelve True solo si la foto
    que queda en disco es POSTERIOR al correo mas nuevo.

    POR QUE EXISTE. Lo eligio Daniel el 20-ago-2026 -'cuando guarda el Excel, sigue:
    baja el Detalle de Orden, cruza y publica'- y el 21 volvio a fallar por no
    estar: el correo se guardo 18:32 y la foto seguia siendo la de las 06:57, con
    CERO pendientes del dia. Publico 31.246 unidades contra 116.467 reales.

    LA VARA ES EL CORREO MAS NUEVO, no el de hoy a secas. Si comercial manda una
    correccion a las 21:40 y la foto es de las 19:10, la foto vuelve a quedar vieja
    y se baja de nuevo. Solo se saltea la bajada cuando de verdad no cambio nada, y
    ahi se ahorran los ocho minutos.

    SI FALLA, NO SE PUBLICA. La bajada le cede el paso al robot del stock -codigo
    3- y ademas puede fallar por mil motivos; en todos la respuesta es la misma que
    puso Daniel para el cruce: no se pisa el pendiente bueno con uno peor. Queda el
    del dia anterior y el correo reintenta en la vuelta siguiente.
    """
    vara = hora_correo()
    log('   correo mas nuevo   %s' % _reloj(vara))
    log('   foto del WMS       %s' % _reloj(hora_foto()))
    if vara and hora_foto() > vara:
        log('   la foto ya es posterior al correo: no hace falta bajarla de nuevo')
        return True

    bajador = os.path.join(AQUI, 'picking_y_orden.py')
    if not os.path.isfile(bajador):
        log('No esta picking_y_orden.py, no hay con que bajar la foto: %s'
            % bajador, 'ERROR')
        return False

    log('Bajando del WMS la foto de hoy (365 dias, unos 8 minutos)...')
    cod = -1
    try:
        cod = subprocess.run([sys.executable, bajador, '--solo-pendientes'],
                             timeout=ESPERA_BAJADA).returncode
    except Exception as e:
        log('No se pudo correr picking_y_orden.py (%s: %s)'
            % (type(e).__name__, str(e)[:140]), 'ERROR')
    if cod == 0:
        log('Foto nueva bajada - %s' % _reloj(hora_foto()))
    elif cod == 3:
        log('El WMS estaba ocupado con otro robot y esta bajada le cede el paso.',
            'AVISO')
    else:
        log('La bajada FALLO (codigo %s). Mirar logs/picking_orden_*.log' % cod,
            'ERROR')

    if not vara:
        # Sin correos no hay con que comparar, y alcanza con que el archivo exista:
        # el cruce de mas abajo tampoco va a dar nada y se corta ahi.
        return hora_foto() > 0
    if hora_foto() > vara:
        return True
    log('La foto del WMS sigue siendo ANTERIOR al correo (%s contra %s).'
        % (_reloj(hora_foto()), _reloj(vara)), 'ERROR')
    return False


def armar(hoy):
    guias, cabecera, IQ, IG = leer_correos()
    gen, rims, colec = leer_maestro()
    rutas = leer_rutas()

    if not os.path.isfile(PENDIENTES):
        raise SystemExit('No esta el pendiente del WMS: %s' % PENDIENTES)

    hoy_d = datetime.strptime(hoy, '%Y-%m-%d').date()

    f = io.open(PENDIENTES, encoding='utf-8-sig', newline='', errors='replace')
    r = csv.reader(f, delimiter=';')
    try:
        next(r)
    except StopIteration:
        raise SystemExit('El pendiente del WMS esta vacio.')

    # UNA SOLA LINEA POR (orden, articulo, destino). El archivo trae repetidas y
    # sumarlas agrego 5.714 unidades de la nada el 19-ago-2026.
    vistas = set()
    por_guia = collections.defaultdict(float)
    por_sku = collections.defaultdict(lambda: [0.0, 0.0])
    tiendas = collections.defaultdict(lambda: [set(), 0.0])
    r_rims = collections.defaultdict(lambda: [set(), 0.0])
    r_col = collections.defaultdict(lambda: [set(), 0.0])
    r_pri = collections.defaultdict(lambda: [set(), 0.0])
    r_gen = collections.defaultdict(float)
    # (zona, fila) -> [accesorios, calzado]. La fila es DIA + TURNO en Lima y el
    # transportista en provincia: asi lo arma comercial en su dinamica.
    r_rut = collections.defaultdict(lambda: [0.0, 0.0])
    rut_sin = [0.0, set()]
    r_ant = collections.defaultdict(lambda: [set(), 0.0])
    ord_dentro, ord_fuera = set(), set()
    und_dentro = und_fuera = 0.0
    lineas = repetidas = 0
    sin_maestro = set()

    def tramo(mes, dia):
        try:
            d = datetime(hoy_d.year, mes, dia).date()
        except Exception:
            return 'sin fecha'
        x = (hoy_d - d).days
        if x < 0:
            x = 0
        return ('hoy' if x == 0 else '1 dia' if x == 1 else '2 a 3 dias' if x <= 3
                else '4 a 7 dias' if x <= 7 else '8 a 15 dias' if x <= 15
                else 'mas de 15 dias')

    for row in r:
        if len(row) < 14:
            continue
        if row[4].strip() not in ESTADOS:
            continue
        orden = limpio(row[1])
        sku = limpio(row[5])
        dest = limpio(row[13])
        clave = (orden, sku, dest)
        if clave in vistas:
            repetidas += 1
            continue
        vistas.add(clave)
        pend = num(row[6]) - num(row[9])

        if orden not in guias:
            ord_fuera.add(orden)
            und_fuera += max(0.0, pend)
            continue
        ord_dentro.add(orden)
        und_dentro += max(0.0, pend)
        por_sku[sku][0] += num(row[6])
        por_sku[sku][1] += num(row[9])
        if pend <= 0:
            continue
        lineas += 1
        por_guia[orden] += pend

        fila, mes, dia = guias[orden]
        def campo(i):
            return str(fila[i]).strip() if i < len(fila) and fila[i] is not None else ''
        tienda = ('%s %s' % (campo(1), campo(2))).strip() or '(sin tienda)'
        prioridad = campo(3) or '(sin prioridad)'
        base = sku.split('-')[0]
        rr = rims.get(sku) or rims.get(base) or '(sin Maestro)'
        cc = colec.get(sku) or colec.get(base) or '(sin coleccion)'
        gg = gen.get(sku) or gen.get(base) or '(sin Maestro)'
        if rr == '(sin Maestro)':
            sin_maestro.add(sku)

        tiendas[tienda][0].add(orden); tiendas[tienda][1] += pend
        r_rims[rr][0].add(orden); r_rims[rr][1] += pend
        r_col[cc][0].add(orden); r_col[cc][1] += pend
        r_pri[prioridad][0].add(orden); r_pri[prioridad][1] += pend
        r_gen[gg] += pend

        # EL CORTE POR RUTA. Al codigo de tienda del correo se le pone 50 delante
        # para encontrarlo en el maestro de rutas.
        cod = campo(1)
        info = rutas.get('50' + cod.lstrip('0').zfill(3)) if cod else None
        if info:
            zona, prov, turno, dsp = info
            if zona.startswith('LIMA'):
                clave = ('LIMA', ('%s %s' % (dsp, turno)).strip() or '(sin ruta)')
            else:
                clave = ('PROVINCIA', prov or '(sin transportista)')
            r_rut[clave][0 if gg != 'Footwear' else 1] += pend
        elif rutas:
            rut_sin[0] += pend
            rut_sin[1].add(cod)

        t = tramo(mes, dia)
        r_ant[t][0].add(orden); r_ant[t][1] += pend

    f.close()

    # LA GUARDA. Un cruce roto da casi cero y no avisa: el WMS envuelve los codigos
    # como formula y el correo los escribe pelados, asi que basta un cambio de
    # formato para que no calce ninguno. Antes que publicar un pendiente vacio,
    # no se publica nada.
    total_ord = len(ord_dentro) + len(ord_fuera)
    cruce = (len(ord_dentro) / float(total_ord)) if total_ord else 0.0
    if total_ord and cruce < MINIMO_CRUCE:
        raise SystemExit(
            'EL CRUCE NO CUADRA: solo %d de %d ordenes del WMS figuran en algun correo '
            '(%.0f%%). No se publica nada; queda el pendiente anterior.'
            % (len(ord_dentro), total_ord, 100 * cruce))
    if sin_maestro:
        log('%d articulos no estan en el Maestro: sus cortes van a "(sin Maestro)"'
            % len(sin_maestro), 'AVISO')

    ORDEN_ANT = ['hoy', '1 dia', '2 a 3 dias', '4 a 7 dias', '8 a 15 dias',
                 'mas de 15 dias', 'sin fecha']

    def lista(d, limite=None):
        filas = [{'k': k, 'ped': len(v[0]), 'und': int(round(v[1]))}
                 for k, v in d.items() if v[1] > 0]
        filas.sort(key=lambda x: -x['und'])
        return filas[:limite] if limite else filas

    dias_vieja = 0
    for t, v in r_ant.items():
        if t == 'mas de 15 dias' and v[1] > 0:
            dias_vieja = 16
    if not dias_vieja:
        for i, t in enumerate(ORDEN_ANT):
            if t in r_ant and r_ant[t][1] > 0:
                dias_vieja = [0, 1, 3, 7, 15, 99, 0][i]

    datos = {
        'fecha': hoy,
        'generado': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'totales': {
            'pedidos': len(por_guia),
            'tiendas': len(tiendas),
            'articulos': len([s for s, v in por_sku.items() if v[0] - v[1] > 0]),
            'unidades': int(round(sum(por_guia.values()))),
            'lineas': lineas,
            'diasMasVieja': dias_vieja,
        },
        'origen': {
            'abiertoWms': {'ordenes': total_ord,
                           'unidades': int(round(und_dentro + und_fuera))},
            'mandado': {'ordenes': len(ord_dentro), 'unidades': int(round(und_dentro))},
            'noLiberado': {'ordenes': len(ord_fuera), 'unidades': int(round(und_fuera))},
        },
        'antiguedad': [{'k': t, 'ped': len(r_ant[t][0]), 'und': int(round(r_ant[t][1]))}
                       for t in ORDEN_ANT if t in r_ant and r_ant[t][1] > 0],
        'tiendas': lista(tiendas),
        'rims': lista(r_rims),
        'coleccion': lista(r_col),
        'prioridad': lista(r_pri),
        'gender': [{'k': k, 'und': int(round(v))}
                   for k, v in sorted(r_gen.items(), key=lambda x: -x[1]) if v > 0],
        'repetidasDescartadas': repetidas,
        # EL CUADRO DE COMERCIAL. Mismo corte que la dinamica del asistente:
        # zona -> ruta, partido en calzado y no calzado.
        'rutas': [{'z': k[0], 'k': k[1], 'acc': int(round(v[0])),
                   'cal': int(round(v[1])), 'und': int(round(v[0] + v[1]))}
                  for k, v in sorted(r_rut.items(), key=lambda x: -(x[1][0] + x[1][1]))],
        'rutasSinCruce': {'und': int(round(rut_sin[0])),
                          'tiendas': len(rut_sin[1])},
    }
    return datos, guias, cabecera, IQ, por_guia, por_sku


# ══════════════════════════════════════════════════════════════════════════════
#  4. EL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def excel(ruta, cabecera, IQ, guias, por_guia, por_sku):
    """Dos hojas.

    `Detalle` tiene LA CARA DEL CORREO DE COMERCIAL: las mismas columnas, en el
    mismo orden. Lo unico que cambia es la cantidad, que pasa a ser lo que falta
    por atender. Decision de Daniel, 20-ago-2026: *"debe ser tal cual el archivo
    que manda comercial, solo que las cantidades deberian variar"*. Las guias ya
    atendidas del todo no salen: no aportan nada.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detalle'
    relleno = PatternFill('solid', fgColor='1F3864')
    negrita = Font(bold=True, color='FFFFFF')

    ws.append(cabecera)
    for c in ws[1]:
        c.fill = relleno
        c.font = negrita
        c.alignment = Alignment(horizontal='center')
    for g, q in sorted(por_guia.items(), key=lambda x: -x[1]):
        if g not in guias or q <= 0:
            continue
        fila = list(guias[g][0])
        while len(fila) < len(cabecera):
            fila.append('')
        fila[IQ] = int(round(q))
        ws.append(fila[:len(cabecera)])
    ws.freeze_panes = 'A2'
    for i, ancho in enumerate([9, 8, 27, 17, 13, 12, 11, 9, 13, 21, 11], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws2 = wb.create_sheet('Resumen')
    ws2.append(['Código de artículo', 'Cantidad solicitada', 'Cantidad asignada', 'Pendiente'])
    for c in ws2[1]:
        c.fill = relleno
        c.font = negrita
    for s in sorted(por_sku):
        sol, asig = por_sku[s]
        if sol - asig <= 0:
            continue
        ws2.append([s, int(round(sol)), int(round(asig)), int(round(sol - asig))])
    ws2.freeze_panes = 'A2'
    for i, ancho in enumerate([22, 20, 20, 14], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    wb.save(ruta)
    return ruta


# ══════════════════════════════════════════════════════════════════════════════
#  5. PUBLICAR
# ══════════════════════════════════════════════════════════════════════════════

def publicar_datos(datos, intentos=3):
    cuerpo = json.dumps(datos, ensure_ascii=False).encode('utf-8')
    url = '%s/%s?date=%s' % (WEB_DATOS_API, AREA, datos['fecha'])
    for i in range(1, intentos + 1):
        try:
            p = urllib.request.Request(url, data=cuerpo, method='POST')
            p.add_header('Content-Type', 'application/json')
            p.add_header('X-Robot-Token', ROBOT_TOKEN)
            with urllib.request.urlopen(p, timeout=300) as resp:
                json.loads(resp.read().decode('utf-8'))
            log('Publicado en la plataforma: %.1f KB' % (len(cuerpo) / 1024.0))
            return True
        except Exception as e:
            if i < intentos:
                log('Intento %d: no se pudo publicar (%s), se reintenta'
                    % (i, type(e).__name__), 'AVISO')
            else:
                log('No se pudo publicar: %s: %s' % (type(e).__name__, str(e)[:160]), 'ERROR')
    return False


def publicar_pedidos(por_sku, intentos=3):
    """DEJA EL PENDIENTE EN LA TARJETA *PEDIDOS* DE ZONA BUFFER -> ARCHIVO.

    Lo pidio Daniel el 21-ago-2026: *"una vez que el robot termine de hacer el
    pendiente, lo tiene que publicar en la zona de buffer, en archivos de buffer"*.
    Hasta hoy esa tarjeta se llenaba subiendo un CSV a mano.

    EL FORMATO NO ES CAPRICHO. La web publica esa area REDUCIDA a tres columnas
    -ver DEMANDA_EN_LA_NUBE en `csvHub_v6.js`- porque el archivo entero son 30
    columnas y 50.333 filas, 58 MB, y el motor del buffer lee solo tres. Ademas hay
    una guarda que RECHAZA lo que no venga reducido, asi que mandar el archivo
    crudo seria como no mandar nada. Los nombres son los canonicos, los mismos que
    busca el motor.

    SOLO LOS QUE TIENEN PENDIENTE. Se publican los 2.844 SKU con solicitada mayor
    que asignada, que son los mismos que cuenta el reporte y los mismos que van a
    la hoja Resumen del Excel: 116.474 unidades el 21-ago. Si se mandaran todos los
    SKU tocados -5.697- el motor daria igual, pero la tarjeta diria un numero que no
    coincide con ningun otro lado.

    NO LLEVA ?date=. Se publica igual que lo hace el navegador, sin fecha, y el
    servidor lo guarda bajo el dia de HOY. Poner la fecha a mano abriria la puerta a
    que el robot y la web escribieran en dos renglones distintos.

    SI DANIEL LO QUITA, ESTO NO SE LO DEVUELVE EN EL ACTO: el sello dice que el
    pendiente de hoy ya salio, asi que el robot no vuelve a armarlo salvo que entre
    una correccion de comercial. Es lo acordado -*"quiero correr solamente
    replenishment o solamente otras solicitudes, para eso te pido poder borrar el
    archivo"*-.
    """
    filas = []
    for s in sorted(por_sku):
        sol, asig = por_sku[s]
        if sol - asig <= 0:
            continue
        filas.append({'Código de artículo': s,
                      'Cantidad solicitada': int(round(sol)),
                      'Cantidad asignada': int(round(asig))})
    if not filas:
        log('El pendiente no tiene ni un articulo: NO se toca la tarjeta PEDIDOS.',
            'AVISO')
        return False

    cuerpo = json.dumps(filas, ensure_ascii=False).encode('utf-8')
    url = '%s/%s' % (WEB_DATOS_API, AREA_PEDIDOS)
    und = sum(f['Cantidad solicitada'] - f['Cantidad asignada'] for f in filas)
    for i in range(1, intentos + 1):
        try:
            p = urllib.request.Request(url, data=cuerpo, method='POST')
            p.add_header('Content-Type', 'application/json')
            p.add_header('X-Robot-Token', ROBOT_TOKEN)
            with urllib.request.urlopen(p, timeout=300) as resp:
                json.loads(resp.read().decode('utf-8'))
            log('Zona Buffer > Archivo > PEDIDOS: %s articulos, %s unidades (%.1f KB)'
                % (format(len(filas), ',d'), format(int(und), ',d'), len(cuerpo) / 1024.0))
            return True
        except Exception as e:
            if i < intentos:
                log('Intento %d: no se pudo dejar PEDIDOS (%s), se reintenta'
                    % (i, type(e).__name__), 'AVISO')
            else:
                log('NO se pudo dejar el archivo en PEDIDOS (%s: %s). El pendiente SI '
                    'quedo publicado en su submodulo; lo que falta es la tarjeta del '
                    'buffer.' % (type(e).__name__, str(e)[:120]), 'ERROR')
    return False


def subir_excel(ruta, fecha, intentos=3):
    """Sube el Excel al modulo Descargas.

    NO ES MULTIPART. El servidor espera los bytes crudos como
    `application/octet-stream` y los datos en la direccion. Armado como multipart
    el archivo entra igual pero **sin nombre y sin tipo**: el 20-ago-2026 quedo
    subido como `archivo.xlsx` / tipo `archivo`, que ademas le habria peleado el
    cupo a los otros. Es la misma forma que usa `subir_a_la_web` de
    `generar_slotting.py`, que lleva meses andando.

    EL `tipo` IMPORTA: el servidor guarda SIETE DE CADA TIPO, no siete del modulo.
    Sin el, este archivo se repartiria el cupo con el Slotting y los stocks.
    """
    try:
        with io.open(ruta, 'rb') as fh:
            datos = fh.read()
    except Exception as e:
        log('No se pudo leer el Excel: %s' % e, 'ERROR')
        return False

    parametros = urllib.parse.urlencode({
        'nombre': os.path.basename(ruta),
        'fecha': fecha,
        'usuario': 'robot',
        'tipo': 'Pendiente',
    })
    url = '%s/descargas?%s' % (WEB_ARCHIVOS_API, parametros)

    for i in range(1, intentos + 1):
        try:
            p = urllib.request.Request(url, data=datos, method='POST')
            p.add_header('Content-Type', 'application/octet-stream')
            with urllib.request.urlopen(p, timeout=300) as resp:
                r = json.loads(resp.read().decode('utf-8'))
            if r.get('status') == 'success':
                log('Excel subido a Descargas: %s (%.0f KB), quedan %s guardados'
                    % (os.path.basename(ruta), len(datos) / 1024.0, r.get('guardados')))
                return True
            raise RuntimeError(r.get('message', 'respuesta inesperada del servidor'))
        except Exception as e:
            if i < intentos:
                log('Intento %d: no se pudo subir el Excel (%s: %s), se reintenta'
                    % (i, type(e).__name__, str(e)[:120]), 'AVISO')
            else:
                log('No se pudo subir el Excel: %s: %s'
                    % (type(e).__name__, str(e)[:160]), 'ERROR')
    return False


def main():
    probar = '--probar' in sys.argv
    sin_bajar = '--sin-bajar' in sys.argv
    hoy = arg('--fecha', datetime.now().strftime('%Y-%m-%d'))

    log('=' * 62)
    log('PENDIENTE DE DESPACHO  ·  %s' % hoy)
    log('=' * 62)

    # LA FOTO PRIMERO. Solo para el dia de hoy: con --fecha de un dia pasado la foto
    # del WMS no sirve -es de ahora, no de aquel dia- y bajarla serian ocho minutos
    # tirados. Con --probar no se toca el WMS.
    if not probar and not sin_bajar and hoy == datetime.now().strftime('%Y-%m-%d'):
        if not refrescar_pendientes():
            log('')
            log('NO SE PUBLICA NADA. Sin una foto del WMS posterior al correo el '
                'pendiente sale corto y pisaria al bueno; queda el del dia anterior. '
                'El correo lo vuelve a intentar en la proxima vuelta.', 'ERROR')
            return 2
        log('')

    datos, guias, cabecera, IQ, por_guia, por_sku = armar(hoy)
    t = datos['totales']
    o = datos['origen']
    log('')
    log('   pedidos          %s' % format(t['pedidos'], ',d'))
    log('   tiendas          %s' % format(t['tiendas'], ',d'))
    log('   articulos        %s' % format(t['articulos'], ',d'))
    log('   POR ATENDER      %s unidades' % format(t['unidades'], ',d'))
    log('')
    log('   el WMS abre      %s ordenes / %s unidades'
        % (format(o['abiertoWms']['ordenes'], ',d'), format(o['abiertoWms']['unidades'], ',d')))
    log('   comercial mando  %s ordenes / %s unidades'
        % (format(o['mandado']['ordenes'], ',d'), format(o['mandado']['unidades'], ',d')))
    log('   nunca libero     %s ordenes / %s unidades  <- queda fuera'
        % (format(o['noLiberado']['ordenes'], ',d'), format(o['noLiberado']['unidades'], ',d')))

    if probar:
        if '--excel' in sys.argv:
            # El Excel se escribe pero NO se sube: sirve para mirar una corrida sin
            # tocar el pendiente bueno que ya esta publicado.
            ruta = os.path.join(AQUI, 'Pendiente PRUEBA %s.xlsx'
                                % datetime.strptime(hoy, '%Y-%m-%d').strftime('%d-%m-%y'))
            excel(ruta, cabecera, IQ, guias, por_guia, por_sku)
            log('')
            log('Excel de prueba: %s' % ruta)
        log('')
        log('MODO PROBAR: no se publica nada ni se sube ningun archivo.')
        return 0

    nombre = 'Pendiente %s.xlsx' % datetime.strptime(hoy, '%Y-%m-%d').strftime('%d-%m-%y')
    ruta = os.path.join(AQUI, nombre)
    excel(ruta, cabecera, IQ, guias, por_guia, por_sku)
    ok1 = publicar_datos(datos)
    ok2 = subir_excel(ruta, hoy)
    ok3 = publicar_pedidos(por_sku)
    try:
        os.remove(ruta)
    except Exception:
        pass
    if ok1 and ok2 and ok3:
        # EL SELLO ES PARA QUE EL CORREO SEPA QUE YA NO HACE FALTA REINTENTAR.
        # `correo_guias.py` se despierta cada media hora hasta las 23:00; sin esto
        # solo volveria a armar el pendiente si entrara OTRO correo, y una noche en
        # que el WMS estuviera ocupado el dia se quedaria sin pendiente y nadie se
        # enteraria hasta la mañana.
        try:
            os.makedirs(os.path.dirname(SELLO), exist_ok=True)
            io.open(SELLO, 'w', encoding='utf-8').write(hoy)
        except Exception as e:
            log('No se pudo dejar el sello (%s). El pendiente SI se publico; lo '
                'unico que pasa es que el correo va a volver a armarlo.'
                % type(e).__name__, 'AVISO')
    log('')
    log('LISTO · datos %s · excel %s · pedidos %s'
        % ('OK' if ok1 else 'FALLO', 'OK' if ok2 else 'FALLO', 'OK' if ok3 else 'FALLO'))
    return 0 if (ok1 and ok2 and ok3) else 1


if __name__ == '__main__':
    """NADA SE MUERE EN SILENCIO. Misma guarda que `correo_guias.py`: el mensaje de
    SystemExit sale por stderr y una tarea programada no lo ve. Queda en el log."""
    try:
        codigo = main()
    except SystemExit as e:
        codigo = e.code
        if isinstance(codigo, str):
            log(codigo, 'ERROR')
            codigo = 1
    except KeyboardInterrupt:
        log('Cortado a mano.', 'AVISO')
        codigo = 1
    except Exception:
        log('SE CAYO SIN AVISAR:', 'ERROR')
        for linea in traceback.format_exc().rstrip().splitlines():
            log('   ' + linea, 'ERROR')
        codigo = 1
    sys.exit(codigo)
