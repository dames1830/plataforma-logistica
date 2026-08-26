# -*- coding: utf-8 -*-
"""
Genera el reporte Slotting a partir de los archivos que baja el robot de Oracle WMS.

Reemplaza el proceso manual de Power Query del libro
"Replenishment Dia - UssGeralFordv2.xlsm" (hoja del reporte de slotting).

Fuentes:
    scraping Stock\\Stock Activo\\Stock Activo DD-MM-AA.csv     (lo baja el robot)
    scraping Stock\\Stock Reserva\\Stock Reserva DD-MM-AA.xlsx  (lo baja el robot)
    scraping Stock\\Archivos\\Maestro_Articulos.xlsx            (lo actualiza Daniel)
    scraping Stock\\Archivos\\Marcas.xlsx                       (lo actualiza Daniel)

Salida:
    scraping Stock\\Slotting\\Slotting DD-MM-AA.xlsx
        hoja "Slotting" : tabla dinámica lista para usar
        hoja "Datos"    : la tabla plana que la alimenta

Reglas, tal como estaban en el Power Query original:
  - Activo:  Área -> NIVEL, Artículo -> SKU, Cantidad actual -> QTY
  - Reserva: sólo sucursal 50008; SKU = PRODUCTO, o ARTICULO si PRODUCTO viene vacío
  - Se omiten las ubicaciones que empiezan con CDBUFFER-C
  - MZN03-01/02/03/07 -> "Zona Industrial", MZN03-04/05/06 -> "Zona Marie Claire"
  - Qty Zona    = NIVEL en AND, MZN01..MZN04, PARED, SEL
    Qty Buffer  = NIVEL CDBUFFER
    Qty Reserva = NIVEL ALTO, salvo ubicaciones SEL-14
  - La talla sale del final de la descripción (...BUBBLEGUMMERS-1-23 -> 23)
  - Marca: MarcaStd del Maestro traducida con Marcas.xlsx
  - Lo que no está en el Maestro queda con guión
"""

import csv
import decimal
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────── Configuración ───────────────────────────────

def _base_onedrive():
    """
    La carpeta de OneDrive donde están los archivos. SE BUSCA, NO SE ESCRIBE A MANO.

    En la laptop el usuario de Windows es 'dames' y en el servidor 'Administrator', así que
    una ruta fija sirve en una máquina y falla en la otra. El 05-ago-2026 se copió al
    servidor la versión de la laptop y el robot salió con "Falta: Stock Activo, Falta:
    Maestro_Articulos.xlsx" aunque los cuatro archivos estaban en su lugar: buscaba en
    C:\\Users\\dames, que en el servidor no existe.

    Se prueban las candidatas en orden y se usa la primera que exista de verdad.
    """
    candidatas = [
        os.environ.get("OneDrive"),                          # lo que dice el propio Windows
        os.environ.get("OneDriveCommercial"),
        os.path.join(os.path.expanduser("~"), "OneDrive"),   # el usuario que esté corriendo
        r"C:\Users\Administrator\OneDrive",                  # el servidor
        r"C:\Users\dames\OneDrive",                          # la laptop
    ]
    for c in candidatas:
        if not c:
            continue
        ruta = os.path.join(c, "danielames.bata", "scraping Stock")
        if os.path.isdir(ruta):
            return ruta
    # Ninguna existe: se devuelve la del usuario actual para que el error diga dónde buscó.
    return os.path.join(os.path.expanduser("~"), "OneDrive", "danielames.bata", "scraping Stock")


BASE = _base_onedrive()
DIR_ACTIVO = os.path.join(BASE, "Stock Activo")
DIR_RESERVA = os.path.join(BASE, "Stock Reserva")
DIR_ARCHIVOS = os.path.join(BASE, "Archivos")
DIR_SALIDA = os.path.join(BASE, "Slotting")

MAESTRO = os.path.join(DIR_ARCHIVOS, "Maestro_Articulos.xlsx")
MARCAS = os.path.join(DIR_ARCHIVOS, "Marcas.xlsx")

LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")

SUCURSAL = 50008
NIVELES_ZONA = {"AND", "MZN01", "MZN02", "MZN03", "MZN04", "PARED", "SEL"}
ZONA_INDUSTRIAL = ("MZN03-01", "MZN03-02", "MZN03-03", "MZN03-07")
ZONA_MARIE_CLAIRE = ("MZN03-04", "MZN03-05", "MZN03-06")

# Un SKU válido es 7 dígitos, guion, un dígito, guion y la talla
SKU_VALIDO = re.compile(r"^\d{7}-\d-\d+$")
# La talla está al final de la descripción, después de un patrón -N-
TALLA = re.compile(r"-([1-9])-([A-Z0-9.ÁÉÍÓÚÑ]+)\s*$", re.I)

SIN_DATO = "-"

# ─────────────────────── Publicación en la plataforma web ───────────────────────
# El archivo terminado se sube tal cual a la web, para que los asistentes lo bajen
# desde Inventario > Descargas y lo abran en Excel con su tabla dinámica.
#
# Encendido desde el 01-ago-2026, cuando se publicó la v29.0005 del backend con
# los endpoints /api/archivos. Antes de esa fecha tenía que estar apagado: cada
# intento habría dado error y marcado la corrida como fallida sin que nada
# estuviera realmente mal.
WEB_SUBIR = True
WEB_API = "https://logistics-backend-wv0x.onrender.com/api/archivos"
# Todo lo descargable vive junto, en un módulo propio con buscador. Antes iba a
# "inventario", y los archivos que quedaron ahí los mueve el backend al arrancar.
WEB_MODULO = "descargas"

# "produccion" usa la base real; "beta" la de pruebas. Comparten servidor pero no
# base de datos. Para mandar una corrida a pruebas sin tocar este archivo:
#     set PULSE_ENTORNO=beta
WEB_ENTORNO = os.environ.get("PULSE_ENTORNO", "produccion")

# ── LOS STOCKS A LA NUBE ─────────────────────────────────────────────────────
#
# Además del reporte, el robot publica el Stock Activo y el de Reserva como DATOS,
# para que la plataforma los lea desde cualquier PC. Antes había que cargarlos a
# mano en cada computadora: por eso el 02-ago-2026 dos PC dieron papeles distintos
# —una tenía el stock del día y la otra el de hacía semanas— y por eso la reserva
# que usaba el cálculo llevaba un mes sin actualizarse.
#
# SE SUBEN SOLO LAS COLUMNAS QUE LA PLATAFORMA USA, Y EN EL MISMO ORDEN. Parte del
# código las lee POR POSICIÓN y no por nombre: la 1 es el artículo y la 2 la
# descripción de donde sale la talla. Mover o quitar una de las primeras seis las
# rompe en silencio, sin ningún error a la vista. De 33 columnas quedan 6, y el
# área pasa de 30,7 MB a unos 6.
#
# Van con ?date=MASTER para que cada corrida REEMPLACE a la anterior. Sin eso el
# servidor guardaría un snapshot por día y llenaría su disco de 1 GB en un mes.
WEB_SUBIR_STOCKS = True
WEB_DATOS_API = "https://logistics-backend-wv0x.onrender.com/api/logistics"
# EL TOKEN DEL ROBOT. Desde v29.0415 el servidor puede EXIGIR credencial para
# escribir datos (ver EXIGIR_TOKEN_ESCRITURA en backend/main.py). El robot no tiene
# sesion, asi que lleva su propio token, leido del entorno del Contabo -NUNCA escrito
# aca, o estaria publico en el repo-. Si la variable no esta, se manda vacio y el
# servidor, mientras el candado siga apagado, lo deja pasar igual.
ROBOT_TOKEN = os.environ.get('ROBOT_TOKEN', '')

AREA_ACTIVO = "almacenaje_activo"
AREA_RESERVA = "analisis_sku_reserva"

# ── EL CAJÓN DE LA HORA ──────────────────────────────────────────────────────
#
# Las dos áreas de arriba son la FOTO DEL TURNO y no se mueven hasta la corrida
# siguiente: sobre ellas se calculan el Replenishment, la Zona Buffer, las tareas de
# almacenaje y la meta de Limpieza del Buffer C. Estas dos, en cambio, las reescribe
# `stock_por_hora.py` cada hora, y las leen únicamente los reportes de avance —el mapa
# de calor del Slotting y el Cumplimiento del turno—.
#
# La corrida principal las publica TAMBIÉN, con las mismas filas que acaba de leer y
# sin bajar nada de nuevo. Si no lo hiciera, a las 19:00 el reporte del turno arrancaría
# comparando la foto nueva contra un stock de hasta media hora antes —de cuando el turno
# día todavía estaba trabajando— y el primer avance de la noche saldría con pares que no
# almacenó nadie de la noche.
#
# El nombre del activo es histórico: nació para el mapa de calor. No se cambia porque ya
# está elegido para esquivar dos reglas de csvHub_v6.js —AREA_CANONICA, que lo repartiría
# a buffer_activo y compañía, y el disparo de updateTablaTallas() para todo lo que termina
# en _activo o _reserva—. Por eso la reserva es `reserva_hora` y no `stock_hora_reserva`.
AREA_ACTIVO_HORA = "layout_stock_hora"
AREA_RESERVA_HORA = "reserva_hora"

# LA FOTO DEL BUFFER C AL ARRANCAR EL TURNO.
#
# El stock va con ?date=MASTER, o sea que la corrida de las 06:00 pisa la de las
# 19:00 y al día siguiente ya no se sabe con cuánto arrancó la noche. Sin ese dato
# no hay meta para la Limpieza de Buffer C.
#
# Guardar el stock entero dos veces sería 12 MB por día. Acá va SOLO el Buffer C,
# que son unas 160 líneas —8 KB—: artículo y cantidad, nada más. Un año son 3 MB.
#
# Y va la LISTA, no el total: restar totales no sirve. La noche del 10-ago el
# Buffer C cerró con MÁS de lo que empezó —1.759 contra 1.820— porque entraron 930
# pares nuevos, y la resta daba −61 cuando el equipo había sacado 869. La cuenta
# es artículo por artículo.
AREA_BUFFER_C = "buffer_c_arranque"

# LAS PALETAS ALTAS AL ARRANCAR EL TURNO.
#
# Misma idea que el Buffer C y por el mismo motivo: sin la foto del arranque no hay
# contra qué comparar, y la del arranque se pierde a las 06:00.
#
# Y la misma trampa, que Daniel señaló antes de que la viéramos en los datos: durante
# el turno también SUBE mercadería. La noche del 10-ago bajaron 97 paletas y subieron
# 24; restar los totales daba 49. La cuenta es POR PALETA (LPN): cuáles de las que
# estaban arriba ya no están arriba.
#
# "Arriba" es NIVEL=ALTO, que en este almacén es exactamente el selectivo (SEL-*).
# Las otras filas del reporte son MERMA, DEV, RECEP, INS y AEREO, que no son paletas
# de reserva. De las 97 que bajaron esa noche, 73 salieron del reporte y 24 estaban
# en MZM-TRANS-00-01, o sea bajadas y en tránsito al mezzanine: por eso no alcanza
# con mirar las que desaparecen.
#
# Son unas 1.550 paletas, 40 KB por noche, 15 MB al año.
AREA_RESERVA_ARRANQUE = "reserva_arranque"

# ── LAS FOTOS DEL CIERRE ─────────────────────────────────────────────────────
#
# El reporte del turno mide contra la foto de la hora mientras la jornada corre, pero
# a las 06:30 la jornada cambia y ese cajón va con MASTER: a media mañana ya es otra
# foto y no hay contra qué recalcular. Sin estas dos, el avance de la noche se queda
# congelado en la última medición que alcanzó a guardar la pantalla.
#
# LA NOCHE DEL 17-AGO-2026 LO DEJÓ A LA VISTA: la última foto fue la de las 05:36, el
# turno cerró a las 06:30, y la Bajada de paletas quedó clavada en 47 de 166 con la
# separación en 97 de 2.189. Estaban escritas en la memoria del proyecto como si el
# robot ya las guardara; el código nunca las tuvo — la corrida de la mañana decía
# "las fotos del arranque no se tocan" y se iba sin dejar nada.
#
# Van con la fecha de la JORNADA QUE TERMINA, no la del día en que se corre.
AREA_BUFFER_C_CIERRE = "buffer_c_cierre"
AREA_RESERVA_CIERRE = "reserva_cierre"

# ── NO PUBLICAR UN STOCK QUE NO SEA DE ESTA CORRIDA ──────────────────────────
#
# El 06-ago-2026 la descarga de las 19:00 no produjo archivo. Este generador buscó
# "el más reciente del día", encontró el de las 08:23 y LO PUBLICÓ COMO SI FUERA EL DE
# LAS 19:00: mismo contenido, marca de hora nueva. Nadie se enteró.
#
# Con ese stock viejo Daniel corrió la ola de las 19:28, que volvió a generar tareas
# sobre mercadería que el turno día ya había almacenado — 870 pares que el operario iba
# a ir a buscar y no estaban. Las hojas ya estaban repartidas cuando se descubrió.
#
# Ahora hay dos defensas:
#   1. El robot le pasa por parámetro el archivo EXACTO que acaba de bajar (--activo,
#      --reserva). Sin adivinar cuál es "el más reciente".
#   2. Si igual hay que buscarlo -corrida a mano-, el archivo no puede tener más de
#      MAX_HORAS_STOCK horas. Si las tiene, NO se publica y se dice por qué.
#
# Para reprocesar un archivo viejo a propósito está --igualmente.
MAX_HORAS_STOCK = 3

# ── LAS DOS CORRIDAS DEL DÍA ─────────────────────────────────────────────────
#
# Desde el 06-ago-2026 el robot corre a las 06:00 y a las 19:00, para que el turno día
# no tenga que trabajar con el stock de anoche. Las dos hacen lo mismo salvo en una
# cosa: LOS STOCKS COMO ARCHIVO DESCARGABLE SOLO LOS DEJA LA CORRIDA DE LA NOCHE.
#
# El motivo es de Daniel y es simple: en Descargas se baja el stock con el que cerró el
# día, y tener ahí también el de la mañana no le sirve a nadie — serían dos archivos
# casi iguales y habría que mirar la hora para saber cuál es cuál.
#
# Lo que sí hacen las dos: bajar los archivos a OneDrive (con la hora en el nombre, así
# no se pisan) y publicar el stock como DATO, que es de donde salen las tareas. Ahí sí
# manda siempre el último, y a la noche el de la noche reemplaza al de la mañana.
HORA_CORTE_NOCHE = 12


def es_corrida_de_la_noche():
    """
    De mediodía en adelante es la corrida de la noche. El corte va al mediodía y no a
    las 19:00 a propósito: si una noche el robot arranca tarde o se lo corre a mano a
    las 21:00, sigue siendo la de la noche.

    Se puede forzar con la variable PULSE_TURNO, para cuando hay que correr el robot a
    mano fuera de hora y sí se quieren los archivos en Descargas:
        set PULSE_TURNO=noche     -> se comporta como la corrida de la noche
        set PULSE_TURNO=manana    -> se comporta como la de la mañana
    """
    forzado = os.environ.get("PULSE_TURNO", "").strip().lower()
    if forzado in ("noche", "night", "n"):
        return True
    if forzado in ("manana", "mañana", "dia", "día", "m", "d"):
        return False
    return datetime.now().hour >= HORA_CORTE_NOCHE

# ─────────────────────────────────────────────────────────────────────────────
# LA TABLA DE TALLAS
#
# La talla no es un dato: el WMS no la manda como columna, está metida al final del
# texto de la descripción —"...BUBBLEGUMMERS-1-18"— y hay que deducirla leyendo. Eso
# se venía haciendo en cada pantalla y con reglas distintas, y las que tenían que
# cruzarse no se encontraban: en una polera talla M, el generador de tareas guardaba
# "M" y el cálculo del papel decía "sin talla", así que la hoja salía con todo al piso
# y el destino en blanco, sin ningún aviso.
#
# El robot es el mejor lugar para armarla: es el único que tiene los DOS stocks juntos.
# La web puede no haber visto nunca un artículo que solo está en reserva.
#
# ES ACUMULATIVA, y eso importa por cuatro razones:
#   - la talla de un SKU no cambia nunca, así que no hay nada que rehacer;
#   - un artículo que se agotó y salió del stock conserva su talla para cuando vuelva;
#   - si una noche el robot no corre, la tabla de ayer sigue siendo válida;
#   - y una talla corregida a mano no se pisa sola en la corrida siguiente.
WEB_SUBIR_TALLAS = True
AREA_TALLAS = "tabla_tallas"
TIPO_TALLAS = "Tabla de Tallas"
# Cuántas versiones guarda Descargas. Daniel pidió seis. Como solo se publica cuando
# de verdad cambió algo, seis versiones son seis cambios, no seis días.
TALLAS_EN_DESCARGAS = 6
COLS_TALLAS = ["SKU", "TALLA"]

# Las seis primeras del CSV, tal cual vienen. El ORDEN es parte del contrato.
COLS_ACTIVO = ["Área", "Artículo", "Descripción de artículo",
               "Ubicación", "Cantidad actual", "Cantidad asignada"]

# Para el Excel que se descarga, en el orden en que se lee de corrido. Van sin UBI_KEY ni
# ES_ALTO, que son de uso interno de la plataforma y a quien abre la planilla no le dicen
# nada. Los datos que consulta la web sí llevan las ocho.
COLS_RESERVA = ["UBICACION", "LPN", "PRODUCTO", "DESCRIPCION", "CANTIDAD", "NIVEL"]

# Las tres últimas se llaman distinto a los campos de valor de la dinámica
# ("Qty Buffer", "Qty Zona", "Qty Reserva") porque Excel no acepta que un campo
# de valor tenga el mismo nombre que su columna de origen.
COLUMNAS = ["Articulo", "UBICACION", "SKU", "Tallas", "Marcas",
            "Gender RIMS", "Temporada", "NIVEL",
            "Buffer", "Zona", "Reserva"]

# columna de origen -> nombre que se muestra en la dinámica
VALORES = [("Buffer", "Qty Buffer"), ("Zona", "Qty Zona"), ("Reserva", "Qty Reserva")]

_log_file = None


def log(msg, nivel="INFO"):
    linea = "[%s] [%-5s] %s" % (datetime.now().strftime("%H:%M:%S"), nivel, msg)
    try:
        print(linea)
    except UnicodeEncodeError:
        print(linea.encode("ascii", "replace").decode("ascii"))
    if _log_file:
        try:
            with io.open(_log_file, "a", encoding="utf-8") as f:
                f.write(linea + "\n")
        except Exception:
            pass


# ─────────────────────────────── Lectura ───────────────────────────────

def zona_especial(ubi):
    """MZN03 se muestra por zona comercial en vez del código crudo."""
    if ubi.startswith(ZONA_INDUSTRIAL):
        return "Zona Industrial"
    if ubi.startswith(ZONA_MARIE_CLAIRE):
        return "Zona Marie Claire"
    return None


def extraer_talla(desc):
    """La talla viene pegada al final de la descripción: ...BUBBLEGUMMERS-1-23"""
    if not desc:
        return SIN_DATO
    m = TALLA.search(str(desc).strip())
    return m.group(2).strip() if m else SIN_DATO


def leer_activo(ruta):
    filas = []
    with io.open(ruta, encoding="utf-8-sig", errors="replace") as fh:
        lector = csv.reader(fh, delimiter=";")
        next(lector, None)
        for row in lector:
            if len(row) < 5 or not row[1]:
                continue
            ubi = row[3].strip()
            z = zona_especial(ubi)
            try:
                qty = float(row[4] or 0)
            except ValueError:
                qty = 0.0
            filas.append({
                "NIVEL": row[0].strip(),
                "SKU": row[1].strip(),
                "DESC": row[2].strip(),
                "UBICACION": "%s - %s" % (z, ubi) if z else ubi,
                "QTY": qty,
            })
    return filas


def leer_reserva(ruta):
    filas = []
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for f in ws.iter_rows(min_row=4, values_only=True):
        if f[0] is None:
            continue
        try:
            if int(f[0]) != SUCURSAL:
                continue
        except (ValueError, TypeError):
            continue
        if f[7] is None:
            continue
        # Cuando PRODUCTO viene vacío se usa el ARTICULO, que es lo que se hacía a mano
        articulo = str(f[7]).strip()
        producto = "" if f[8] is None else str(f[8]).strip()
        try:
            qty = float(f[10] or 0)
        except (ValueError, TypeError):
            qty = 0.0
        filas.append({
            "NIVEL": str(f[1] or "").strip(),
            "SKU": producto if producto else articulo,
            "DESC": str(f[9] or "").strip(),
            "UBICACION": str(f[4] or "").strip(),
            "QTY": qty,
        })
    wb.close()
    return filas


def leer_maestro(ruta):
    """CodArticulo -> (MarcaStd, Gender RIMS, Temporada)"""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mapa = {}
    for f in ws.iter_rows(min_row=2, values_only=True):
        if not f[1]:
            continue
        cod = str(f[1]).strip().zfill(7)
        mapa[cod] = (
            str(f[8] or SIN_DATO).strip(),    # I MarcaStd
            str(f[3] or SIN_DATO).strip(),    # D Gender RIMS
            str(f[14] or SIN_DATO).strip(),   # O Temporada
        )
    wb.close()
    return mapa


def leer_marcas(ruta):
    """MarcaStd -> Marcas"""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mapa = {}
    for f in ws.iter_rows(min_row=2, values_only=True):
        if f[0]:
            mapa[str(f[0]).strip().upper()] = str(f[1] or SIN_DATO).strip()
    wb.close()
    return mapa


# ─────────────────────────────── Armado ───────────────────────────────

def construir(filas, maestro, marcas):
    """Aplica exclusiones, reparte las cantidades y agrega los datos del artículo."""
    resultado = []
    fuera = {"CDBUFFER-C": 0, "SKU inválido": 0, "NIVEL no usado": 0, "ALTO en SEL-14": 0}
    sin_maestro = set()

    for x in filas:
        ubi_up = x["UBICACION"].upper()

        if ubi_up.startswith("CDBUFFER-C"):
            fuera["CDBUFFER-C"] += 1
            continue
        if not SKU_VALIDO.match(x["SKU"]):
            fuera["SKU inválido"] += 1
            continue

        nivel, qty = x["NIVEL"], x["QTY"]
        buf = zona = res = 0.0

        if nivel == "CDBUFFER":
            buf = qty
        elif nivel in NIVELES_ZONA:
            zona = qty
        elif nivel == "ALTO":
            if ubi_up.startswith("SEL-14"):
                fuera["ALTO en SEL-14"] += 1
                continue
            res = qty
        else:
            fuera["NIVEL no usado"] += 1
            continue

        articulo = x["SKU"][:7]
        datos = maestro.get(articulo)
        if datos:
            marca_std, gender, temporada = datos
            marca = marcas.get(marca_std.upper(), marca_std) or SIN_DATO
        else:
            sin_maestro.add(articulo)
            marca = gender = temporada = SIN_DATO

        resultado.append([
            articulo, x["UBICACION"], x["SKU"], extraer_talla(x["DESC"]),
            marca, gender, temporada, nivel,
            buf, zona, res,
        ])

    return resultado, fuera, sin_maestro


# ─────────────────────────────── Escritura ───────────────────────────────

def escribir_datos(ruta, filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    ws.append(COLUMNAS)
    cab = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="1C2B3A")
    for c in range(1, len(COLUMNAS) + 1):
        celda = ws.cell(row=1, column=c)
        celda.font = cab
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center")

    for f in filas:
        ws.append(f)

    anchos = [11, 34, 16, 8, 18, 20, 14, 14, 12, 12, 12]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"

    for fila in ws.iter_rows(min_row=2, min_col=9, max_col=11):
        for celda in fila:
            celda.number_format = "#,##0"

    wb.save(ruta)


def _pid_de(excel):
    """PID de esa instancia de Excel, para poder cerrarla si se queda colgada."""
    try:
        import ctypes
        pid = ctypes.c_ulong()
        ctypes.windll.user32.GetWindowThreadProcessId(int(excel.Hwnd), ctypes.byref(pid))
        return pid.value
    except Exception:
        return 0


def _matar(pid):
    """Cierra a la fuerza ese Excel, y sólo ese, si sobrevivió al Quit."""
    if not pid:
        return
    try:
        # Los dos filtros juntos: aunque Windows hubiera reciclado el PID, no se
        # mata nada que no sea Excel.
        subprocess.run(
            ["taskkill", "/F", "/FI", "PID eq %d" % pid, "/FI", "IMAGENAME eq EXCEL.EXE"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=30,
        )
    except Exception:
        pass


def abrir_excel_propio():
    """
    Abre una instancia de Excel propia, invisible y aparte de la que el usuario
    tenga abierta.

    DispatchEx obliga a arrancar un proceso nuevo. Con Dispatch o EnsureDispatch,
    Windows devuelve el Excel que el usuario ya tiene abierto, y ahí pasan dos
    cosas malas: el reporte se arma a la vista, porque Visible=False no puede
    ocultar una ventana que tiene libros del usuario, y el Quit() del final le
    cierra esos libros.
    """
    import win32com.client as win32

    # Enlace tardío a propósito: los constantes de Excel que usa este script ya
    # están escritos como números, y así no depende de la caché de makepy, que
    # bajo el Programador de tareas se corrompe con facilidad.
    excel = win32.DispatchEx("Excel.Application")
    pid = _pid_de(excel)

    # Una instancia recién creada no tiene libros abiertos. Si los tiene, es la
    # del usuario: se corta acá y no se la toca ni se la cierra.
    if excel.Workbooks.Count:
        raise RuntimeError(
            "Excel devolvió una instancia con %d libro(s) abiertos; se cancela para "
            "no interferir con el usuario" % excel.Workbooks.Count)

    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    try:
        excel.AskToUpdateLinks = False
    except Exception:
        pass
    log("Excel propio abierto en segundo plano (proceso %s)" % (pid or "?"))
    return excel, pid


def agregar_dinamica(ruta, n_filas):
    """
    Crea la tabla dinámica con Excel. Es la única forma de generar una dinámica
    nativa: ninguna librería de Python las escribe desde cero.
    Si Excel no está disponible, el archivo igual queda con la hoja Datos usable.
    """
    excel, pid = abrir_excel_propio()
    libro = None
    try:
        libro = excel.Workbooks.Open(ruta)

        for h in list(libro.Worksheets):
            if h.Name == "Slotting":
                h.Delete()

        hd = libro.Worksheets("Datos")
        hs = libro.Worksheets.Add(Before=hd)
        hs.Name = "Slotting"

        # Argumentos por posición, no por nombre: con enlace tardío pywin32 no
        # acepta argumentos con nombre.
        rango = "Datos!R1C1:R%dC%d" % (n_filas + 1, len(COLUMNAS))
        cache = libro.PivotCaches().Create(1, rango)                    # SourceType=xlDatabase
        tabla = cache.CreatePivotTable("Slotting!R4C1", "SlottingPivot")  # destino, nombre

        # Primero los campos, después el formato: al revés Excel rechaza los cambios
        CAMPOS_FILA = ["Articulo", "UBICACION", "SKU", "Tallas",
                       "Marcas", "Gender RIMS", "Temporada"]
        for campo in CAMPOS_FILA:
            tabla.PivotFields(campo).Orientation = 1        # xlRowField

        tabla.PivotFields("NIVEL").Orientation = 3          # xlPageField

        for origen, etiqueta in VALORES:
            df = tabla.AddDataField(tabla.PivotFields(origen), etiqueta, -4157)  # xlSum
            df.NumberFormat = "#,##0"

        # Diseño tabular: cada campo en su propia columna, como el original
        tabla.RowAxisLayout(1)          # xlTabularRow
        try:
            tabla.RepeatAllLabels(2)    # xlRepeatLabels
        except Exception:
            pass                        # no está en todas las versiones de Excel
        tabla.ColumnGrand = False       # sin total de columnas
        tabla.RowGrand = True           # total general al pie

        # Sólo el artículo lleva subtotal, igual que en el reporte manual
        for campo in CAMPOS_FILA[1:]:
            try:
                tabla.PivotFields(campo).Subtotals = tuple([False] * 12)
            except Exception:
                pass

        hs.Range("A1").Value = "REPORTE SLOTTING"
        hs.Range("A1").Font.Size = 14
        hs.Range("A1").Font.Bold = True
        hs.Range("A2").Value = "Generado el %s desde Stock Activo, Stock Reserva y Maestro de Artículos" % \
                               datetime.now().strftime("%d/%m/%Y %H:%M")
        hs.Range("A2").Font.Italic = True

        hd.Visible = 0                  # xlSheetHidden
        hs.Activate()
        libro.Save()
        return True
    finally:
        # El cierre nunca puede tumbar la corrida: si algo falló, el archivo ya
        # está en disco con la hoja Datos. SaveChanges=False porque lo que había
        # que guardar se guardó arriba.
        try:
            if libro is not None:
                libro.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        # Si Excel se colgó, Quit() no lo cierra y queda un proceso zombi que
        # rompe la corrida del día siguiente. Se lo cierra por PID.
        time.sleep(2)
        _matar(pid)


def crear_dinamica(ruta, n_filas, intentos=2):
    """Arma la dinámica; reintenta una vez, porque Excel a veces corta la
    conexión en el primer intento y al segundo levanta sin problema."""
    for intento in range(1, intentos + 1):
        try:
            agregar_dinamica(ruta, n_filas)
            log("Tabla dinámica lista")
            return True
        except Exception as e:
            detalle = "%s: %s" % (type(e).__name__, str(e)[:200])
            if intento < intentos:
                log("Intento %d: no se pudo crear la dinámica (%s), se reintenta..."
                    % (intento, detalle), "WARN")
                time.sleep(10)
            else:
                log("No se pudo crear la dinámica: %s" % detalle, "ERROR")
    return False


def a_fecha_iso(fecha):
    """31-07-26 -> 2026-07-31.

    El servidor ordena los archivos por este campo para decidir cuál es el más
    viejo y cuál se borra. En DD-MM-AA el orden alfabético no es el cronológico
    ('01-08-26' iría antes que '31-07-26') y rotaría el archivo equivocado.
    """
    try:
        return datetime.strptime(fecha, "%d-%m-%y").strftime("%Y-%m-%d")
    except ValueError:
        return datetime.now().strftime("%Y-%m-%d")


def subir_a_la_web(ruta, fecha, tipo="", intentos=3, guardar=0):
    """
    Sube el archivo terminado a la plataforma. Devuelve si salió bien.

    El `tipo` agrupa las versiones del mismo archivo a lo largo de los días: el servidor
    guarda SIETE de cada tipo, no siete del módulo. Sin eso, el Slotting y los dos stocks
    se repartirían el mismo cupo y quedarían dos días de cada uno. Si va vacío, el servidor
    lo deduce del nombre quitándole la fecha.
    """
    if not WEB_SUBIR:
        log("Subida a la web desactivada (WEB_SUBIR = False), se omite")
        return True

    # NINGÚN ARCHIVO DE LA CORRIDA DE LA MAÑANA VA A LA WEB. Es decisión de Daniel: en
    # Descargas se baja lo del cierre del día, y tener también lo de la mañana serían
    # archivos casi iguales entre los que hay que elegir mirando la hora.
    #
    # El corte va acá adentro y no en cada llamada A PROPÓSITO: así vale para el
    # Slotting, para los dos stocks, para la tabla de tallas y para cualquier archivo
    # que se agregue después, sin que haya que acordarse de ponerle el if.
    #
    # OJO: esto NO frena los DATOS. Los stocks se publican igual de mañana —con
    # subir_datos(), que es otra función— porque de ahí salen las tareas del turno día.
    if not es_corrida_de_la_noche():
        log("Corrida de la mañana: '%s' no va a Descargas (los archivos los deja la "
            "corrida de la noche)" % (tipo or os.path.basename(ruta)))
        return True

    with io.open(ruta, "rb") as fh:
        datos = fh.read()

    parametros = {
        "nombre": os.path.basename(ruta),
        "fecha": a_fecha_iso(fecha),
        "usuario": "robot",
        "tipo": tipo,
    }
    # Cuántas versiones guardar de este tipo. Sin esto el servidor usa las siete de
    # siempre; un servidor viejo que no conozca el parámetro simplemente lo ignora.
    if guardar:
        parametros["guardar"] = guardar
    url = "%s/%s?%s" % (WEB_API, WEB_MODULO, urllib.parse.urlencode(parametros))

    for intento in range(1, intentos + 1):
        try:
            pedido = urllib.request.Request(url, data=datos, method="POST")
            pedido.add_header("Content-Type", "application/octet-stream")
            if WEB_ENTORNO == "beta":
                pedido.add_header("X-Environment", "beta")

            # El servidor puede estar dormido y tardar casi un minuto en despertar.
            with urllib.request.urlopen(pedido, timeout=300) as resp:
                respuesta = json.loads(resp.read().decode("utf-8"))

            if respuesta.get("status") == "success":
                log("Subido a la web (%s): %s MB, quedan %s archivos guardados%s"
                    % (respuesta.get("entorno", WEB_ENTORNO), respuesta.get("mb"),
                       respuesta.get("guardados"),
                       ", se borró el más viejo" if respuesta.get("borrados") else ""))
                return True
            raise RuntimeError(respuesta.get("message", "respuesta inesperada del servidor"))

        except Exception as e:
            detalle = "%s: %s" % (type(e).__name__, str(e)[:200])
            if intento < intentos:
                log("Intento %d: no se pudo subir (%s), se reintenta..." % (intento, detalle), "WARN")
                time.sleep(20)
            else:
                log("No se pudo subir el archivo a la web: %s" % detalle, "ERROR")
                log("El archivo igual quedó en OneDrive, se puede subir a mano", "WARN")
    return False


def _txt(v):
    """El valor tal cual lo dejaba el navegador: texto sin espacios de sobra."""
    return "" if v is None else str(v).strip()


def _lpn(v):
    """
    El LPN viene del Excel como número y hay que dejarlo igual que lo dejaba la web.
    JavaScript escribe 201003000000003780; str() en Python daría 2.01e+17, que no es
    el mismo dato y no cruzaría con nada.

    Y NO ALCANZA CON int(): un LPN tiene 18 dígitos y no entra exacto en un número de
    computadora, así que int() se lleva los últimos por delante — 201003000000003780
    salía 201003000000003776. Son pocos (4 de 17.802 el 05-ago-2026) pero ese número se
    IMPRIME en el Excel de reposición para que el operario encuentre la paleta, y un
    LPN que no existe no se encuentra. repr() da la forma corta que representa a ese
    número, la misma que escribe JavaScript, y Decimal la lee sin volver a redondear.
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(decimal.Decimal(repr(v))))
    return str(v).strip()


def datos_activo_web(ruta):
    """El Stock Activo con las seis columnas que usa la plataforma, en su orden."""
    filas = []
    with io.open(ruta, encoding="utf-8-sig", errors="replace") as fh:
        lector = csv.reader(fh, delimiter=";")
        next(lector, None)          # el encabezado
        for row in lector:
            if len(row) < 6 or not row[1]:
                continue
            filas.append(dict(zip(COLS_ACTIVO, [c.strip() for c in row[:6]])))
    return filas


def foto_buffer_c(filas_activo):
    """
    El Buffer C de esta corrida, artículo por artículo.

    Es prepack, y el generador de tareas lo deja fuera a propósito, así que limpiarlo
    no deja rastro en ninguna tarea: la única forma de medirlo es comparar dos fotos
    del stock. Esta es la del arranque del turno.
    """
    def qty_de(f):
        try:
            return float(str(f.get("Cantidad actual") or 0).replace(",", ""))
        except ValueError:
            return 0.0

    por_articulo = {}
    total = 0
    for f in filas_activo:
        ubi = (f.get("Ubicación") or "").strip().upper()
        if not ubi.startswith("CDBUFFER-C"):
            continue
        art = (f.get("Artículo") or "").strip()
        if not art:
            continue
        qty = qty_de(f)
        if qty <= 0:
            continue
        por_articulo[art] = por_articulo.get(art, 0) + qty
        total += qty

    # LO QUE ESOS MISMOS SKU TIENEN FUERA DEL BUFFER C, dentro del activo.
    #
    # Es la LINEA DE BASE del avance de Limpieza de Buffer C. Sin ella no se puede
    # distinguir un par que llego a su destino de uno que quedo encajado en un LPN
    # todavia sin matricular, y el avance se infla con trabajo a medias: la noche del
    # 12-ago-2026 salieron 1.110 pares del C y solo 138 aparecieron en otra ubicacion.
    #
    # Y hace falta la BASE, no solo la foto de despues, porque un SKU puede estar en el
    # Buffer C y en el mezzanine a la vez: sin el punto de partida, lo que ya estaba
    # afuera se contaria como destino.
    #
    # La misma funcion arma la foto del arranque (19:00) y la del cierre (06:30), asi
    # que las dos quedan con su base y la jornada se puede medir despues de cerrada.
    # Son unos 90 SKU, menos de 2 KB por foto.
    fuera = {}
    for f in filas_activo:
        ubi = (f.get("Ubicación") or "").strip().upper()
        if ubi.startswith("CDBUFFER-C"):
            continue
        art = (f.get("Artículo") or "").strip()
        if art not in por_articulo:
            continue
        qty = qty_de(f)
        if qty > 0:
            fuera[art] = fuera.get(art, 0) + qty

    ahora = time.localtime()
    return {
        "fecha": time.strftime("%Y-%m-%d", ahora),
        "hora": time.strftime("%H:%M", ahora),
        "pares": int(round(total)),
        "articulos": len(por_articulo),
        "detalle": {a: int(round(q)) for a, q in por_articulo.items()},
        "fuera": {a: int(round(q)) for a, q in fuera.items()},
    }


def paletas_altas(filas_reserva):
    """
    Las paletas que están ARRIBA, una entrada por LPN con sus pares.

    Se cuenta por LPN y no por línea: una paleta trae varios artículos y sería varias
    líneas. Y se mira ES_ALTO —NIVEL=ALTO, que acá es el selectivo— porque el resto
    del reporte es MERMA, DEV, recepción e inspección, que no son paletas de reserva.
    """
    por_lpn = {}
    for f in filas_reserva:
        if not f.get("ES_ALTO"):
            continue
        lpn = str(f.get("LPN") or "").strip()
        if not lpn:
            continue
        try:
            q = float(f.get("CANTIDAD") or 0)
        except (TypeError, ValueError):
            q = 0.0
        por_lpn[lpn] = por_lpn.get(lpn, 0) + q
    return por_lpn


def por_codigo_de(filas_reserva):
    """{LPN: {articulo: pares}} de las paletas altas. El 36% trae mas de un articulo."""
    por = {}
    for f in filas_reserva:
        if not f.get("ES_ALTO"):
            continue
        lpn = str(f.get("LPN") or "").strip()
        art = str(f.get("PRODUCTO") or "").strip()
        if not lpn or not art:
            continue
        try:
            q = float(f.get("CANTIDAD") or 0)
        except (TypeError, ValueError):
            q = 0.0
        if q <= 0:
            continue
        por.setdefault(lpn, {})
        por[lpn][art] = por[lpn].get(art, 0) + q
    return {l: {a: int(round(q)) for a, q in c.items()} for l, c in por.items()}


def foto_reserva(filas_reserva):
    """
    La foto de las paletas altas al arrancar el turno, para medir la bajada.

    LLEVA `porCodigo` DESDE EL 22-ago-2026, y esa falta costo un reporte entero. La
    separacion de mercaderia se mide por ARTICULO, y como esta foto solo traia el total
    por paleta, la pantalla iba a buscar el arranque al area viva `analisis_sku_reserva`.
    Durante la noche funciona; a las 07:07 el robot de la maana la pisa, y la jornada
    cerrada quedaba restando la foto de la maana contra si misma. La noche del 21-ago
    dio 0 de 311 cuando se habia separado el 100%.
    """
    por_lpn = paletas_altas(filas_reserva)
    ahora = time.localtime()
    return {
        "fecha": time.strftime("%Y-%m-%d", ahora),
        "hora": time.strftime("%H:%M", ahora),
        "paletas": len(por_lpn),
        "pares": int(round(sum(por_lpn.values()))),
        "detalle": {l: int(round(q)) for l, q in por_lpn.items()},
        "porCodigo": por_codigo_de(filas_reserva),
    }


def foto_reserva_cierre(filas_reserva, fecha):
    """
    La misma foto, pero del CIERRE y ABIERTA POR CÓDIGO.

    El `porCodigo` no está de más: el 36% de las paletas trae más de un artículo —720
    de 1.981 la noche del 17-ago-2026— y la separación se mide por artículo. Con solo
    el total de la paleta habría que repartir a prorrata, y esa noche eso perdía 192
    unidades de 2.165.

    `fecha` llega de afuera porque esta foto pertenece a la jornada que TERMINÓ, y la
    corrida que la toma ya es del día siguiente.
    """
    por_lpn = paletas_altas(filas_reserva)
    return {
        "fecha": fecha,
        "hora": time.strftime("%H:%M", time.localtime()),
        "paletas": len(por_lpn),
        "pares": int(round(sum(por_lpn.values()))),
        "detalle": {l: int(round(q)) for l, q in por_lpn.items()},
        "porCodigo": por_codigo_de(filas_reserva),
    }


def jornada_que_termina():
    """
    La fecha de la noche que acaba de cerrar, vista desde la corrida de la mañana.

    La jornada arranca a las 19:00 y termina a las 06:30 del día siguiente, así que a
    las 07:00 la que cerró es la de AYER. Se usa el mismo corte de mediodía que
    `es_corrida_de_la_noche()` para que las dos cuenten la misma jornada.
    """
    return time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))


def datos_reserva_web(ruta):
    """
    El Stock Reserva con el MISMO formato que dejaba la carga a mano de la web
    (el mapeo vive en csvHub_v6.js), para que las pantallas que ya lo leen no
    noten el cambio. No se filtra por sucursal ni se sacan MERMA y DEV: la web
    guardaba todo y son sus consumidores los que filtran. Cambiar eso acá les
    movería los totales sin avisar.

    Fila 1 el título, la 2 vacía, la 3 los encabezados: los datos van desde la 4.
    La carga a mano arrancaba en la 3 y se llevaba la fila de encabezados como si
    fuera un dato; por eso el código que la lee descarta el PRODUCTO que dice
    "PRODUCTO". Acá ya no se manda, y ese descarte simplemente no encuentra nada.
    """
    filas = []
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for f in ws.iter_rows(min_row=4, values_only=True):
        if len(f) < 11:
            continue
        ubi = _txt(f[4])
        producto = _txt(f[8])
        if not ubi and not producto:
            continue
        nivel = _txt(f[1]).upper()
        try:
            qty = float(f[10] or 0)
        except (ValueError, TypeError):
            qty = 0.0
        filas.append({
            "NIVEL": nivel,
            "ES_ALTO": ("ALTO" in nivel) or nivel == "A",
            "PRODUCTO": producto,
            "CANTIDAD": qty,
            "UBICACION": ubi,
            "UBI_KEY": re.sub(r"[^A-Z0-9]", "", ubi.upper()),
            "LPN": _lpn(f[5]),
            "DESCRIPCION": _txt(f[9]),
        })
    wb.close()
    return filas


def escribir_xlsx(ruta, columnas, filas):
    """
    Deja las filas en un .xlsx que se abre como cualquier Excel.

    El Stock Activo sale de Oracle en CSV con punto y coma, y al abrirlo en Excel se ve todo
    apretado en una sola columna, así que el asistente terminaba peleándose con el asistente
    de importación. Acá va ya en columnas. Se escribe en modo write_only porque son 31.000
    filas y así no se levanta la hoja entera en memoria.
    """
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Datos")
    ws.append(list(columnas))
    for f in filas:
        ws.append([f.get(c, "") for c in columnas])
    wb.save(ruta)
    wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# TABLA DE TALLAS
# ─────────────────────────────────────────────────────────────────────────────

# La MISMA regla que extractTalla() en csvHub_v6.js. Si las dos se separan volvemos al
# problema que esto viene a resolver, así que cualquier cambio va en los dos lados.
#
# Busca guion, un dígito del 1 al 9, guion y la talla al final: "...WEINBRENNER-1-38"
# da 38 y "...WEINBRENNER-1-M" da M. Acepta letras a propósito: hay ropa en tallas S,
# M, L y XL, y una regla que solo aceptara números las dejaría a todas sin talla.
_RE_TALLA = re.compile(r"-([1-9])-([A-Z0-9.ÁÉÍÓÚÑ]+)$", re.IGNORECASE)


def extraer_talla(desc):
    """La talla que hay al final de la descripción, o None si no se puede leer."""
    d = _txt(desc)
    if not d:
        return None
    m = _RE_TALLA.search(d)
    if m:
        return m.group(2).strip().upper()
    # Respaldo: el anteúltimo trozo es un dígito suelto ("... - 1 - 38")
    partes = d.split("-")
    if len(partes) >= 3:
        previo = partes[-2].strip()
        if len(previo) == 1 and "1" <= previo <= "9":
            return partes[-1].strip().upper()
    return None


def tallas_de(filas, campo_sku, campo_desc):
    """Recorre un stock y devuelve {SKU: talla} con las que se pudieron leer."""
    salida = {}
    for f in filas:
        sku = _txt(f.get(campo_sku))
        talla = extraer_talla(f.get(campo_desc))
        if sku and talla:
            salida[sku] = talla
    return salida


def bajar_area(area, fecha="MASTER"):
    """
    Lee un área de la plataforma. Devuelve lo que haya guardado, o None.

    Es la pareja de `subir_datos`: hace falta desde que el robot no solo publica sino
    que también necesita leer —el cierre de la jornada lee `plan_buffer` para saber
    qué códigos congelar—. Nunca lanza: si el servidor no contesta, quien llama sigue
    con lo que pueda y lo dice en el log.
    """
    url = "%s/%s?date=%s&t=%d" % (WEB_DATOS_API, area, fecha, int(time.time()))
    try:
        pedido = urllib.request.Request(url)
        if WEB_ENTORNO == "beta":
            pedido.add_header("X-Environment", "beta")
        with urllib.request.urlopen(pedido, timeout=300) as resp:
            cuerpo = json.loads(resp.read().decode("utf-8"))
        if isinstance(cuerpo, dict) and "data" in cuerpo:
            return cuerpo["data"]
        return cuerpo
    except Exception as e:
        log("No se pudo leer el área %s (%s: %s)" % (area, type(e).__name__, str(e)[:120]), "WARN")
    return None


def bajar_tabla_tallas():
    """
    La tabla que ya está publicada. Devuelve {} si no hay o si el servidor no contesta.

    Que devuelva vacío NO es lo mismo que perderla: quien llama solo AGREGA, así que en
    el peor caso se vuelve a subir lo que se leyó hoy y lo viejo se recupera mañana. Lo
    que nunca hay que hacer es reemplazar la tabla con lo de una sola corrida.
    """
    url = "%s/%s?t=%d" % (WEB_DATOS_API, AREA_TALLAS, int(time.time()))
    try:
        pedido = urllib.request.Request(url)
        if WEB_ENTORNO == "beta":
            pedido.add_header("X-Environment", "beta")
        with urllib.request.urlopen(pedido, timeout=300) as resp:
            cuerpo = json.loads(resp.read().decode("utf-8"))
        datos = cuerpo.get("data", cuerpo) if isinstance(cuerpo, dict) else cuerpo
        if isinstance(datos, list):
            return dict((_txt(d.get("SKU")), _txt(d.get("TALLA")))
                        for d in datos if isinstance(d, dict) and d.get("SKU"))
        if isinstance(datos, dict):
            return dict((str(k), str(v)) for k, v in datos.items())
    except Exception as e:
        log("No se pudo leer la tabla de tallas publicada (%s: %s), se sigue con lo de hoy"
            % (type(e).__name__, str(e)[:120]), "WARN")
    return {}


def publicar_tallas(act_web, res_web, fecha, carpeta_tmp):
    """
    Arma la tabla acumulada, la publica y la deja en Descargas SOLO si cambió.

    Devuelve (todo_ok, hubo_cambios).
    """
    if not WEB_SUBIR_TALLAS:
        log("Tabla de tallas desactivada (WEB_SUBIR_TALLAS = False), se omite")
        return True, False

    leidas = {}
    leidas.update(tallas_de(act_web, "Artículo", "Descripción de artículo"))
    leidas.update(tallas_de(res_web, "PRODUCTO", "DESCRIPCION"))
    if not leidas:
        log("No se pudo leer ninguna talla de los stocks de hoy", "WARN")
        return False, False

    previa = bajar_tabla_tallas()
    nuevas = dict((k, v) for k, v in leidas.items() if k not in previa)
    cambiadas = dict((k, v) for k, v in leidas.items() if k in previa and previa[k] != v)

    # LAS CORREGIDAS A MANO MANDAN. Si un SKU ya tenía talla y hoy se lee otra, casi
    # siempre es porque alguien la arregló en el servidor —la descripción del WMS no
    # cambia sola—. Se avisa, pero no se pisa.
    if cambiadas:
        log("%d tallas distintas a las publicadas: se respeta lo que ya estaba"
            % len(cambiadas), "WARN")
        for k in list(cambiadas)[:5]:
            log("   %s: publicada '%s', leída hoy '%s'" % (k, previa[k], cambiadas[k]))

    acumulada = dict(previa)
    acumulada.update(nuevas)

    if not nuevas:
        log("Tabla de tallas sin novedades: %s SKU, no se vuelve a publicar"
            % format(len(acumulada), ",d"))
        return True, False

    log("Tabla de tallas: %s SKU nuevos, quedan %s en total"
        % (format(len(nuevas), ",d"), format(len(acumulada), ",d")))

    filas = [{"SKU": k, "TALLA": acumulada[k]} for k in sorted(acumulada)]
    ok = subir_datos(AREA_TALLAS, filas)

    # A Descargas SOLO cuando hubo novedades — es lo que pidió Daniel: si hoy no cambió
    # nada, no tiene sentido bajarse otra vez la misma planilla.
    try:
        ruta_x = os.path.join(carpeta_tmp, "Tabla de Tallas %s.xlsx" % fecha)
        escribir_xlsx(ruta_x, COLS_TALLAS, filas)
        ok = subir_a_la_web(ruta_x, fecha, tipo=TIPO_TALLAS,
                            guardar=TALLAS_EN_DESCARGAS) and ok
    except Exception as e:
        log("No se pudo dejar la tabla de tallas en Descargas: %s: %s"
            % (type(e).__name__, str(e)[:200]), "ERROR")
        ok = False

    return ok, True


def subir_datos(area, filas, intentos=3, fecha=None):
    """
    Publica un área de datos en la plataforma. Devuelve si salió bien.

    Va aparte de subir_a_la_web porque no es lo mismo: aquello deja un ARCHIVO
    para bajar, esto deja DATOS que la aplicación consulta.
    """
    if not WEB_SUBIR_STOCKS:
        log("Publicación de stocks desactivada (WEB_SUBIR_STOCKS = False), se omite")
        return True
    if not filas:
        log("%s: no hay filas para publicar, se omite" % area, "WARN")
        return False

    cuerpo = json.dumps(filas, ensure_ascii=False).encode("utf-8")
    url = "%s/%s?date=%s" % (WEB_DATOS_API, area, fecha or "MASTER")
    mb = len(cuerpo) / (1024.0 * 1024.0)

    for intento in range(1, intentos + 1):
        try:
            pedido = urllib.request.Request(url, data=cuerpo, method="POST")
            pedido.add_header("Content-Type", "application/json")
            pedido.add_header('X-Robot-Token', ROBOT_TOKEN)
            if WEB_ENTORNO == "beta":
                pedido.add_header("X-Environment", "beta")

            # El servidor puede estar dormido y tardar casi un minuto en despertar
            with urllib.request.urlopen(pedido, timeout=300) as resp:
                respuesta = json.loads(resp.read().decode("utf-8"))

            if respuesta.get("status") in (None, "success"):
                log("%s publicado: %s filas, %.2f MB" % (area, format(len(filas), ",d"), mb))
                return True
            raise RuntimeError(respuesta.get("message", "respuesta inesperada del servidor"))

        except Exception as e:
            detalle = "%s: %s" % (type(e).__name__, str(e)[:200])
            if intento < intentos:
                log("Intento %d: no se pudo publicar %s (%s), se reintenta..."
                    % (intento, area, detalle), "WARN")
                time.sleep(20)
            else:
                log("No se pudo publicar %s: %s" % (area, detalle), "ERROR")
    return False


def publicar(origen, destino):
    """
    Deja el archivo terminado en OneDrive. Reintenta por si está abierto.

    SE COPIA, NO SE MUEVE. Mover un archivo a una carpeta de OneDrive, estando los dos
    en el disco C, es un simple cambio de nombre para Windows: el archivo aparece en la
    carpeta sin que OneDrive lo vea nacer, y queda con la ficha de sincronización rota.
    Se ve así: flechitas girando para siempre, nunca llega a la nube, y después ni
    PowerShell lo puede leer —"acceso denegado"— aunque el Explorador sí.

    Del 05 al 07-ago-2026 los TRES Slotting que dejó el servidor quedaron atrapados así,
    tres de tres. Los dos stocks, que Playwright escribe directo en la carpeta con
    save_as(), no falló ni uno. Esa era toda la diferencia.

    copy2 crea el archivo de nuevo, byte a byte, que es como llegan los stocks, y
    conserva la fecha de modificación para que siga valiendo como la foto de su corrida.
    La copia temporal no se borra acá: de eso se encarga quien llama, que elimina la
    carpeta tmp entera cuando esto sale bien.
    """
    for intento in range(1, 4):
        try:
            if os.path.exists(destino):
                os.remove(destino)
            shutil.copy2(origen, destino)
            return True
        except OSError as e:
            if intento < 3:
                log("No se pudo escribir en OneDrive (%s), se reintenta..."
                    % str(e)[:120], "WARN")
                time.sleep(10)
            else:
                log("No se pudo dejar el archivo en OneDrive: %s" % str(e)[:200], "ERROR")
                log("Puede que esté abierto en Excel. El archivo quedó en %s" % origen, "WARN")
    return False


# ─────────────────────────────── Principal ───────────────────────────────

def buscar_archivo(carpeta, patron, fecha):
    """
    El archivo de esa fecha; si hay varios, EL MÁS RECIENTE. Si no hay ninguno de ese
    día, devuelve el último que haya y avisa que no es el pedido.

    Desde que el robot corre dos veces —06:00 y 19:00— el nombre lleva la hora
    ("Stock Activo 06-08-26 1900.csv") y hay dos archivos por día. El Slotting tiene que
    armarse con el último, que es el que refleja el almacén ahora.

    Los archivos viejos no llevan hora: por eso se busca POR FECHA y no por nombre
    exacto, así los de antes y los de ahora se encuentran igual.
    """
    if not os.path.isdir(carpeta):
        return None, False
    ext = os.path.splitext(patron)[1].lower()
    base = os.path.splitext(patron % fecha)[0]        # "Stock Activo 06-08-26"
    delDia = [os.path.join(carpeta, f) for f in os.listdir(carpeta)
              if f.lower().endswith(ext) and os.path.splitext(f)[0].startswith(base)]
    if delDia:
        # Por nombre y no por fecha de modificación: copiar la carpeta o resincronizar
        # OneDrive le cambia el mtime a todos y elegiría cualquiera. En el nombre, la
        # hora va al final, así que el mayor alfabético es el más tarde.
        return max(delDia), True
    candidatos = [os.path.join(carpeta, f) for f in os.listdir(carpeta) if f.lower().endswith(ext)]
    if not candidatos:
        return None, False
    return max(candidatos, key=os.path.getmtime), False


def edad_horas(ruta):
    """Cuántas horas hace que se escribió el archivo. Infinito si no se puede saber."""
    try:
        return (time.time() - os.path.getmtime(ruta)) / 3600.0
    except Exception:
        return float("inf")


def run(fecha=None, ruta_act_dada=None, ruta_res_dada=None, igualmente=False):
    global _log_file
    os.makedirs(LOG_DIR, exist_ok=True)
    _log_file = os.path.join(LOG_DIR, "slotting_%s.log" % datetime.now().strftime("%Y-%m-%d_%H%M%S"))
    inicio = time.time()

    fecha = fecha or datetime.now().strftime("%d-%m-%y")
    log("=" * 58)
    log("REPORTE SLOTTING - %s" % fecha)
    log("=" * 58)

    # Si el robot dijo qué archivo bajó, se usa ese y no se busca nada.
    if ruta_act_dada:
        ruta_act, exacto_a = ruta_act_dada, True
        log("Stock Activo: el que acaba de bajar el robot (%s)" % os.path.basename(ruta_act))
    else:
        ruta_act, exacto_a = buscar_archivo(DIR_ACTIVO, "Stock Activo %s.csv", fecha)
    if ruta_res_dada:
        ruta_res, exacto_r = ruta_res_dada, True
        log("Stock Reserva: el que acaba de bajar el robot (%s)" % os.path.basename(ruta_res))
    else:
        ruta_res, exacto_r = buscar_archivo(DIR_RESERVA, "Stock Reserva %s.xlsx", fecha)

    faltan = []
    if not ruta_act:
        faltan.append("Stock Activo")
    if not ruta_res:
        faltan.append("Stock Reserva")
    if not os.path.exists(MAESTRO):
        faltan.append("Maestro_Articulos.xlsx en la carpeta Archivos")
    if not os.path.exists(MARCAS):
        faltan.append("Marcas.xlsx en la carpeta Archivos")
    if faltan:
        for f in faltan:
            log("Falta: %s" % f, "ERROR")
        # DÓNDE ESTUVO BUSCANDO. Sin esto, "Falta: Stock Activo" no distingue entre que el
        # archivo no esté y que el robot esté mirando la carpeta equivocada, que fue lo que
        # pasó el 05-ago al copiar al servidor la versión de la laptop.
        log("Se buscó en: %s" % BASE, "ERROR")
        if not os.path.isdir(BASE):
            log("Esa carpeta NO EXISTE en esta máquina (%s, usuario %s)"
                % (os.environ.get("COMPUTERNAME", "?"), os.environ.get("USERNAME", "?")), "ERROR")
        return 1

    if not exacto_a:
        log("No hay Stock Activo del %s, se usa %s" % (fecha, os.path.basename(ruta_act)), "WARN")
    if not exacto_r:
        log("No hay Stock Reserva del %s, se usa %s" % (fecha, os.path.basename(ruta_res)), "WARN")

    # EL SEGURO: un stock viejo no se publica.
    #
    # Se mira SIEMPRE, incluso cuando el archivo lo pasó el robot: si su descarga se colgó
    # a medias y dejó el de la corrida anterior, la edad lo delata igual.
    viejos = []
    for etiqueta, ruta in (("Stock Activo", ruta_act), ("Stock Reserva", ruta_res)):
        h = edad_horas(ruta)
        log("%-14s %s  (escrito hace %.1f h)" % (etiqueta, os.path.basename(ruta), h))
        if h > MAX_HORAS_STOCK:
            viejos.append((etiqueta, os.path.basename(ruta), h))

    if viejos and not igualmente:
        log("=" * 58, "ERROR")
        log("NO SE PUBLICA: el stock no es de esta corrida", "ERROR")
        for etiqueta, nombre, h in viejos:
            log("   %s -> %s tiene %.1f horas (el limite son %d)"
                % (etiqueta, nombre, h, MAX_HORAS_STOCK), "ERROR")
        log("Publicarlo dejaria a la web con una foto vencida y con la hora de ahora,", "ERROR")
        log("que es lo que paso el 06-ago-2026: la ola de la noche se corrio con el", "ERROR")
        log("stock de la manana y genero tareas sobre mercaderia ya almacenada.", "ERROR")
        log("Si de verdad quiere reprocesar un archivo viejo: --igualmente", "ERROR")
        log("=" * 58, "ERROR")
        return 4
    if viejos:
        log("Se publica un stock viejo porque se pidio --igualmente", "WARN")

    log("Leyendo Stock Activo...")
    filas = leer_activo(ruta_act)
    n_act = len(filas)
    log("Leyendo Stock Reserva...")
    filas += leer_reserva(ruta_res)
    log("Filas leídas: %s (Activo %s + Reserva %s)" %
        (format(len(filas), ",d"), format(n_act, ",d"), format(len(filas) - n_act, ",d")))

    log("Leyendo Maestro de Artículos y tabla de Marcas...")
    maestro = leer_maestro(MAESTRO)
    marcas = leer_marcas(MARCAS)
    log("Maestro: %s artículos   Marcas: %s equivalencias" %
        (format(len(maestro), ",d"), format(len(marcas), ",d")))

    log("Aplicando reglas...")
    datos, fuera, sin_maestro = construir(filas, maestro, marcas)

    log("Filas del reporte: %s" % format(len(datos), ",d"))
    for k, v in sorted(fuera.items(), key=lambda x: -x[1]):
        if v:
            log("   fuera por %-16s %s" % (k, format(v, ",d")))
    if sin_maestro:
        log("Artículos sin Maestro: %s (van con guión)" % format(len(sin_maestro), ",d"), "WARN")

    buf = sum(f[8] for f in datos)
    zona = sum(f[9] for f in datos)
    res = sum(f[10] for f in datos)
    log("-" * 46)
    log("Qty Buffer  : %14s" % format(int(buf), ",d"))
    log("Qty Zona    : %14s" % format(int(zona), ",d"))
    log("Qty Reserva : %14s" % format(int(res), ",d"))
    log("TOTAL       : %14s" % format(int(buf + zona + res), ",d"))
    log("-" * 46)

    if not datos:
        log("No quedaron filas para el reporte", "ERROR")
        return 1

    os.makedirs(DIR_SALIDA, exist_ok=True)
    salida = os.path.join(DIR_SALIDA, "Slotting %s.xlsx" % fecha)

    # El reporte se arma en una carpeta local y recién al final se mueve a
    # OneDrive. Que Excel trabaje en vivo sobre OneDrive es la causa habitual de
    # que se corte la conexión a mitad del armado, y además así OneDrive no
    # sincroniza un archivo a medio hacer.
    carpeta_tmp = tempfile.mkdtemp(prefix="slotting_")
    temporal = os.path.join(carpeta_tmp, "Slotting %s.xlsx" % fecha)

    log("Escribiendo la tabla de datos...")
    escribir_datos(temporal, datos)

    log("Creando la tabla dinámica con Excel (en segundo plano)...")
    dinamica_ok = crear_dinamica(temporal, len(datos))
    if not dinamica_ok:
        log("El archivo igual queda con la hoja Datos, se puede armar la dinámica a mano", "WARN")

    log("Publicando el archivo en OneDrive...")
    publicado = publicar(temporal, salida)
    final = salida if publicado else temporal
    if publicado:
        shutil.rmtree(carpeta_tmp, ignore_errors=True)

    log("Subiendo el archivo a la web...")
    subido = subir_a_la_web(final, fecha, tipo="Slotting")

    # Los stocks van DESPUÉS del reporte y a propósito: si algo falla acá, el
    # Slotting del día ya quedó publicado y el asistente no se queda sin su archivo.
    stocks_ok = True
    if WEB_SUBIR_STOCKS:
        log("Publicando los stocks para toda la red...")
        act_web = datos_activo_web(ruta_act)
        res_web = datos_reserva_web(ruta_res)
        stocks_ok = subir_datos(AREA_ACTIVO, act_web) and stocks_ok
        stocks_ok = subir_datos(AREA_RESERVA, res_web) and stocks_ok
        if not stocks_ok:
            log("Los stocks no se publicaron: la web va a seguir usando los anteriores", "WARN")

        # Y las MISMAS filas al cajón de la hora, para que el turno arranque con los dos
        # relojes en cero. No se baja nada de nuevo: son las que ya están en memoria.
        #
        # A propósito NO entran en `stocks_ok`: si esto falla, el robot de la hora lo
        # arregla solo en la próxima corrida y no tiene sentido marcar en rojo la corrida
        # del turno por un reporte de avance que se repone en 60 minutos.
        log("Dejando el cajón de la hora en sincronía con el arranque...")
        if not (subir_datos(AREA_ACTIVO_HORA, act_web) and subir_datos(AREA_RESERVA_HORA, res_web)):
            log("El cajón de la hora no se pudo sincronizar; lo repone stock_por_hora.py "
                "en la próxima corrida", "WARN")

        # SOLO EN LA CORRIDA DE LA NOCHE: es la foto con la que arranca el turno, y es
        # contra ella que se mide después cuánto se limpió del Buffer C. La de la mañana
        # ya es el resultado, no el punto de partida, y guardarla borraría la buena.
        if es_corrida_de_la_noche():
            foto = foto_buffer_c(act_web)
            log("Guardando el Buffer C del arranque: %s pares en %s artículos"
                % (foto["pares"], foto["articulos"]))
            # CON FECHA, no con MASTER: si se pisara, al dia siguiente no habria con
            # que comparar y el reporte del turno se quedaria sin historico. Son 2,7 KB
            # por noche, unos 3 MB al ano.
            if not subir_datos(AREA_BUFFER_C, foto, fecha=foto["fecha"]):
                log("El Buffer C del arranque no se guardó: mañana el reporte del turno "
                    "no va a tener meta para la limpieza", "WARN")

            # Y las paletas altas, para poder medir la bajada durante la noche.
            pal = foto_reserva(res_web)
            log("Guardando las paletas del arranque: %s paletas altas, %s pares, "
                "%s paletas abiertas por codigo"
                % (format(pal["paletas"], ",d"), format(pal["pares"], ",d"),
                   format(len(pal.get("porCodigo") or {}), ",d")))
            if not subir_datos(AREA_RESERVA_ARRANQUE, pal, fecha=pal["fecha"]):
                log("Las paletas del arranque no se guardaron: el reporte del turno no va "
                    "a poder medir la bajada de paletas", "WARN")
        else:
            # LA CORRIDA DE LA MAÑANA CIERRA LA NOCHE. No toca las fotos del arranque
            # —esas son el punto de partida y pisarlas dejaría al turno sin con qué
            # comparar— pero sí guarda las del CIERRE, que es lo que le faltaba al
            # reporte para que una jornada terminada siga midiéndose después de las
            # 06:30. Van con la fecha de la jornada que cerró, no la de hoy.
            log("Corrida de la mañana: las fotos del arranque no se tocan")
            jornada = jornada_que_termina()

            cierre_c = foto_buffer_c(act_web)
            cierre_c["fecha"] = jornada
            log("Guardando el Buffer C del cierre de la jornada %s: %s pares en %s artículos"
                % (jornada, cierre_c["pares"], cierre_c["articulos"]))
            if not subir_datos(AREA_BUFFER_C_CIERRE, cierre_c, fecha=jornada):
                log("El Buffer C del cierre no se guardó: la limpieza de esa noche se "
                    "queda con el último avance que alcanzó a medir la pantalla", "WARN")

            cierre_r = foto_reserva_cierre(res_web, jornada)
            log("Guardando las paletas del cierre: %s paletas altas, %s pares"
                % (format(cierre_r["paletas"], ",d"), format(cierre_r["pares"], ",d")))
            if not subir_datos(AREA_RESERVA_CIERRE, cierre_r, fecha=jornada):
                log("Las paletas del cierre no se guardaron: la bajada de paletas y la "
                    "separación de esa noche quedan sin avance medido", "WARN")

        # Y los mismos datos como Excel descargable. Se arman en la carpeta temporal y se
        # borran al terminar: en OneDrive ya están los originales de Oracle.
        # De mañana no se suben: subir_a_la_web() lo corta, y acá igual se arman porque
        # de paso se actualiza la tabla de tallas, que sí va todos los días.
        log("Dejando los stocks en Descargas...")
        tmp_x = tempfile.mkdtemp(prefix="stocks_")
        try:
            for nombre, cols, datos_x in (
                    ("Stock Activo", COLS_ACTIVO, act_web),
                    ("Stock Reserva", COLS_RESERVA, res_web)):
                ruta_x = os.path.join(tmp_x, "%s %s.xlsx" % (nombre, fecha))
                escribir_xlsx(ruta_x, cols, datos_x)
                mb = os.path.getsize(ruta_x) / (1024.0 * 1024.0)
                log("   %s: %s filas, %.2f MB" % (nombre, format(len(datos_x), ",d"), mb))
                stocks_ok = subir_a_la_web(ruta_x, fecha, tipo=nombre) and stocks_ok

            # LA TABLA DE TALLAS, con los dos stocks ya leídos. Va acá adentro a
            # propósito: necesita act_web y res_web, y reutiliza la carpeta temporal.
            log("Actualizando la tabla de tallas...")
            tallas_ok, hubo = publicar_tallas(act_web, res_web, fecha, tmp_x)
            stocks_ok = tallas_ok and stocks_ok
        except Exception as e:
            log("No se pudieron dejar los stocks en Descargas: %s: %s"
                % (type(e).__name__, str(e)[:200]), "ERROR")
            stocks_ok = False
        finally:
            shutil.rmtree(tmp_x, ignore_errors=True)

    # ── EVOLUCIÓN DEL ARTÍCULO ────────────────────────────────────────────────
    # Va al final y a propósito NO cambia el código de salida: es un reporte de
    # estudio, no algo que el turno necesite. Si falla, la pantalla sigue mostrando
    # el estudio de ayer con su fecha a la vista, así que se nota sin romper nada.
    #
    # Tiene que correr DESPUÉS de bajar los stocks: lee la foto de hoy recién dejada
    # en OneDrive y la agrega a su acumulado. La primera corrida tarda unos dos
    # minutos porque relee las 76 fotos; las siguientes son segundos.
    try:
        log("Actualizando la evolución del artículo...")
        import generar_evolucion
        if generar_evolucion.main(solo_calcular=False, log_externo=log) == 0:
            log("Evolución del artículo publicada")
        else:
            log("La evolución del artículo no se pudo publicar; el reporte se queda "
                "con el estudio anterior", "WARN")
    except Exception as e:
        log("No se pudo actualizar la evolución del artículo: %s: %s"
            % (type(e).__name__, str(e)[:200]), "WARN")

    # ── ROTACIÓN Y PERMANENCIA ────────────────────────────────────────────────
    # El FSN del almacén más el aging. Va acá por lo mismo que la evolución: necesita
    # la foto de hoy recién dejada en OneDrive, y NO cambia el código de salida —es un
    # reporte de estudio, no algo que el turno necesite—. Si falla, la pantalla sigue
    # mostrando el estudio de ayer con su fecha a la vista.
    #
    # Tarda unos minutos: relee las fotos de la ventana de 3 meses y, para saber desde
    # cuándo está cada artículo, también las anteriores.
    try:
        log("Actualizando rotación y permanencia...")
        import generar_rotacion
        if generar_rotacion.main(solo_calcular=False, log_externo=log) == 0:
            log("Rotación y permanencia publicada")
        else:
            log("Rotación y permanencia no se pudo publicar; el reporte se queda con "
                "el estudio anterior", "WARN")
    except Exception as e:
        log("No se pudo actualizar rotación y permanencia: %s: %s"
            % (type(e).__name__, str(e)[:200]), "WARN")

    # ── LA FOTO DEL DIA DE LA RESERVA ─────────────────────────────────────────
    # Va al final y DESPUES de que los stocks quedaron publicados: la foto se arma con
    # el stock de reserva que hay en el SERVIDOR, asi que corriendo antes guardaria la
    # de anoche con la fecha de hoy.
    #
    # Antes esto lo hacia el navegador, y solo cuando alguien entraba a Analisis
    # Reserva: el dia que nadie abriera esa pantalla quedaba un agujero en el
    # calendario, y no se recuperaba porque el stock ya lo habia pisado el del dia
    # siguiente. Daniel, 22-ago-2026: *"que el robot guarde la foto al terminar el
    # ancla, se tiene que guardar eso"*.
    #
    # NO cambia el codigo de salida, igual que la evolucion y la rotacion: es un
    # historico de estudio, no algo que el turno necesite esta noche. Y no pisa una
    # foto que ya este.
    try:
        log("Guardando la foto del dia de la reserva...")
        import subprocess
        r = subprocess.run([sys.executable, os.path.join(os.path.dirname(os.path.abspath(__file__)), "foto_reserva.py")],
                           timeout=900)
        log("Foto de reserva: %s" % ("guardada" if r.returncode == 0
            else "no se pudo, ver logs/foto_reserva.log"),
            "INFO" if r.returncode == 0 else "WARN")
    except Exception as e:
        log("No se pudo guardar la foto de reserva: %s: %s"
            % (type(e).__name__, str(e)[:200]), "WARN")

    mb = os.path.getsize(final) / (1024.0 * 1024.0)
    log("=" * 58)
    log("LISTO en %.1f minutos - %.2f MB" % ((time.time() - inicio) / 60.0, mb))
    log(final)
    log("=" * 58)

    # Código 3: los datos están, pero algo del reporte no quedó completo. Así el
    # Programador de tareas marca la corrida en rojo en vez de darla por buena.
    # Los stocks cuentan igual que el resto: si no llegaron a la nube, la web se
    # queda con los del día anterior y nadie se entera hasta que un papel sale mal.
    return 0 if (dinamica_ok and publicado and subido and stocks_ok) else 3


if __name__ == "__main__":
    # Se acepta como antes la fecha suelta ("07-08-26"), y además los archivos exactos
    # que el robot acaba de bajar. Así el generador no tiene que adivinar cuál es el bueno.
    _fecha = None
    _act = _res = None
    _igualmente = False
    for _arg in sys.argv[1:]:
        if _arg.startswith("--activo="):
            _act = _arg.split("=", 1)[1]
        elif _arg.startswith("--reserva="):
            _res = _arg.split("=", 1)[1]
        elif _arg == "--igualmente":
            _igualmente = True
        elif not _arg.startswith("--"):
            _fecha = _arg
    try:
        sys.exit(run(_fecha, _act, _res, _igualmente))
    except Exception as e:
        log("Error no controlado: %s: %s" % (type(e).__name__, str(e)[:300]), "ERROR")
        sys.exit(1)
