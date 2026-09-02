# -*- coding: utf-8 -*-
"""
Validación previa del reporte Slotting.
Aplica las reglas extraídas del Power Query y muestra los totales, para
compararlos contra el Excel que se arma a mano antes de construir nada.
"""
import csv, io, os, re, sys
from collections import Counter, defaultdict
import openpyxl

BASE = r"C:\Users\Administrator\OneDrive\danielames.bata\scraping Stock"
FECHA = "30-07-26"
ACT = os.path.join(BASE, "Stock Activo", "Stock Activo %s.csv" % FECHA)
RES = os.path.join(BASE, "Stock Reserva", "Stock Reserva %s.xlsx" % FECHA)

NIVELES_ZONA = {"AND", "MZN01", "MZN02", "MZN03", "MZN04", "PARED", "SEL"}
SKU_VALIDO = re.compile(r"^\d{7}-\d-\d+$")

ZONA_IND = ("MZN03-01", "MZN03-02", "MZN03-03", "MZN03-07")
ZONA_MC = ("MZN03-04", "MZN03-05", "MZN03-06")


def zona_de(ubi):
    if ubi.startswith(ZONA_IND):
        return "Zona Industrial"
    if ubi.startswith(ZONA_MC):
        return "Zona Marie Claire"
    return None


filas = []
desc = Counter()

# ── Stock Activo ────────────────────────────────────────────────
with io.open(ACT, encoding="utf-8-sig", errors="replace") as fh:
    r = csv.reader(fh, delimiter=";")
    next(r)
    for row in r:
        if len(row) < 5 or not row[1]:
            continue
        nivel, sku, descripcion, ubi = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()
        try:
            qty = float(row[4] or 0)
        except ValueError:
            qty = 0
        z = zona_de(ubi)
        if z:
            ubi = "%s - %s" % (z, ubi)
        filas.append({"origen": "ACT", "NIVEL": nivel, "SKU": sku, "UBICACION": ubi, "DESC": descripcion, "QTY": qty})

# ── Stock Reserva ───────────────────────────────────────────────
wb = openpyxl.load_workbook(RES, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
for f in ws.iter_rows(min_row=4, values_only=True):
    if f[0] is None or str(f[0]).strip() != "50008":
        if f[0] != 50008:
            continue
    if f[7] is None:
        continue
    art = str(f[7]).strip()
    prod = "" if f[8] is None else str(f[8]).strip()
    sku = prod if prod else art
    try:
        qty = float(f[10] or 0)
    except (ValueError, TypeError):
        qty = 0
    filas.append({"origen": "RES", "NIVEL": str(f[1] or "").strip(), "SKU": sku,
                  "UBICACION": str(f[4] or "").strip(), "DESC": str(f[9] or "").strip(), "QTY": qty})
wb.close()

print("Filas leídas          : %s  (Activo %s + Reserva %s)" % (
    format(len(filas), ",d"),
    format(sum(1 for x in filas if x["origen"] == "ACT"), ",d"),
    format(sum(1 for x in filas if x["origen"] == "RES"), ",d")))

# ── Exclusiones ─────────────────────────────────────────────────
fuera = Counter()
limpias = []
for x in filas:
    if x["UBICACION"].upper().startswith("CDBUFFER-C"):
        fuera["ubicación CDBUFFER-C"] += 1
        continue
    if not SKU_VALIDO.match(x["SKU"]):
        fuera["SKU inválido (roto o sin talla)"] += 1
        continue
    limpias.append(x)

print()
print("Excluidas:")
for k, v in fuera.most_common():
    print("   %-34s %s" % (k, format(v, ",d")))
print("Filas que quedan      : %s" % format(len(limpias), ",d"))

# ── Reparto en las tres cantidades ──────────────────────────────
buf = zona = res = 0.0
nb = nz = nr = 0
sin_clasificar = Counter()

for x in limpias:
    n, u, q = x["NIVEL"], x["UBICACION"], x["QTY"]
    if n == "CDBUFFER":
        buf += q; nb += 1
    elif n in NIVELES_ZONA:
        zona += q; nz += 1
    elif n == "ALTO":
        if u.upper().startswith("SEL-14"):
            sin_clasificar["ALTO en SEL-14 (excluido)"] += 1
        else:
            res += q; nr += 1
    else:
        sin_clasificar[n or "(vacío)"] += 1

print()
print("=" * 52)
print("TOTALES PARA COMPARAR CONTRA TU EXCEL")
print("=" * 52)
print("   Qty Buffer  : %14s   (%s filas)" % (format(int(buf), ",d"), format(nb, ",d")))
print("   Qty Zona    : %14s   (%s filas)" % (format(int(zona), ",d"), format(nz, ",d")))
print("   Qty Reserva : %14s   (%s filas)" % (format(int(res), ",d"), format(nr, ",d")))
print("   " + "-" * 44)
print("   TOTAL       : %14s" % format(int(buf + zona + res), ",d"))
print()
print("Filas descartadas por NIVEL no usado:")
for k, v in sin_clasificar.most_common(12):
    print("   %-30s %s" % (k, format(v, ",d")))
