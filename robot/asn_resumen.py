# -*- coding: utf-8 -*-
"""
EL RESUMEN DEL ASN PARA LA PANTALLA DE RECEPCION.

Lee los seis archivos que dejo `asn_web_report.py` -uno por mes, 64 MB en
total- y publica a la plataforma solo lo que la pantalla necesita: unos pocos
KB. Daniel no abre los Excel; abre la web.

QUE PUBLICA (area `asn_recepcion`)
    · enviado contra recibido, con el calzado separado del resto
    · el cumplimiento mes a mes
    · el estado de los ASN, con los cancelados
    · si llego lo que dijo el ASN: completo, parcial, sin recibir, de mas
    · la lista de los ASN parciales, que son los que hay que perseguir

COMO SEPARA EL CALZADO
    Cruzando con el Maestro por la columna `G. Gender`: Footwear es calzado y
    todo lo demas no. No alcanza con "estar en el Maestro": ahi adentro tambien
    hay Non Footwear, Non Commercial y Promotions. La `Caja H30`, que es el
    articulo con mas pendiente de todos, esta en el Maestro marcada como
    Non Commercial.

CORRE DETRAS DEL ROBOT DEL ASN, a las 04:30.
"""
import os
import re
import sys
import json
import time
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

API = "https://logistics-backend-wv0x.onrender.com/api/logistics"
AREA = "asn_recepcion"
# Cuantas filas de detalle se publican. Medido: 500 articulos son 89 KB y todos
# los 11.307 serian 1.117 KB. Ver el comentario del paquete.
TOPE_ARTICULOS = 500
TOPE_POR_MES = 50

# CUANDO LLEGA. Daniel eligio el 03-sep-2026: los dias cercanos COMPLETOS -que son
# sobre los que se prepara el almacen- y solo los mas grandes de los lejanos.
DIAS_CERCA = 14
TOPE_DIA_LEJOS = 25
# La credencial del robot vive en el entorno del Contabo y en Render, NUNCA en el repo.
ROBOT_TOKEN = os.environ.get("ROBOT_TOKEN", "")

CARPETA_ASN = None          # se resuelve con wms_automation_final
MAESTRO = None
# La marca por codigo, para la tabla. Se llena al leer el Maestro en construir().
MARCA_DE = {}

_LOG_EXTERNO = None


def log(mensaje, nivel="INFO"):
    if _LOG_EXTERNO:
        _LOG_EXTERNO(mensaje, nivel)
        return
    # flush=True: corriendo como tarea programada la salida va a un archivo, y
    # sin esto Python la guarda en memoria hasta el final. El log se ve vacio los
    # tres minutos que dura la corrida y parece que el robot no arranco.
    print("[%s] [%-5s] %s" % (datetime.now().strftime("%H:%M:%S"), nivel, mensaje),
          flush=True)


# ── DE DONDE VIENE CADA ASN ─────────────────────────────────────────────────
#
# Las etiquetas cortas van en el orden en que Daniel las nombra.
TIPOS = ('importacion', 'nacional', 'inversa', 'devolucion', 'reingreso',
         'traslado', 'materiales', 'sin_clasificar')

# EL TIPO VIENE CODIFICADO EN `cust_field_1`, no como texto.
#
# Comprobado el 03-sep-2026 sobre los 384 ASN de septiembre: el codigo coincide
# al 100% con el tipo que deduce el patron del numero, y la moneda de
# `cust_field_2` lo refuerza sola.
#
#     codigo  moneda   tipo             ASN   coincidencia con el numero
#       24     USD     importacion      161      161/161  = 100%
#       23     PEN     nacional          18       18/18   = 100%
#       56/89   -      inversa          191      191/191  = 100%
#       30      -      devolucion        14       14/14   = 100%
#
# 56 y 89 son los dos codigos de la inversa; los dos caen en ASN que empiezan
# con T, que es lo que Daniel confirmo como logistica inversa. Que los separa
# todavia no se sabe, asi que los dos van al mismo tipo y no se inventa nada.
# La tabla completa, medida sobre los 16.404 ASN de los seis meses el 03-sep-2026.
# Septiembre solo mostraba 5 de los 11 codigos.
#
#   codigo   ASN     que es                     como se comprobo
#     24    6.000    importacion                100% del patron del numero
#     56    3.917    inversa                    100%
#     30    2.557    devolucion                  97%
#     89    2.027    inversa                    100%
#     23    1.584    nacional                   100%
#     16       76    inversa                    100%
#     36       12    reingreso  "cambio de calidad"
#     63        3    reingreso  "cambio de calidad a 2da"
#     79        2    reingreso  "devolucion por acuerdo comercial"
#     A4        1    reingreso  "reingreso de ordenes Falabella"
#     10       98    NO SE MAPEA: se contradice solo -61% nacional, 39%
#                    importacion segun el numero-. Cae al patron, que acierta
#                    el 97,4% de las veces.
# CUANTO TARDA LO ANUNCIADO EN ENTRAR DE VERDAD, en dias.
#
# Daniel, 03-sep-2026: *"si la orden dice 142 mil y solo hay 25 mil, esta mandando
# un parcial. No es que de golpe te mande los 142.597, te va a estar mandando
# parciales"*. Asi que la fecha del ASN es la del ANUNCIO y la cantidad es el
# techo de la orden, no lo que baja del camion ese dia.
#
# Medido el 03-sep-2026 sobre 1,3 millones de lineas YA RECIBIDAS, con la fecha
# de recepcion que recien esa bajada trajo en los seis archivos. Es la MEDIANA:
# el promedio lo arrastran los ASN colgados de 500 dias.
#
#     tipo          lineas   mediana   el 25%   el 75%
#     inversa      521.238     15 d       8       29
#     nacional     424.583      4 d       0       14
#     importacion  226.073     10 d       6       16
#     reingreso     45.106     19 d
#     devolucion    12.248    155 d      66      155   <- otro mundo
#
# Solo el 8% entra ANTES de la fecha anunciada, asi que la fecha sirve de
# referencia y lo que llega, llega despues.
DEMORA_DIAS = {
    'importacion': 10, 'nacional': 4, 'inversa': 15, 'reingreso': 19,
    'devolucion': 155, 'traslado': 12, 'materiales': 12, 'sin_clasificar': 12,
}

POR_CODIGO = {
    '23': 'nacional',
    '24': 'importacion',
    '30': 'devolucion',
    '56': 'inversa',
    '89': 'inversa',
    '16': 'inversa',
    '36': 'reingreso',
    '63': 'reingreso',
    '79': 'reingreso',
    'A4': 'reingreso',
}

# Y por si algun dia el campo trae el texto en vez del codigo.
DEL_WMS = {
    'IMP': 'importacion', 'IMPORTACION': 'importacion', 'IMPORTACION IMP': 'importacion',
    'NAC': 'nacional', 'NACIONAL': 'nacional',
    'DEV': 'inversa', 'DEVOLUCION': 'inversa', 'LOGISTICA INVERSA': 'inversa',
}

_PRE = re.compile(r'^([A-Za-z]+)')

# EL NUMERO DE ASN ENCODEA LA ORDEN DE COMPRA Y LA SOCIEDAD:
#     20260533602BA.8817454  ->  orden 2026-05336-02, sociedad BA
# Se descubrio el 03-sep-2026 buscando el expediente que Daniel paso por correo.
ORDEN_EN_EL_ASN = re.compile(r'^(\d{4})(\d{5})(\d{2})([A-Z]{2})')


def _hora(v):
    """HH:MM:SS de la fecha de recepcion, venga como fecha de Excel o como texto."""
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M:%S")
    t = str(v or "").strip()
    return t[11:19] if len(t) >= 19 and t[10:11] == " " else ""


def _fecha10(v):
    """AAAA-MM-DD, venga como fecha de Excel o como texto."""
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    t = str(v or '').strip()[:10]
    return t if len(t) == 10 and t[4:5] == '-' else ''


def tipo_por_el_numero(asn):
    """El tipo deducido del numero, medido sobre 16.175 ASN el 03-sep-2026.

    Cubre el 100%: no hay ASN que no caiga en alguna de estas formas.
    """
    a = (asn or '').strip()
    if not a:
        return 'sin_clasificar'
    m = _PRE.match(a)
    if m:
        pre = m.group(1).upper()
        if pre == 'T':
            return 'inversa'
        if pre in ('B', 'F'):
            return 'devolucion'
        if pre in ('G', 'RA'):
            return 'traslado'
        if pre == 'OS':
            return 'materiales'
        return 'sin_clasificar'
    # empieza con el ano. El punto separa la importacion de lo nacional:
    # con punto son +37 dias de anticipacion y proveedor de 10 digitos; sin
    # punto se registra DESPUES de la fecha anunciada y llega completo.
    return 'importacion' if '.' in a else 'nacional'


def tipo_del_wms(valor):
    """Lo que diga el campo personalizado: primero el codigo, despues el texto."""
    v = (valor or '').strip().upper()
    if not v:
        return None
    if v in POR_CODIGO:
        return POR_CODIGO[v]
    v = v.replace('\u00d3', 'O').replace('\u00cd', 'I').replace('\u00c1', 'A')
    for k, t in DEL_WMS.items():
        if v == k or v.startswith(k + ' ') or v.endswith(' ' + k):
            return t
    return None


def limpio(texto):
    """Saca los bytes rotos que trae el WMS en algunas descripciones.

    `9906681-1-01` llega como "KIT DE AFILIACION CAT?LOGO": donde va la A con
    tilde hay un byte 0x81 suelto, que Python arrastra como sustituto sin
    pareja. En pantalla sale un rombo negro, y ademas ROMPE a cualquiera que
    vuelva a codificar el JSON -paso al releer el paquete ya publicado-.
    """
    if not texto:
        return ""
    if any(0xD800 <= ord(c) <= 0xDFFF for c in texto):
        return "".join("?" if 0xD800 <= ord(c) <= 0xDFFF else c for c in texto)
    return texto


def rutas():
    global CARPETA_ASN, MAESTRO
    import wms_automation_final as wms
    base = wms._base_onedrive()                     # ...\scraping Stock
    CARPETA_ASN = os.path.join(base, "ASN")
    MAESTRO = os.path.join(os.path.dirname(base), "Maestro_Articulos.xlsx")
    if not os.path.isfile(MAESTRO):
        alterno = os.path.join(base, "Archivos", "Maestro_Articulos.xlsx")
        if os.path.isfile(alterno):
            MAESTRO = alterno


def af(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def fecha_de(v):
    if isinstance(v, datetime):
        return v
    s = str(v or "").strip()
    for f in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    return None


def leer_maestro():
    """Los codigos que son calzado, y ademas la MARCA y el gender de cada uno.

    LA COLUMNA DE LA MARCA SE LLAMA `Marcas`, EN PLURAL. Buscarla como "Marca" no
    la encuentra y la pantalla sale con todo en "(sin marca)": paso en la primera
    medicion. Es la misma que usa el robot de picking.
    """
    from openpyxl import load_workbook
    wb = load_workbook(MAESTRO, read_only=True, data_only=True)
    h = wb[wb.sheetnames[0]]
    it = h.iter_rows(values_only=True)
    enc = [str(c).strip() if c is not None else "" for c in next(it)]

    def col(*nombres):
        for n in nombres:
            for i, c in enumerate(enc):
                if c.lower().replace(".", "").replace(" ", "") == n.lower().replace(".", "").replace(" ", ""):
                    return i
        return None

    iC, iG = enc.index("CodArticulo"), enc.index("G. Gender")
    iM = col("Marcas", "MarcaStd")
    calzado = set()
    ficha = {}
    total = 0
    for f in it:
        if iC >= len(f) or f[iC] is None:
            continue
        total += 1
        cod = str(f[iC]).strip()
        gen = str(f[iG]).strip() if iG < len(f) and f[iG] else ""
        if gen == "Footwear":
            calzado.add(cod.lstrip("0"))
        ficha[cod[:7]] = (gen or "(sin gender)",
                          (str(f[iM]).strip() if iM is not None and iM < len(f) and f[iM] else "")
                          or "(sin marca)")
    wb.close()
    log("Maestro: %s articulos, %s de calzado, marca en la columna %s"
        % ("{:,}".format(total), "{:,}".format(len(calzado)), enc[iM] if iM is not None else "?"))
    return calzado, ficha


def construir():
    from openpyxl import load_workbook
    calzado, ficha = leer_maestro()
    # La marca por codigo, para la tabla del ASN. Sale de la misma ficha.
    global MARCA_DE
    MARCA_DE = {c: v[1] for c, v in ficha.items()}

    clase = defaultdict(lambda: {"lineas": 0, "env": 0.0, "rec": 0.0})
    mes = defaultdict(lambda: {"env": 0.0, "rec": 0.0})
    estado = defaultdict(lambda: {"asn": set(), "env": 0.0, "rec": 0.0})
    asn_env, asn_rec, asn_info = defaultdict(float), defaultdict(float), {}
    # DE DONDE VIENE CADA ASN. Un ASN aparece en muchas lineas, asi que el tipo se
    # decide una vez -la primera- y las cantidades se suman aparte.
    asn_tipo = {}
    asn_codigo = {}
    # LA FILA DE LA TABLA: una por articulo dentro de cada ASN. Es el nivel al
    # que se consulta -"que trae el expediente", "que paso con este ASN", "donde
    # esta este articulo"-. Las lineas sueltas no se guardan: el LPN y la talla
    # no se buscan desde la web y serian ~250 MB en vez de 22.
    tabla = defaultdict(lambda: {"env": 0.0, "rec": 0.0, "n": 0, "desc": ""})
    tabla_meta = {}
    tipo_acum = defaultdict(lambda: {"asn": set(), "env": 0.0, "rec": 0.0})
    # Lo RECIBIDO, por el mes en que entro de verdad al sistema. Es la unica
    # fecha que responde 'que se recibio entre tal y tal dia': ni la de envio
    # ni la de creacion sirven para eso.
    recep = defaultdict(lambda: {'unid': 0.0, 'asn': set(), 'lineas': 0})
    sin_fecha_rec = {'unid': 0.0, 'lineas': 0}
    archivos = []

    # ══════════════════════════════════════════════════════════════════════════
    # EL DETALLE QUE PIDIO DANIEL EL 03-sep-2026
    # ══════════════════════════════════════════════════════════════════════════
    #
    # *"necesito un reporte que me diga que articulo esta llegando, que marca esta
    # llegando... en el mes a mes, darle clic y ver que es lo que esta faltando.
    # Esa es la idea de tener el ASN"*.
    #
    # Todo esto ya estaba en los archivos —traen `Articulo` y `Descripcion`— y se
    # tiraba: hasta hoy solo se publicaban totales.
    #
    # SE ACUMULA TODO Y SE RECORTA AL FINAL, no al vuelo: recortar mientras se lee
    # obligaria a decidir con la mitad de los datos y el que mas falta puede
    # aparecer en el ultimo archivo.
    art = defaultdict(lambda: {"e": 0.0, "r": 0.0, "n": 0, "d": "", "m": "", "g": ""})
    marca = defaultdict(lambda: {"e": 0.0, "r": 0.0, "n": 0})
    por_mes_art = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))

    # ── CUANDO LLEGA CADA COSA ────────────────────────────────────────────────
    #
    # Daniel, 03-sep-2026: *"necesito saber la fecha aproximada en que va a llegar
    # y los SKU y marcas, porque si no, no puedo preparar el almacen"*.
    #
    # Se mide LINEA POR LINEA, no por articulo: cada linea trae su propia fecha y
    # un mismo codigo llega en varias tandas distintas.
    #
    # SE SEPARA LO VENCIDO. Medido el 03-sep: de 6.852.012 pendientes, el 37,6%
    # tiene la fecha YA PASADA. Pintar eso junto a lo futuro seria decirle que va
    # a llegar algo que ya fallo una vez.
    por_dia = defaultdict(lambda: defaultdict(float))
    por_dia_marca = defaultdict(lambda: defaultdict(float))
    por_dia_tipo = defaultdict(lambda: defaultdict(float))
    venc_edad = defaultdict(float)
    cuando = {"futuro": 0.0, "vencido": 0.0, "sin_fecha": 0.0,
              "lin_fut": 0, "lin_ven": 0, "lin_sin": 0}
    HOY = datetime.now().date()

    for nombre in sorted(f for f in os.listdir(CARPETA_ASN) if f.lower().endswith(".xlsx")):
        etiqueta = nombre.replace("ASN ", "").replace(".xlsx", "")
        ruta = os.path.join(CARPETA_ASN, nombre)
        wb = load_workbook(ruta, read_only=True, data_only=True)
        hh = wb[wb.sheetnames[0]]
        it = hh.iter_rows(values_only=True)
        idx = None
        for fila in it:
            if fila and any(c is not None for c in fila):
                txt = [str(c).strip() if c is not None else "" for c in fila]
                if "Número de ASN" in txt:
                    idx = {n: i for i, n in enumerate(txt) if n}
                    break
        if idx is None:
            log("%s sin encabezado; se salta" % nombre, "WARN")
            wb.close()
            continue

        iN, iA = idx.get("Número de ASN"), idx.get("Artículo")
        iE, iR = idx.get("Cantidad enviada"), idx.get("Cantidad recibida")
        iS, iF, iD = idx.get("Estado"), idx.get("Fecha de envío"), idx.get("Descripción")
        # La columna se agrego el 01-sep-2026. Los archivos bajados antes no la
        # traen: se acepta que falte y el bloque queda vacio hasta la proxima bajada.
        iV = idx.get("Fecha de recepción") or idx.get("verified_ts")
        # El Tipo ASN, codificado. Se agrego al web report el 03-sep-2026; los
        # archivos bajados antes no la traen y el tipo sale del numero.
        iT = idx.get("cust_field_1")
        # QUIEN RECIBIO y A QUE HORA. La columna del usuario se agrego al informe
        # el 03-sep-2026 para poder medir la productividad de recepcion por
        # persona, como ya se mide la de picking. Los archivos bajados antes no la
        # traen: se acepta que falte y esas filas quedan sin usuario.
        iU = idx.get("verified_user")

        n = 0
        for fila in it:
            if not fila or iN is None or iN >= len(fila) or fila[iN] is None:
                continue
            asn = str(fila[iN]).strip()
            if not asn:
                continue
            n += 1
            env, rec = af(fila[iE]), af(fila[iR])
            cod = str(fila[iA]).strip() if iA is not None and iA < len(fila) and fila[iA] else ""
            cl = "calzado" if cod.split("-")[0].lstrip("0") in calzado else "no_calzado"

            c = clase[cl]
            c["lineas"] += 1
            c["env"] += env
            c["rec"] += rec

            mes[etiqueta]["env"] += env
            mes[etiqueta]["rec"] += rec

            # EL TIPO SE DECIDE UNA VEZ POR ASN. Manda el codigo del WMS; si no
            # esta -o es uno de los que se contradice- entra el patron del numero,
            # que acierta el 97,4% comparado con el codigo.
            if asn not in asn_tipo:
                # `cod_tipo`, NO `cod`. Aca decia `cod`, que es el codigo de
                # ARTICULO de esta misma fila y se usa mas abajo: cada vez que
                # aparecia un ASN nuevo, su articulo quedaba reemplazado por el
                # codigo de tipo. Por eso el reporte mostraba articulos llamados
                # "24", "23" o "56" con descripcion de calzado real.
                cod_tipo = ""
                if iT is not None and iT < len(fila) and fila[iT] is not None:
                    cod_tipo = str(fila[iT]).strip()
                asn_codigo[asn] = cod_tipo
                asn_tipo[asn] = tipo_del_wms(cod_tipo) or tipo_por_el_numero(asn)
            t = tipo_acum[asn_tipo[asn]]
            t["asn"].add(asn)
            t["env"] += env
            t["rec"] += rec

            # la fila de la tabla
            f = tabla[(asn, cod)]
            f["env"] += env
            f["rec"] += rec
            f["n"] += 1
            if not f["desc"] and iD is not None and iD < len(fila) and fila[iD]:
                f["desc"] = limpio(str(fila[iD]).strip()[:70])
            if asn not in tabla_meta:
                m = ORDEN_EN_EL_ASN.match(asn)
                tabla_meta[asn] = (
                    str(fila[idx["Proveedor"]]).strip()
                    if "Proveedor" in idx and idx["Proveedor"] < len(fila)
                    and fila[idx["Proveedor"]] else "",
                    "%s-%s-%s" % (m.group(1), m.group(2), m.group(3)) if m else "",
                    m.group(4) if m else "",
                    asn_tipo[asn],
                    str(fila[iS]).strip() if iS is not None and iS < len(fila) and fila[iS] else "",
                    _fecha10(fila[iF]) if iF is not None and iF < len(fila) else "",
                    _fecha10(fila[iV]) if iV is not None and iV < len(fila) else "",
                    # LA HORA VA APARTE de la fecha, no pegada: si fuera
                    # "2026-09-02 19:53", el filtro por dia <= "2026-09-02"
                    # dejaria fuera todo lo de ese dia. Separadas, los dos
                    # filtros siguen siendo una comparacion simple.
                    _hora(fila[iV]) if iV is not None and iV < len(fila) else "",
                    str(fila[iU]).strip() if iU is not None and iU < len(fila) and fila[iU] else "",
                )

            # el detalle por articulo, por marca y por mes
            if cod:
                gen, mk = ficha.get(cod[:7], ("(no esta en el Maestro)", "(sin marca)"))
                a = art[cod]
                a["e"] += env
                a["r"] += rec
                a["n"] += 1
                if not a["d"]:
                    a["d"] = limpio(str(fila[iD]).strip()[:70]
                                    if iD is not None and iD < len(fila) and fila[iD] else "")
                    a["m"] = mk
                    a["g"] = gen
                k = marca[mk]
                k["e"] += env
                k["r"] += rec
                k["n"] += 1
                x = por_mes_art[etiqueta][cod]
                x[0] += env
                x[1] += rec

                # lo que todavia no llego, con la fecha que anuncia el ASN
                pend = env - rec
                if pend > 0.5:
                    fe = (fecha_de(fila[iF])
                          if iF is not None and iF < len(fila) else None)
                    if fe is None:
                        cuando["sin_fecha"] += pend
                        cuando["lin_sin"] += 1
                    elif fe.date() >= HOY:
                        cuando["futuro"] += pend
                        cuando["lin_fut"] += 1
                        d = fe.date().isoformat()
                        por_dia[d][cod] += pend
                        por_dia_marca[d][mk] += pend
                        por_dia_tipo[d][asn_tipo[asn]] += pend
                    else:
                        cuando["vencido"] += pend
                        cuando["lin_ven"] += 1
                        dd = (HOY - fe.date()).days
                        venc_edad["1a7" if dd <= 7 else "8a30" if dd <= 30
                                  else "31a90" if dd <= 90 else "mas90"] += pend

            est = str(fila[iS]).strip() if iS is not None and iS < len(fila) and fila[iS] else "(sin estado)"
            e = estado[est]
            e["asn"].add(asn)
            e["env"] += env
            e["rec"] += rec

            asn_env[asn] += env
            asn_rec[asn] += rec

            if rec > 0:
                fv = fecha_de(fila[iV]) if iV is not None and iV < len(fila) else None
                if fv is None:
                    sin_fecha_rec["unid"] += rec
                    sin_fecha_rec["lineas"] += 1
                else:
                    r = recep["%04d-%02d" % (fv.year, fv.month)]
                    r["unid"] += rec
                    r["asn"].add(asn)
                    r["lineas"] += 1
            if asn not in asn_info:
                fe = fecha_de(fila[iF]) if iF is not None and iF < len(fila) else None
                asn_info[asn] = {
                    "estado": est,
                    "envio": fe.strftime("%d/%m/%Y") if fe else "",
                }
        wb.close()
        archivos.append({"mes": etiqueta, "lineas": n,
                         "mb": round(os.path.getsize(ruta) / 1048576.0, 1)})
        log("%s: %s lineas" % (nombre, "{:,}".format(n)))

    # ── llego lo que dijo el ASN ────────────────────────────────────────
    cumpl = {k: {"asn": 0, "env": 0.0, "rec": 0.0}
             for k in ("completo", "parcial", "sin_recibir", "de_mas")}
    parciales = []
    for asn, env in asn_env.items():
        rec = asn_rec[asn]
        if rec == 0 and env > 0:
            k = "sin_recibir"
        elif rec > env:
            k = "de_mas"
        elif rec == env and env > 0:
            k = "completo"
        else:
            k = "parcial"
        cumpl[k]["asn"] += 1
        cumpl[k]["env"] += env
        cumpl[k]["rec"] += rec
        if k == "parcial":
            info = asn_info.get(asn, {})
            parciales.append({
                "asn": asn, "envio": info.get("envio", ""), "estado": info.get("estado", ""),
                "enviado": env, "recibido": rec, "falta": env - rec,
                "cumple": round(100.0 * rec / env, 1) if env else 0,
            })
    parciales.sort(key=lambda p: -p["falta"])

    # ── el detalle, ya recortado ─────────────────────────────────────────────
    con_falta = sorted(((c, v) for c, v in art.items() if v["e"] - v["r"] > 0.5),
                       key=lambda x: -(x[1]["e"] - x[1]["r"]))
    # ── EL MES TIENE QUE DECIR LO MISMO QUE EL CUADRO DE ARRIBA ──────────────
    #
    # Daniel, 03-sep-2026: *"en el mes a mes, el 26 del cuatro dice que falta 160
    # mil: necesito darle clic y ver que es lo que esta faltando"*. Ese 160.773 es
    # el NETO de abril -enviado menos recibido- del cuadro "meses", que es el que
    # el esta mirando cuando lo dice.
    #
    # Sumando solo los faltantes se obtiene 187.691, porque deja fuera los 26.918
    # que llegaron DE MAS. Dos filas pegadas diciendo 160 mil y 187 mil del mismo
    # abril es exactamente lo que el detecta con la calculadora, y con razon: no
    # hay forma de saber cual de las dos es la buena.
    #
    # Asi que el mes publica las tres, y las tres cierran entre si:
    #
    #     falta (neto)  =  faltaBruta  -  sobra
    #     160.773       =  187.691     -  26.918
    #
    # `falta` es la que cuadra con el cuadro de arriba y con el de marcas.
    por_mes = {}
    for m in sorted(por_mes_art):
        todos = list(por_mes_art[m].items())
        arts = [(c, v) for c, v in todos if v[0] - v[1] > 0.5]
        if not arts:
            continue
        arts.sort(key=lambda x: -(x[1][0] - x[1][1]))
        bruta = sum(v[0] - v[1] for _, v in arts)
        sobra = sum(v[1] - v[0] for _, v in todos if v[1] - v[0] > 0.5)
        por_mes[m] = {
            "articulos": len(arts),
            "falta": round(bruta - sobra),
            "faltaBruta": round(bruta),
            "sobra": round(sobra),
            "sobraArticulos": sum(1 for _, v in todos if v[1] - v[0] > 0.5),
            "top": [{"cod": c, "marca": art[c]["m"],
                     "env": round(v[0]), "rec": round(v[1]), "falta": round(v[0] - v[1])}
                    for c, v in arts[:TOPE_POR_MES]],
        }

    # ── EL CALENDARIO DE LO QUE VIENE ────────────────────────────────────────
    #
    # Los dias de aca a DIAS_CERCA van COMPLETOS -son sobre los que se prepara el
    # almacen- y los lejanos recortados. Cada dia dice si esta completo o no: una
    # lista recortada que no avisa se lee como la lista entera.
    dias_llegada = []
    for d in sorted(por_dia):
        arts = sorted(por_dia[d].items(), key=lambda x: -x[1])
        cerca = (datetime.strptime(d, "%Y-%m-%d").date() - HOY).days <= DIAS_CERCA
        top = arts if cerca else arts[:TOPE_DIA_LEJOS]
        # LA FECHA PROBABLE DE ENTRADA. Un dia puede mezclar tipos, asi que la
        # demora se pondera por unidades en vez de tomar la del tipo mayor: un
        # dia con 90% de importacion y 10% de devolucion no entra en 155 dias.
        mezcla = por_dia_tipo.get(d) or {}
        totd = sum(mezcla.values())
        demora = (sum(DEMORA_DIAS.get(t, 12) * q for t, q in mezcla.items()) / totd
                  if totd else 12)
        entra = (datetime.strptime(d, "%Y-%m-%d").date()
                 + timedelta(days=int(round(demora)))).isoformat()
        dias_llegada.append({
            "dia": d,
            "entra": entra,
            "demora": int(round(demora)),
            "tipos": [{"t": t, "u": round(q)} for t, q in
                      sorted(mezcla.items(), key=lambda x: -x[1])],
            "u": round(sum(v for _, v in arts)),
            "n": len(arts),
            "marcas": [{"m": m, "u": round(q)} for m, q in
                       sorted(por_dia_marca[d].items(), key=lambda x: -x[1])],
            "top": [{"cod": c, "marca": art[c]["m"], "u": round(q)} for c, q in top],
            "completo": len(top) == len(arts),
        })

    # La proxima fecha de cada codigo, para la columna CUANDO del cuadro de
    # articulos: *"puedes anadir ahi, en esos reportes que ya tienes, la fecha"*.
    proxima = {}
    for d in sorted(por_dia):
        for c in por_dia[d]:
            if c not in proxima:
                proxima[c] = d

    cuando_llega = {
        "hoy": HOY.isoformat(),
        "diasCerca": DIAS_CERCA,
        "futuro": round(cuando["futuro"]),
        "vencido": round(cuando["vencido"]),
        "sinFecha": round(cuando["sin_fecha"]),
        "lineasFuturo": cuando["lin_fut"],
        "lineasVencido": cuando["lin_ven"],
        "dias": dias_llegada,
        # Para que la pantalla pueda decir de donde sale la fecha de entrada.
        "demoraDias": DEMORA_DIAS,
        "vencidoEdad": {k: round(v) for k, v in venc_edad.items()},
    }

    # LA TABLA VIAJA APARTE, no dentro del paquete: son 76.658 filas y el paquete
    # es lo que se baja el navegador entero. Se cuelga del objeto para que main()
    # la mande al endpoint, y se quita antes de publicar.
    paquete = {
        "_tabla": [tabla, tabla_meta],
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "archivos": archivos,
        "totales": {
            "lineas": sum(c["lineas"] for c in clase.values()),
            "asn": len(asn_env),
            "enviado": sum(c["env"] for c in clase.values()),
            "recibido": sum(c["rec"] for c in clase.values()),
        },
        "clase": {k: {"lineas": v["lineas"], "enviado": v["env"], "recibido": v["rec"]}
                  for k, v in clase.items()},
        "meses": [{"mes": m, "enviado": d["env"], "recibido": d["rec"]}
                  for m, d in sorted(mes.items())],
        "recepciones": [{"mes": m, "unidades": d["unid"], "asn": len(d["asn"]),
                         "lineas": d["lineas"]}
                        for m, d in sorted(recep.items())],
        "recibido_sin_fecha": sin_fecha_rec,
        "estados": sorted(
            [{"estado": k, "asn": len(v["asn"]), "enviado": v["env"], "recibido": v["rec"]}
             for k, v in estado.items()],
            key=lambda x: -x["enviado"]),
        "cumplimiento": cumpl,
        # LOS PARCIALES, TODOS. Antes se publicaban 60 de 122 y Daniel pidio un
        # filtro sobre ellos: filtrar la mitad de la lista no sirve. Son chicos
        # -18 KB los 139- y son justo los que hay que perseguir.
        # ── DE DONDE VIENE LO QUE LLEGA ──────────────────────────────────────
        #
        # Daniel, 03-sep-2026: *"falta anadir importacion, nacional, logistica
        # inversa y otras cosas al reporte"*.
        #
        # Se publica tambien de donde salio cada clasificacion, porque no es lo
        # mismo un dato del sistema que una deduccion: la pantalla lo dice.
        "tipos": sorted([{
            "tipo": k,
            "asn": len(v["asn"]),
            "enviado": round(v["env"]),
            "recibido": round(v["rec"]),
            "falta": round(v["env"] - v["rec"]),
            "cumple": round(100.0 * v["rec"] / v["env"], 1) if v["env"] else 0,
        } for k, v in tipo_acum.items()], key=lambda x: -x["falta"]),
        "tipoFuente": {
            "porCodigo": sum(1 for a in asn_tipo if tipo_del_wms(asn_codigo.get(a, ""))),
            "porNumero": sum(1 for a in asn_tipo if not tipo_del_wms(asn_codigo.get(a, ""))),
        },
        "parciales": parciales,
        "parciales_total": len(parciales),
        "parciales_falta": sum(p["falta"] for p in parciales),

        # ── EL DETALLE ────────────────────────────────────────────────────────
        #
        # LOS TOPES SALEN DE UNA MEDICION. Son 23.533 articulos distintos y 11.307
        # con algo pendiente: publicarlos todos son 1.117 KB al navegador, y eso
        # es lo que ya hizo lenta la web una vez. Cortando por los que mas faltan,
        # 500 articulos + 50 por mes + las 13 marcas dan 183 KB, que Daniel aprobo
        # mirando la maqueta.
        #
        # LA PANTALLA DICE SIEMPRE CUANTOS QUEDARON FUERA. Una tabla recortada que
        # no avisa se lee como la lista completa.
        "articulosDistintos": len(art),
        "articulosConFalta": len(con_falta),
        # `desc` ya no viaja en cada fila: se repetia en articulos, en cada mes y
        # en cada dia del calendario. Ahora va UNA vez en `desc`, mas abajo.
        "articulos": [{
            "cod": c, "marca": v["m"], "gender": v["g"],
            "env": round(v["e"]), "rec": round(v["r"]),
            "falta": round(v["e"] - v["r"]), "lineas": v["n"],
            "prox": proxima.get(c, ""),
        } for c, v in con_falta[:TOPE_ARTICULOS]],
        "marcas": sorted([{
            "marca": m, "env": round(v["e"]), "rec": round(v["r"]),
            "falta": round(v["e"] - v["r"]), "lineas": v["n"],
            "cumple": round(100.0 * v["r"] / v["e"], 1) if v["e"] else 0,
        } for m, v in marca.items()], key=lambda x: -x["env"]),
        "porMes": por_mes,
        "cuandoLlega": cuando_llega,
    }

    # ── LAS DESCRIPCIONES, UNA SOLA VEZ ──────────────────────────────────────
    #
    # Repetirlas en cada fila era la mitad del peso del paquete, y un mismo
    # codigo aparece en el cuadro de articulos, en varios meses y en varios dias.
    codigos = set(a["cod"] for a in paquete["articulos"])
    for m in por_mes.values():
        codigos.update(a["cod"] for a in m["top"])
    for d in dias_llegada:
        codigos.update(a["cod"] for a in d["top"])
    paquete["desc"] = {c: art[c]["d"] for c in sorted(codigos) if art[c]["d"]}

    for k in ("articulos", "marcas", "parciales", "desc"):
        log("   %-11s %6.0f KB"
            % (k, len(json.dumps(paquete[k], ensure_ascii=False).encode("utf-8")) / 1024.0))
    log("   %-11s %6.0f KB (%d dias, %d completos)"
        % ("cuandoLlega",
           len(json.dumps(cuando_llega, ensure_ascii=False).encode("utf-8")) / 1024.0,
           len(dias_llegada), sum(1 for d in dias_llegada if d["completo"])))
    return paquete


LOTE_TABLA = 4000        # 76.658 filas de un saque son 15 MB en una peticion


def cargar_tabla(tabla, meta, marca_de):
    """Manda la tabla al servidor en tres pasos.

    NO ES CRITICO: si esto falla, el paquete ya se publico y la pantalla sigue
    andando como hasta hoy. Por eso devuelve True/False en vez de reventar.
    """
    filas = []
    for (asn, cod), f in tabla.items():
        m = meta.get(asn, ("", "", "", "", "", "", "", "", ""))
        filas.append([asn, cod, f["desc"], marca_de(cod), m[0], m[1], m[2], m[3],
                      m[4], m[5], m[6], m[7], m[8],
                      int(round(f["env"])), int(round(f["rec"])), f["n"]])
    log("tabla del ASN: %s filas" % "{:,}".format(len(filas)))
    # A LOS DOS ENTORNOS, igual que el paquete: produccion y beta tienen bases
    # distintas y la tabla vive en la base.
    # SE IMPORTA ADENTRO, como hace `subir()`. El modulo vive junto a este
    # archivo en el servidor y no esta importado arriba: darlo por presente hizo
    # que la primera carga fallara con "name 'publicar_area' is not defined".
    from publicar_area import pedir_json

    ok_prod = False
    for nombre, cabecera in (("produccion", None), ("beta", "beta")):
        try:
            pedir_json("/api/asn/carga", {"paso": "inicio"}, cabecera)
            for i in range(0, len(filas), LOTE_TABLA):
                pedir_json(
                    "/api/asn/carga",
                    {"paso": "lote", "filas": filas[i:i + LOTE_TABLA]}, cabecera)
            r = pedir_json("/api/asn/carga", {"paso": "fin"}, cabecera)
            if (r or {}).get("status") == "ok":
                log("   tabla cargada en %s: %s filas"
                    % (nombre, "{:,}".format(r.get("filas", 0))))
                ok_prod = ok_prod or nombre == "produccion"
                continue
            log("   la carga en %s no termino: %s" % (nombre, str(r)[:110]), "WARN")
        except Exception as e:
            log("   no se pudo cargar la tabla en %s: %s" % (nombre, str(e)[:110]), "WARN")
    if not ok_prod:
        log("   el paquete SI se publico; la pantalla sigue andando", "WARN")
    return ok_prod


def subir(paquete, intentos=3):
    """Sube el paquete a PRODUCCION Y A BETA.

    ANTES SOLO IBA A PRODUCCION, y por eso beta decia "todavia no hay ASN
    publicado" — Daniel, 03-sep-2026: *"no veo nada"*. No era que el robot
    fallara: es que nunca le habia mandado nada a beta, asi que no habia forma de
    probar ningun cambio de esta pantalla antes de soltarlo a produccion.

    Se usa `publicar_area.publicar`, que es el que ya sabe mandar a los dos y que
    usan los demas robots. Escribirlo otra vez aca es como se separan las dos
    verdades.

    VA CON `date=MASTER`, y no es opcional: el area guarda un OBJETO y no filas
    sueltas. Sin ese parametro el servidor responde 200 igual y lo guarda como
    lista vacia — la pantalla queda en blanco y nada avisa.
    """
    from publicar_area import publicar
    return publicar(AREA, paquete, "MASTER",
                    log=lambda t, n="INFO": log(t, n), intentos=intentos)


def main(log_externo=None):
    global _LOG_EXTERNO
    if log_externo:
        _LOG_EXTERNO = log_externo
    log("=" * 58)
    log("RESUMEN DEL ASN PARA LA WEB")
    log("=" * 58)
    rutas()
    if not os.path.isdir(CARPETA_ASN):
        log("no existe la carpeta %s" % CARPETA_ASN, "ERROR")
        return 2
    p = construir()
    t = p["totales"]
    log("%s lineas · %s ASN · enviado %s · recibido %s"
        % ("{:,}".format(t["lineas"]), "{:,}".format(t["asn"]),
           "{:,.0f}".format(t["enviado"]), "{:,.0f}".format(t["recibido"])))
    if p.get("recepciones"):
        log("recepciones: %d meses con fecha; %s unidades sin fecha de recepcion"
            % (len(p["recepciones"]), "{:,.0f}".format(p["recibido_sin_fecha"]["unid"])))
    else:
        log("los archivos todavia no traen la fecha de recepcion", "WARN")
    log("parciales: %s ASN, faltan %s unidades"
        % ("{:,}".format(p["parciales_total"]), "{:,.0f}".format(p["parciales_falta"])))
    # LA TABLA SALE DEL PAQUETE SIEMPRE, aunque no se suba: si se quedara
    # adentro, `--solo-calcular` escribiria un JSON con las 76.658 filas.
    tabla, tabla_meta = p.pop("_tabla", (None, None))

    if "--solo-calcular" in sys.argv:
        destino = os.path.join(os.path.dirname(os.path.abspath(__file__)), "asn_resumen.json")
        with open(destino, "w", encoding="utf-8") as f:
            json.dump(p, f, ensure_ascii=False, indent=1)
        log("guardado en %s (no se publico)" % destino)
        return 0
    # ── LA TABLA, para poder consultar los seis meses desde la web ───────────
    #
    # Va DESPUES de publicar el paquete y no antes: si la tabla falla, el paquete
    # ya esta arriba y la pantalla sigue andando como hasta hoy.
    ok = subir(p)
    if tabla and "--sin-tabla" not in sys.argv:
        cargar_tabla(tabla, tabla_meta, lambda c: (MARCA_DE or {}).get(c[:7], ""))
    return 0 if ok else 3


if __name__ == "__main__":
    sys.exit(main())
