# -*- coding: utf-8 -*-
"""
================================================================================
 DISTRIBUCION Y DESPACHO POTENCIAL  -  las dos pantallas de Despacho
================================================================================

Daniel, 05-sep-2026, sobre los bultos que llevan dias parados: *"eso es lo que
quiero detectar. Ahorita lo hacen manualmente. Yo lo que quiero es
automatizarlo"*.

Este robot arma y publica tres cosas:

    distribucion_dia        el cuadro de Retail, los dos pivots de turno x zona,
                            que hay en el patio y en staging, y los varados
    distribucion_detalle    el desglose por articulo de cada bulto, SOLO para el
                            Excel: son 18.000 renglones y no puede viajar con la
                            pantalla, que se abre cien veces al dia
    despacho_potencial_dia  por tienda: patio + staging + el correo de comercial

NO TOCA EL WMS. Lee los archivos que ya bajaron los otros robots, asi que no
compite por la sesion de Oracle ni puede tumbar una corrida ajena. Por eso puede
ir pegado a cualquier hora libre.

--------------------------------------------------------------------------------
 LAS SEIS FUENTES
--------------------------------------------------------------------------------
    scraping Stock/Picking/Picking D-M.csv           lo picado del dia
    scraping Stock/Detalle Orden/*.csv               lo que se pidio
    scraping Stock/OBLPN Embalaje/OBLPN *.csv        los bultos: TODOS los dias
    scraping Stock/Correos Picking/Guias DD.MM.xlsx  lo que mando comercial
    Maestro_Articulos.xlsx                           el gender de cada articulo
    Proyecto web Logistico/RUTAS -  TURNOS.xlsx      que destino es tienda

EL DIA LO MANDA EL OBLPN MAS NUEVO, no la fecha de hoy: si una noche el archivo
no baja, el robot rehace el ultimo dia completo en vez de publicar una pantalla
vacia.

--------------------------------------------------------------------------------
 LAS CINCO REGLAS QUE NO SE PUEDEN CAMBIAR SIN HABLARLO
--------------------------------------------------------------------------------
1. EL PREPACK SE ABRE, TAMBIEN EN EL ACCESORIO. El WMS anota 1 por caja. Los dos
   primeros digitos del sufijo de 5 son los pares, con tope 24. Regla de Daniel,
   05-sep-2026: *"asi lo hacemos tanto en el calzado como en el accesorio"*.

2. `Cantidad de orden original` SE REPITE en cada linea de la misma (orden,
   articulo). Se cuenta UNA vez, con `max`. Sumandola salian 123.910 donde hay
   34.856.

3. PATIO = el contenedor sigue siendo un `PRE`. Cuando embalaje lo pistolea, el
   PRE se cancela y nace el LPN real. Asi el tramo cuadra solo:
   picada = embalada + patio. Restar contra el OBLPN no cuadraba.

4. EL ESTADO MANDA, NO EL NOMBRE DEL LPN. Un `PRE` con estado `Enviado` no esta
   varado. Daniel lo cazo el 05-sep-2026 con el PRE500080001704988.

5. PATIO Y STAGING SON ESTADOS, NO DIAS. Se acumulan hasta que alguien los
   mueve, asi que salen de la foto entera del OBLPN y no del pick de un dia.

--------------------------------------------------------------------------------
 POR QUE HACEN FALTA LOS 33 ARCHIVOS DEL OBLPN
--------------------------------------------------------------------------------
El export del WMS trae SOLO lo que se movio ese dia -en el archivo del 04-09 las
43.256 lineas tienen modificacion 04/09, sin una excepcion-. Y lo varado es justo
lo que nadie toca: un archivo solo no lo ve nunca. Juntandolos todos, la ultima
vez que se vio un bulto y en que estado estaba dice cuantos dias lleva parado.
"""

import collections
import csv
import datetime
import glob
import io
import json
import os
import re
import shutil
import sys
import time

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import publicar_area

csv.field_size_limit(10 ** 9)

AQUI = os.path.dirname(os.path.abspath(__file__))
TEMP = os.path.join(AQUI, "_tmp_distribucion")

# El prepack lleva la talla en CINCO digitos: 5614468-1-06006 contra el suelto
# 5614468-1-03. Es la unica forma de reconocerlo; la ubicacion no alcanza.
FORMA_PREPACK = re.compile(r'^\d{7}-\d-\d{5}$')

# Un bulto parado en cualquiera de estos dos estados es mercaderia detenida.
PARADO = ('Empaquetado', 'En empaquetado')

TRAMOS = ['se movió ayer', '1 día', '2 a 3 días', '4 a 7 días',
          '8 a 14 días', 'más de 14 días']


# ══ EL REGISTRO ═══════════════════════════════════════════════════════════════
#
# Sin esto no hay forma de contestar "corrio o no corrio". El 05-sep-2026 la
# primera corrida sola salio a las 23:00 en vez de las 22:00 y no se pudo saber
# por que: ni el robot ni el registro de tareas de Windows dejaron nada.
CARPETA_LOGS = os.path.join(AQUI, 'logs')
_ARCHIVO_LOG = os.path.join(
    CARPETA_LOGS,
    'distribucion_%s.log' % datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S'))


def log(msg, nivel=''):
    # `publicar_area.publicar` llama al log con un segundo argumento (ERROR o
    # AVISO). Si no se acepta, el robot muere justo al publicar.
    linea = '[%s] %s%s' % (datetime.datetime.now().strftime('%H:%M:%S'),
                           (nivel + ' ') if nivel else '', msg)
    print(linea)
    sys.stdout.flush()
    # QUE EL LOG NO PUEDA TUMBAR LA CORRIDA: si el disco falla se sigue igual.
    try:
        if not os.path.isdir(CARPETA_LOGS):
            os.makedirs(CARPETA_LOGS)
        with io.open(_ARCHIVO_LOG, 'a', encoding='utf-8') as fh:
            fh.write(linea + '\n')
    except Exception:
        pass


def limpiar_logs(dias=30):
    """Los suyos de mas de un mes, fuera. Son 2 KB cada uno, pero uno por dia
       durante un año es basura que despues nadie mira."""
    try:
        corte = time.time() - dias * 86400
        for n in os.listdir(CARPETA_LOGS):
            if n.startswith('distribucion_') and n.endswith('.log'):
                p = os.path.join(CARPETA_LOGS, n)
                if os.path.getmtime(p) < corte:
                    os.remove(p)
    except Exception:
        pass


# ══ LAS CARPETAS ══════════════════════════════════════════════════════════════
#
# Corriendo como tarea programada el usuario es SYSTEM, no Administrator, y su
# perfil es otro: `~/OneDrive` no existe y la variable de entorno tampoco. Por
# eso van tambien las rutas fijas de las dos maquinas, igual que en
# `oblpn_embalaje.py` y `correo_guias.py`.
def base_onedrive():
    for c in (os.environ.get("OneDrive"), os.environ.get("OneDriveCommercial"),
              os.path.join(os.path.expanduser("~"), "OneDrive"),
              os.path.join("C:", os.sep, "Users", "Administrator", "OneDrive"),
              os.path.join("C:", os.sep, "Users", "dames", "OneDrive")):
        if not c:
            continue
        ruta = os.path.join(c, "danielames.bata")
        if os.path.isdir(ruta):
            return ruta
    return None


def L(v):
    """El WMS exporta algunas columnas como `="0123"` para que Excel no se coma
       los ceros de la izquierda. Se limpian aca, en un solo sitio."""
    v = (v or '').strip()
    if v.startswith('="') and v.endswith('"'):
        v = v[2:-1]
    return v.strip()


def N(v):
    v = L(v).replace(',', '')
    try:
        return float(v or 0)
    except ValueError:
        return 0.0


def pares(sku, q):
    """Los pares de verdad. Ver la regla 1 de la cabecera."""
    if not FORMA_PREPACK.match(L(sku)):
        return q
    n = int(L(sku)[-5:][:2])
    return q * n if 0 < n <= 24 else q


def s7(cod):
    c = L(cod).split('-')[0]
    return c.zfill(7)[-7:] if c else ''


def corto(fecha):
    """`04/09/2026` -> `04-09`, que es como se ve en la pantalla."""
    p = fecha.split('/')
    return p[0] + '-' + p[1] if len(p) == 3 else fecha


# ══ EL MAESTRO Y LAS RUTAS ════════════════════════════════════════════════════
#
# EL MAESTRO TIENE TRES COPIAS Y NO SIEMPRE MANDA LA MISMA. Al 05-sep-2026 la
# fresca es la de la raiz (25-ago, 30.175 articulos) y la de `Pruebas Sistema`
# -que era la buena antes- se quedo en el 05-ago con 29.465. Por eso salian
# articulos "sin gender": no estaban en la copia vieja. Se toma SIEMPRE la mas
# reciente de las tres.
#
# Y SE COPIA ANTES DE ABRIR: en OneDrive el archivo puede estar solo en la nube y
# openpyxl lo ve como un zip roto.
def leer_maestro(base):
    copias = [os.path.join(base, 'Maestro_Articulos.xlsx'),
              os.path.join(base, 'Pruebas Sistema', 'Maestro_Articulos.xlsx'),
              os.path.join(base, 'scraping Stock', 'Archivos', 'Maestro_Articulos.xlsx')]
    hay = [p for p in copias if os.path.exists(p)]
    if not hay:
        raise SystemExit('No hay ninguna copia del Maestro.')
    src = max(hay, key=os.path.getmtime)
    dst = os.path.join(TEMP, '_maestro.xlsx')
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    gen, h = {}, None
    for row in ws.iter_rows(values_only=True):
        if h is None:
            h = [str(c or '').strip() for c in row]
            iC, iG = h.index('CodArticulo'), h.index('G. Gender')
            continue
        c = str(row[iC] or '').strip()
        if c:
            gen[s7(c)] = str(row[iG] or '').strip()
    wb.close()
    t = datetime.datetime.fromtimestamp(os.path.getmtime(src))
    log('Maestro: %s (%s) - %d articulos'
        % (os.path.basename(os.path.dirname(src)) or '.', t.strftime('%d-%b'), len(gen)))
    return gen


def leer_rutas(base):
    """RETAIL ES LO QUE ESTA EN EL MAESTRO DE RUTAS. Un destino que no figura ahi
       no es tienda: es catalogo, ecommerce o una devolucion."""
    src = os.path.join(base, 'Proyecto web Logistico', 'RUTAS -  TURNOS.xlsx')
    dst = os.path.join(TEMP, '_rutas.xlsx')
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    T, h = {}, None
    for row in ws.iter_rows(values_only=True):
        if h is None:
            h = [str(c or '').strip().upper() for c in row]
            iC, iT = h.index('CDG'), h.index('TIENDA')
            iZ, iR, iU = h.index('ZONA'), h.index('RUTA2'), h.index('TURNO')
            continue
        if row[iC] is None:
            continue
        k = str(row[iC]).strip().split('.')[0]
        T[k] = {'t': str(row[iT] or '').strip(), 'z': str(row[iZ] or '').strip(),
                'r': str(row[iR] or '').strip(), 'u': str(row[iU] or '').strip()}
    wb.close()
    log('Rutas: %d tiendas' % len(T))
    return T


# ══ QUE DIA SE PUBLICA ════════════════════════════════════════════════════════
def dia_del_oblpn(nombre):
    d, m = os.path.basename(nombre)[6:-4].split('-')
    return datetime.date(2026, int(m), int(d))


def elegir_dia(ss):
    """El dia es el ultimo OBLPN QUE YA TIENE SU ARCHIVO DE PICKING.

       No alcanza con el mas nuevo: el del dia en curso existe desde la
       madrugada con cuatro lineas, y tomarlo deja la pantalla en cero. El
       picking lo baja el Corte del turno cuando la jornada termino, asi que su
       existencia es la senal de que el dia esta cerrado.

       Devuelve (todos los archivos, el del dia, la fecha)."""
    arch = sorted(glob.glob(os.path.join(ss, 'OBLPN Embalaje', 'OBLPN *.csv')),
                  key=dia_del_oblpn)
    if not arch:
        raise SystemExit('No hay ningun archivo de OBLPN.')
    for f in reversed(arch):
        d = dia_del_oblpn(f)
        pick = os.path.join(ss, 'Picking', 'Picking %d-%d.csv' % (d.day, d.month))
        if os.path.exists(pick):
            if f is not arch[-1]:
                log('el OBLPN del %s todavia no tiene su picking: se rehace el %s'
                    % (dia_del_oblpn(arch[-1]).strftime('%d-%m'), d.strftime('%d-%m')))
            # LOS ARCHIVOS POSTERIORES NO ENTRAN. Si se colara el del dia en
            # curso, los bultos de ayer figurarian "vistos hoy" y ninguno saldria
            # varado.
            return [a for a in arch if dia_del_oblpn(a) <= d], f, d
    raise SystemExit('Ningun dia del OBLPN tiene su archivo de picking.')


# ══ 1. EL CUADRO DE RETAIL ════════════════════════════════════════════════════
def cuadro_retail(ss, gen, TIENDAS, fecha, estados):
    """Las cinco columnas de picking salen del archivo de picking del dia; las
       cuatro de despacho vienen de la foto del OBLPN, que es la que dice donde
       esta la mercaderia AHORA. Ver la regla 3 y la nota de los 22 pares."""
    clase = lambda c: 'Footwear' if gen.get(s7(c), '') == 'Footwear' else 'No Footwear'
    dia_csv = '%d-%d' % (fecha.day, fecha.month)          # "4-9", como lo nombra el WMS

    # ── lo pedido, de todos los Detalle Orden ──
    # `Cantidad de orden original` SE REPITE en cada linea: se toma el maximo por
    # (orden, articulo), nunca la suma. Ver la regla 2.
    sol, destDO = {}, {}
    fu = [os.path.join(ss, 'Detalle Orden', 'Detalle Orden Pendientes.csv'),
          os.path.join(ss, 'Detalle Orden', 'Detalle Orden Despachados.csv')]
    fu += sorted(glob.glob(os.path.join(ss, 'Detalle Orden', 'Detalle Orden ??-??.csv')))
    for f in fu:
        try:
            with io.open(f, encoding='utf-8-sig', newline='') as fh:
                for r in csv.DictReader(fh, delimiter=';'):
                    o, c = L(r.get('Número de orden')), L(r.get('Código de artículo'))
                    if not o or not c:
                        continue
                    k = (o, c)
                    q = pares(c, N(r.get('Cantidad de orden original')))
                    if q > sol.get(k, 0.0):
                        sol[k] = q
                    destDO.setdefault(k, L(r.get('Instalación de destino')))
        except OSError:
            pass

    # ── lo picado del dia ──
    ped, dest = {}, {}
    tab = collections.defaultdict(collections.Counter)
    uni = collections.defaultdict(collections.Counter)
    pick = os.path.join(ss, 'Picking', 'Picking %s.csv' % dia_csv)
    if not os.path.exists(pick):
        log('OJO: no hay %s; el cuadro sale sin la parte de picking.'
            % os.path.basename(pick))
    else:
        with io.open(pick, encoding='utf-8-sig', newline='') as fh:
            for r in csv.DictReader(fh, delimiter=';'):
                o, c = L(r['Número de orden']), L(r['Código de artículo'])
                if not o or not c:
                    continue
                k = (o, c)
                ped[k] = max(ped.get(k, 0.0),
                             pares(c, N(r['Cantidad de orden original'])))
                d = L(r['Instalación de destino'])
                dest.setdefault(k, d)
                if r['Estado'] != 'Finalizada' or d not in TIENDAS:
                    continue
                g, q = clase(c), pares(c, N(r['Cantidad asignada']))
                tab[g]['picado'] += q
                if L(r.get('Número de contenedor')).startswith('PRE'):
                    tab[g]['patio'] += q
                else:
                    tab[g]['embalado'] += q

    ordenes = {o for (o, c) in ped}
    retail = lambda k: (dest.get(k) or destDO.get(k) or '') in TIENDAS

    todas = set()
    for k, q in sol.items():
        o, c = k
        if o not in ordenes or k in ped or not retail(k):
            continue
        tab[clase(c)]['pedido'] += q
        uni[o][clase(c)] += q
        todas.add(o)
    for k, q in ped.items():
        o, c = k
        if not retail(k):
            continue
        v = max(q, sol.get(k, 0.0))
        tab[clase(c)]['pedido'] += v
        uni[o][clase(c)] += v
        todas.add(o)
    # el pedido cuenta una sola vez, del lado que mas pares pone
    nPed = collections.Counter(max(v.items(), key=lambda x: x[1])[0]
                               for v in uni.values())

    tabla = []
    for g in ('Footwear', 'No Footwear'):
        v, e = tab[g], estados[g]
        tabla.append({'g': g, 'ped': nPed[g],
                      'qPed': int(v['pedido']), 'qPic': int(v['picado']),
                      'pend': int(v['pedido'] - v['picado']),
                      'emb': int(v['embalado']),
                      # LAS CUATRO DE DESPACHO SALEN DE LA FOTO, para que digan
                      # lo mismo que los dos pivots y que las dos listas.
                      'patio': int(e['patio']), 'stg': int(e['staging']),
                      'car': int(e['Cargado']), 'env': int(e['Enviado'])})
    # LOS CUADROS TIENEN QUE CUADRAR: Daniel suma las filas con la calculadora.
    # La unica diferencia esperable es la del reloj: el picking se baja a las
    # 20:59 y la foto del OBLPN es de las 19:43, asi que un bulto embalado entre
    # medio figura PRE en uno y LPN real en la otra. Si esto crece de unas
    # decenas, es que algo mas cambio.
    pic = sum(f['qPic'] for f in tabla)
    emb = sum(f['emb'] + f['patio'] for f in tabla)
    log('Retail: %d pedidos - picada %d contra embalada+patio %d (%+d por el reloj)'
        % (len(todas), pic, emb, emb - pic))
    return tabla


# ══ 2. LA FOTO: LISTAS, PIVOTS, POTENCIAL Y CONTROL DE PRE ════════════════════
#
# Los cuatro salen del MISMO archivo, asi que se lee UNA vez. Antes eran cuatro
# pasadas por un csv de 43.000 lineas.
def foto_del_dia(ultimo_oblpn, gen, TIENDAS):
    clase = lambda c: 'F' if gen.get(s7(c), '') == 'Footwear' else 'N'
    caja = {'PATIO': {}, 'STAGING': {}}
    pivot = collections.defaultdict(collections.Counter)
    bultos_pivot = collections.defaultdict(set)
    porTienda = collections.defaultdict(collections.Counter)
    detTienda = collections.defaultdict(dict)
    en_bulto = set()                 # las guias que ya estan picadas
    vis = set()
    # DONDE ESTA LA MERCADERIA AHORA, por gender: las cuatro columnas de la
    # banda de Despacho del cuadro grande.
    estados = collections.defaultdict(collections.Counter)

    with io.open(ultimo_oblpn, encoding='utf-8-sig', newline='') as fh:
        for r in csv.DictReader(fh, delimiter=';'):
            q = N(r.get('Cantidad empaquetada'))
            if q <= 0:
                continue
            lpn = L(r.get('Número de LPN'))
            c = L(r.get('Código de artículo'))
            hh = L(r.get('Registro de hora de empaquetado'))
            if (lpn, c, hh) in vis:
                continue
            vis.add((lpn, c, hh))
            d = L(r.get('Instalación de destino'))
            if d not in TIENDAS:
                continue
            estado = L(r.get('Estado de LPN'))
            o = L(r.get('Número de orden'))
            p = pares(c, q)

            g = clase(c)
            gg = 'Footwear' if g == 'F' else 'No Footwear'
            if estado in ('Cargado', 'Enviado'):
                estados[gg][estado] += p
            if estado not in PARADO:
                continue

            etapa = 'PATIO' if lpn.startswith('PRE') else 'STAGING'
            estados[gg]['patio' if etapa == 'PATIO' else 'staging'] += p
            b = caja[etapa].setdefault(lpn, {
                'l': lpn, 'd': d, 't': TIENDAS[d]['t'], 'q': 0.0,
                'dia': L(r.get('Detail Picked Time'))[:10],
                'ped': collections.Counter(), 'it': collections.Counter(), 'ds': {}})
            b['q'] += p
            b['ped'][o] += p
            b['it'][c] += p
            b['ds'][c] = L(r.get('Descripción de artículo'))[:46]
            dd = L(r.get('Detail Picked Time'))[:10]
            if dd and dd < b['dia']:
                b['dia'] = dd

            k = (etapa.lower(), TIENDAS[d]['u'], TIENDAS[d]['z'])
            pivot[k]['cal' if g == 'F' else 'noc'] += p
            bultos_pivot[k].add(lpn)

            porTienda[d][etapa.lower() + g] += p
            dk = ('Patio' if etapa == 'PATIO' else 'Staging', lpn, o)
            db = detTienda[d].setdefault(dk, {'F': 0.0, 'N': 0.0})
            db[g] += p
            en_bulto.add(o)

    # ── las dos listas, ordenadas por lo que mas pesa ──
    listas, arts = {}, {}
    for etapa, nombre in (('PATIO', 'patio'), ('STAGING', 'staging')):
        filas = []
        for b in sorted(caja[etapa].values(), key=lambda x: -x['q']):
            it = []
            for c, q in b['it'].most_common():
                arts.setdefault(c, [b['ds'][c], clase(c)])
                it.append([c, int(q)])
            # EL PEDIDO QUE SE MUESTRA ES EL QUE MAS PESA, no el de numero mas
            # bajo: en el PRE500080002035048 se veia un pedido de 20 pares
            # cuando el que llena el bulto pone 408 de los 449.
            ped = [o for o, _ in b['ped'].most_common()]
            filas.append({'o': ped[0] if len(ped) == 1
                          else ped[0] + ' +' + str(len(ped) - 1),
                          'l': b['l'], 'd': b['d'], 't': b['t'],
                          'f': corto(b['dia']), 'q': int(b['q']), 'i': it})
        listas[nombre] = filas
        log('%-8s %6d pares en %5d bultos'
            % (nombre, sum(f['q'] for f in filas), len(filas)))

    # ── los dos pivots, en el orden del Excel que ellos hacen a mano ──
    ORDEN = [('NOCHE', 'LIMA'), ('DIA', 'LIMA'), ('PROVINCIA', 'PROVINCIA')]
    turnoZona = {}
    for etapa in ('staging', 'patio'):
        filas, T, tb = [], collections.Counter(), set()
        for turno, zona in ORDEN:
            k = (etapa, turno, zona)
            v = pivot.get(k) or collections.Counter()
            filas.append({'turno': turno, 'zona': zona, 'cal': int(v['cal']),
                          'noc': int(v['noc']), 'bultos': len(bultos_pivot[k])})
            T['cal'] += v['cal']
            T['noc'] += v['noc']
            tb |= bultos_pivot[k]
        turnoZona[etapa] = {'filas': filas,
                            'total': {'cal': int(T['cal']), 'noc': int(T['noc']),
                                      'bultos': len(tb)}}
    return listas, arts, turnoZona, porTienda, detTienda, en_bulto, estados


# ══ 3. LOS VARADOS: los 33 archivos juntos ════════════════════════════════════
def varados(archivos, gen, TIENDAS, hoy):
    """La ultima vez que se vio cada bulto y en que estado estaba. Ver la
       explicacion de la cabecera: un archivo solo no ve nunca lo varado."""
    ult = {}
    for f in archivos:
        dia = dia_del_oblpn(f)
        try:
            with io.open(f, encoding='utf-8-sig', newline='') as fh:
                for r in csv.DictReader(fh, delimiter=';'):
                    lpn = L(r.get('Número de LPN'))
                    d = L(r.get('Instalación de destino'))
                    if not lpn or d not in TIENDAS:
                        continue
                    b = ult.get(lpn)
                    if b is not None and dia < b['visto']:
                        continue
                    if b is None or dia > b['visto']:
                        b = {'visto': dia, 'q': 0.0, 'items': collections.Counter(),
                             'ds': {}, 'desc': {}, 'ped': collections.Counter(),
                             'vacio': True}
                        ult[lpn] = b
                    b['estado'] = L(r.get('Estado de LPN'))
                    b['d'] = d
                    b['pick'] = L(r.get('Detail Picked Time'))[:10]
                    b['emp'] = L(r.get('Registro de hora de empaquetado'))[:10]
                    q = N(r.get('Cantidad empaquetada'))
                    if q > 0:
                        b['vacio'] = False
                        c = L(r.get('Código de artículo'))
                        hh = L(r.get('Registro de hora de empaquetado'))
                        if (c, hh) not in b['ds']:
                            b['ds'][(c, hh)] = 1
                            p = pares(c, q)
                            b['q'] += p
                            b['items'][c] += p
                            b['ped'][L(r.get('Número de orden'))] += p
                            b['desc'][c] = L(r.get('Descripción de artículo'))[:46]
        except OSError:
            pass

    listas = {'patio': [], 'staging': []}
    arts = {'patio': {}, 'staging': {}}
    for lpn, b in ult.items():
        if b['vacio'] or b['q'] <= 0:
            continue
        # EL ESTADO MANDA, NO EL NOMBRE DEL LPN. Ver la regla 4.
        if b['estado'] not in PARADO:
            continue
        if lpn.startswith('PRE'):
            etapa, desde = 'patio', b['pick']
        else:
            etapa, desde = 'staging', (b['emp'] or b['pick'])
        it = []
        for c, q in b['items'].most_common():
            arts[etapa].setdefault(
                c, [b['desc'].get(c, ''),
                    'F' if gen.get(s7(c), '') == 'Footwear' else 'N'])
            it.append([c, int(q)])
        ped = [o for o, _ in b['ped'].most_common()]
        t = TIENDAS[b['d']]
        listas[etapa].append({
            'l': lpn, 'o': ped[0] + ('' if len(ped) == 1 else ' +%d' % (len(ped) - 1)),
            'd': b['d'], 't': t['t'], 'z': t['z'], 'r': t['r'],
            'q': int(b['q']), 'dias': (hoy - b['visto']).days,
            'desde': desde, 'i': it})

    def tramo(n):
        if n <= 0:
            return TRAMOS[0]
        if n == 1:
            return TRAMOS[1]
        if n <= 3:
            return TRAMOS[2]
        if n <= 7:
            return TRAMOS[3]
        if n <= 14:
            return TRAMOS[4]
        return TRAMOS[5]

    salida, resumen = {}, {}
    for etapa in ('patio', 'staging'):
        listas[etapa].sort(key=lambda x: (-x['dias'], -x['q']))
        q = collections.Counter()
        n = collections.Counter()
        for b in listas[etapa]:
            q[tramo(b['dias'])] += b['q']
            n[tramo(b['dias'])] += 1
        resumen[etapa] = [{'t': t, 'q': int(q[t]), 'n': n[t]} for t in TRAMOS if n[t]]
        # LA PANTALLA LISTA SOLO LA COLA. Lo que se movio ayer es el trabajo
        # normal del dia y no aporta: lo que hay que mirar es lo que lleva un
        # dia o mas sin que nadie lo toque.
        cola = [b for b in listas[etapa] if b['dias'] >= 1]
        usados = {c for b in cola for c, _ in b['i']}
        salida[etapa] = {'filas': cola,
                         'arts': {k: v for k, v in arts[etapa].items() if k in usados}}
        v = [b for b in cola if b['dias'] >= 4]
        log('varados %-8s %4d bultos con 4 dias o mas (%d pares) - el mas viejo %d dias'
            % (etapa, len(v), sum(b['q'] for b in v),
               cola[0]['dias'] if cola else 0))
    salida['resumen'] = resumen

    # EL CONTROL DE PRE: un PRE que figura despachado no deberia existir, porque
    # tendria que haber pasado por embalaje y cambiado de LPN. Sale de aca y no
    # del archivo del ultimo dia porque un PRE despachado ya no se vuelve a
    # tocar: los dos que encontro Daniel son del 27-mar y del 7-ago.
    salida['controlPRE'] = sorted(
        [{'l': lpn, 'p': (b['ped'].most_common(1) or [('', 0)])[0][0],
          'd': b['d'], 't': TIENDAS[b['d']]['t'], 'q': int(b['q'])}
         for lpn, b in ult.items()
         if lpn.startswith('PRE') and b['estado'] in ('Cargado', 'Enviado')
         and not b['vacio'] and b['q'] > 0],
        key=lambda x: -x['q'])
    log('control: %d PRE despachados sin pasar por embalaje'
        % len(salida['controlPRE']))
    return salida


# ══ 4. EL POTENCIAL DE DESPACHO ═══════════════════════════════════════════════
def potencial(ss, fecha, gen, TIENDAS, porTienda, detTienda, en_bulto):
    """patio + staging + lo que comercial acaba de mandar a picar.

       OJO CON SUMARLOS A CIEGAS: parte del correo YA se pico, y eso ya esta
       contado en patio o en staging. La guia que aparece en un bulto no se
       vuelve a sumar."""
    arch = os.path.join(ss, 'Correos Picking',
                        'Guías %s.xlsx' % fecha.strftime('%d.%m'))
    pisa = collections.Counter()
    if not os.path.exists(arch):
        log('OJO: no esta el correo de comercial del dia (%s); el potencial '
            'sale solo con patio y staging.' % fecha.strftime('%d.%m'))
    else:
        dst = os.path.join(TEMP, '_correo.xlsx')
        shutil.copy2(arch, dst)
        wb = openpyxl.load_workbook(dst, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        h = None
        for row in ws.iter_rows(values_only=True):
            if h is None:
                h = [str(c or '').strip().upper() for c in row]
                iT = h.index('TIEND')
                iE = [i for i, x in enumerate(h) if x.startswith('ETIQUET')][0]
                iQ = [i for i, x in enumerate(h) if x.startswith('SUMA')][0]
                iG = [i for i, x in enumerate(h) if x.startswith('GUIA')][0]
                continue
            if row[iT] is None:
                continue
            # AL CODIGO DE TIENDA DEL CORREO SE LE PONE 50 DELANTE
            d = '50' + str(row[iT]).strip().split('.')[0].zfill(3)
            if d not in TIENDAS:
                continue
            q = float(row[iQ] or 0)
            g = 'F' if str(row[iE] or '').strip().upper() == 'CALZADO' else 'N'
            guia = str(row[iG] or '').strip().split('.')[0]
            if guia in en_bulto:
                pisa[g] += q
                continue
            porTienda[d]['correo' + g] += q
            b = detTienda[d].setdefault(('Correo', '', guia), {'F': 0.0, 'N': 0.0})
            b[g] += q
        wb.close()
        log('correo: %d pares ya estaban en un bulto y no se suman dos veces'
            % sum(pisa.values()))

    filas = []
    for d, v in porTienda.items():
        if sum(v.values()) <= 0:
            continue
        t = TIENDAS[d]
        filas.append({
            'd': d, 't': t['t'], 'z': t['z'], 'r': t['r'],
            'sF': int(v['stagingF']), 'sN': int(v['stagingN']),
            'pF': int(v['patioF']), 'pN': int(v['patioN']),
            'cF': int(v['correoF']), 'cN': int(v['correoN']),
            'det': [[k[0], k[1] or k[2], int(x['F'] + x['N'])]
                    for k, x in sorted(detTienda[d].items(),
                                       key=lambda kv: -(kv[1]['F'] + kv[1]['N']))
                    if x['F'] + x['N'] > 0]})
    for f in filas:
        f['tF'] = f['sF'] + f['pF'] + f['cF']
        f['tN'] = f['sN'] + f['pN'] + f['cN']
    filas.sort(key=lambda x: -(x['tF'] + x['tN']))
    log('potencial: %d tiendas - %d pares'
        % (len(filas), sum(f['tF'] + f['tN'] for f in filas)))
    return filas


# ══ EL ROBOT ══════════════════════════════════════════════════════════════════
def main():
    t0 = time.time()
    log('=' * 62)
    log('DISTRIBUCION Y DESPACHO POTENCIAL  ·  arranca %s'
        % datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S'))
    log('=' * 62)
    limpiar_logs()
    base = base_onedrive()
    if not base:
        raise SystemExit('No encuentro la carpeta de OneDrive.')
    ss = os.path.join(base, 'scraping Stock')
    if not os.path.isdir(TEMP):
        os.makedirs(TEMP)

    archivos, ultimo, fecha = elegir_dia(ss)
    log('dia %s - %d archivos de OBLPN' % (fecha.strftime('%d-%m-%Y'), len(archivos)))

    gen = leer_maestro(base)
    TIENDAS = leer_rutas(base)

    # LA FOTO VA PRIMERO: el cuadro grande usa sus cuatro columnas de despacho.
    listas, arts, turnoZona, porTienda, detTienda, en_bulto, estados = \
        foto_del_dia(ultimo, gen, TIENDAS)
    tabla = cuadro_retail(ss, gen, TIENDAS, fecha, estados)
    var = varados(archivos, gen, TIENDAS, fecha)
    control = var.pop('controlPRE')
    pot = potencial(ss, fecha, gen, TIENDAS, porTienda, detTienda, en_bulto)

    f_txt = fecha.strftime('%d-%m-%Y')

    # ── LO QUE VIAJA CON LA PANTALLA VA SIN EL DESGLOSE POR ARTICULO ──
    # Son 2.400 bultos y el detalle triplica el peso. La pantalla se abre cien
    # veces al dia; el Excel se pide una.
    liviano = lambda fs: [{k: f[k] for k in ('o', 'l', 'd', 't', 'f', 'q') if k in f}
                          for f in fs]
    distribucion = {
        'fecha': f_txt,
        'generado': 'foto del OBLPN %s' % fecha.strftime('%d-%m'),
        'aviso': ('Las columnas de picking salen del archivo de picking del %s. '
                  '<b>Qty patio, Staging, Cargado y Enviado son la foto del OBLPN '
                  'de ese mismo día</b>, así que esas cuatro cambian cuando baje '
                  'el archivo siguiente.' % fecha.strftime('%d-%m')),
        'tabla': tabla,
        'turnoZona': turnoZona,
        'listas': {'patio': liviano(listas['patio']),
                   'staging': liviano(listas['staging'])},
        'varados': var,
        'controlPRE': control,
    }
    detalle = {
        'fecha': f_txt,
        'arts': arts,
        'patio': {f['l']: f['i'] for f in listas['patio'] if f.get('i')},
        'staging': {f['l']: f['i'] for f in listas['staging'] if f.get('i')},
    }
    despacho = {'fecha': f_txt, 'filas': pot}

    # SI LA FOTO SALE VACIA NO SE PUBLICA. Este almacen nunca tiene patio y
    # staging los dos en cero: si pasa, el archivo esta a medias y publicarlo
    # borraria la pantalla. Vale mas dejar el dia anterior a la vista.
    hay = sum(f['patio'] + f['stg'] for f in tabla)
    if hay <= 0:
        log('la foto del %s sale sin un solo bulto en patio ni en staging: '
            'NO SE PUBLICA nada y queda el dia anterior.' % f_txt, 'ERROR')
        return 1

    # `--probar` calcula y deja los tres JSON al lado, sin publicar. Sirve para
    # revisar los numeros antes de que los vea nadie.
    probar = '--probar' in sys.argv
    ok = True
    for area, datos in (('distribucion_dia', distribucion),
                        ('distribucion_detalle', detalle),
                        ('despacho_potencial_dia', despacho)):
        crudo = json.dumps(datos, ensure_ascii=False, separators=(',', ':'))
        n = len(crudo.encode('utf-8'))
        if probar:
            io.open(os.path.join(AQUI, '_prueba_%s.json' % area), 'w',
                    encoding='utf-8').write(crudo)
            log('%-24s calculado, SIN publicar  (%.0f KB)' % (area, n / 1024))
            continue
        # CADA AREA VA POR SEPARADO: si una falla, las otras ya quedaron puestas.
        if publicar_area.publicar(area, datos, 'MASTER'):
            log('%-24s publicado  (%.0f KB)' % (area, n / 1024))
        else:
            log('%-24s *** NO SE PUDO PUBLICAR ***' % area)
            ok = False

    try:
        shutil.rmtree(TEMP)
    except OSError:
        pass
    log('listo en %.1f minutos · termina %s'
        % ((time.time() - t0) / 60.0,
           datetime.datetime.now().strftime('%H:%M:%S')))
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
