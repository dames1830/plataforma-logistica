# -*- coding: utf-8 -*-
"""
ROBOT: el cuadro de PICKING POR DIA (persona x hora, canal y efectividad)

Va ENGANCHADO DETRAS del robot que ya baja el archivo de picking del WMS cada 2 horas
-`ejecutar_picking_hora.bat`-, asi que NO entra al WMS ni descarga nada: lee el archivo
que ese acaba de dejar y publica el cuadro. Daniel, 02-sep-2026: *"ese picking por
hora es el que tienes que agarrar para el modulo de picking dia"*.

Por eso no pide turno al candado del WMS y no puede chocar con ninguna otra
corrida: lo unico que hace es leer un CSV de OneDrive y mandar un JSON.

SE LLAMA `produccion_picking.py` Y NO `picking_por_hora.py` A PROPOSITO: en el servidor ya
existe un `picking_por_hora.py` -el que baja del WMS-, y dos archivos con el
mismo nombre en la carpeta del robot se pisan. El AREA de la plataforma si se llama
`picking_por_hora`, que es lo que leen las pantallas.

Publica en el area `picking_por_hora`, fechada con el DIA DEL ARCHIVO -no con la
fecha de hoy-: el ultimo pase del dia es a las 20:20 y con la hora de la maquina
quedaria estampado el dia siguiente.

PERSONA x HORA DEL PICKING: volumen, efectividad y CANAL.

SIN LA EQUIVALENCIA DEL PREPACK. Daniel, 01-sep-2026: *"no quiero que utilices
la equivalencia del prepack todavia, eso todavia lo veo por comite"*. El suelto y
el prepack se miden POR SEPARADO, cada uno con sus propias lineas y su tiempo.

EL RITMO NO ES EL VOLUMEN. A quien le dan nueve tareas pica mas que quien recibio
una, y eso no dice quien es mas rapido. La efectividad divide las lineas por el
tiempo que esa persona estuvo trabajando EN ESA CLASE -del primer al ultimo
pick-, asi que no depende de cuanto trabajo le tocara.

EL CANAL, con la regla que valido Daniel el 01-sep-2026:
  · SI EL DESTINO ESTA EN EL MAESTRO DE RUTAS, ES TIENDA RETAIL. Sale del propio
    archivo de picking -`Instalacion de destino`-, sin cruzar nada, y cubre el
    100% de las lineas. El codigo pega TAL CUAL: el "50 delante" es para el
    `TIEND` del correo de comercial, no para esto.
  · SI NO ESTA, el canal fino sale del `Tipo de orden` del Detalle de Orden.
    Medido el 31-08: las dos formas coinciden en 99,4%, y donde no coinciden gana
    el maestro de rutas -las 68 lineas sin canal SI eran tienda-.
  · Cada canal que no es tienda tiene su destino fijo: 91891 catalogo, 91890
    tienda virtual, 93173/92458 ecommerce, 81439 industrial, 0019620xx mayorista.

EL CANAL FINO NECESITA TRES FUENTES. `Detalle Orden DD-MM.csv` trae las ordenes
NACIDAS ese dia, no las picadas: pegaba 171 de 1.166. Juntando Pendientes +
Despachados + los diarios se llega al 99,4%.

LAS CUATRO TRAMPAS DEL ARCHIVO, iguales a js/reportes/picking.js:
  1. Estado 'Cancelado' es una COPIA de la linea buena. Solo cuenta 'Finalizada'.
  2. No todo lo que sale son pares: lo corta el Maestro con G. Gender=Footwear.
  3. El prepack cuenta CAJAS: los pares salen de la curva del propio SKU.
  4. El dia sale del CONTENIDO, no del nombre del archivo.
"""
import csv
import io
import json
import os
import re
import shutil
import sys
from collections import defaultdict

import openpyxl

csv.field_size_limit(10 ** 7)

FORMA_PREPACK = re.compile(r'^\d{7}-\d-\d{5}$')
# LAS VEINTICUATRO HORAS, no solo el turno.
#
# Estaba en 8..19 y lo que se movia fuera de ahi se contaba en el total pero NO
# tenia columna donde aparecer: el 28-ago quedaron 275 lineas de picking y 217 de
# embalaje sin fila que las mostrara, y la suma de las horas no daba el total.
#
# Daniel, 02-sep-2026: *"el noventa y cinco por ciento se mueve entre ocho de la
# manana y las siete de la noche, pero hay un minimo que se mueve en la madrugada.
# Necesito las veinticuatro horas"*.
#
# La pantalla no dibuja las 24 siempre: muestra el turno completo y agrega solo
# las horas de afuera que ese dia tuvieron movimiento.
HORAS = list(range(0, 24))
CL = ('cal_suelto', 'cal_prepack', 'no_cal', 'materiales', 'sin_tipo')
TODOS = 'TODOS'
ORDEN_CANAL = ['RETAIL', 'MAYORISTA', 'CATALOGO', 'ECOMMERCE', 'INDUSTRIAL',
               'OTROS', 'SIN CANAL']

# CON DOS PICKS NO SE MIDE UN RITMO, y tampoco alcanza con contar lineas:
# karteaga cerro 37 lineas en 36 SEGUNDOS el 31-08 y salia primero con 4.667
# lineas/hora. Eso no es alguien picando, es una confirmacion en bloque del WMS.
LINEAS_MIN_CELDA = 8
LINEAS_MIN_DIA = 20
SEG_MIN_CELDA = 5 * 60
SEG_MIN_DIA = 15 * 60
SEG_LINEA_MIN = 5
SEG_MUESTRA_CORTA = 60 * 60
# DOS MARCAS PEGADAS SON LA MISMA TANDA. En embalaje el WMS estampa un solo
# instante por caja, asi que sin este puente cada tarea duraria cero y el
# ritmo se iria al cielo. Un hueco mayor que esto si es tiempo parado.
PUENTE_SEG = 15 * 60


def base_onedrive():
    for c in (os.environ.get('OneDrive'),
              'C:' + os.sep + os.path.join('Users', 'Administrator', 'OneDrive'),
              'C:' + os.sep + os.path.join('Users', 'dames', 'OneDrive')):
        if c:
            r = os.path.join(c, 'danielames.bata', 'scraping Stock')
            if os.path.isdir(r):
                return r
    raise SystemExit('no encuentro OneDrive')


def dia_pedido():
    """El dia que vino por `--dia AAAA-MM-DD`, o None.

    EN EL RELLENO EL DIA LO MANDA QUIEN LLAMA. Los archivos viejos del OBLPN
    mezclan hasta doce fechas —el WMS mete lineas empaquetadas antes— y la
    mayoria no siempre es la del archivo: el `OBLPN 01-08.csv` daba 30 de julio.
    El que rellena si sabe de que dia es cada archivo, porque lo dice el nombre.

    En la corrida normal de cada 2 horas no se pasa nada y sigue mandando la
    mayoria de las filas, que ahi es lo correcto.
    """
    a = sys.argv[1:]
    if '--dia' in a:
        i = a.index('--dia')
        if len(a) > i + 1 and re.match(r'^\d{4}-\d{2}-\d{2}$', a[i + 1]):
            return a[i + 1]
    return None


def dia_mayoritario(ruta, columna):
    """EL DIA DEL ARCHIVO ES EL DE LA MAYORIA DE SUS FILAS, no el de la primera.

    Salia de la primera fila con hora valida, y eso es fragil: el OBLPN del
    01-09 traia arriba una linea empaquetada el 31-08 y el cuadro entero quedo
    fechado el 31, tirando las 39.000 filas del dia bueno a `sin hora`. Se vio el
    02-sep-2026 en la primera corrida del robot.

    Se cuenta con una pasada liviana -solo esa columna, sin armar diccionarios-
    y gana la fecha que mas veces aparece. Devuelve AAAA-MM-DD, o None si el
    archivo no tiene ni una fecha legible.
    """
    from collections import Counter
    try:
        f = io.open(ruta, encoding='utf-8-sig', newline='', errors='replace')
        cabeza = f.read(4000)
        f.seek(0)
        r = csv.reader(f, delimiter=';' if cabeza.count(';') > cabeza.count(',') else ',')
        cab = [c.strip() for c in next(r)]
        try:
            i = cab.index(columna)
        except ValueError:
            f.close()
            return None
        cuenta = Counter()
        for x in r:
            if i < len(x):
                mm = re.match(r'^(\d{2})/(\d{2})/(\d{4})\s', str(x[i]).strip().strip('"'))
                if mm:
                    cuenta['%s-%s-%s' % (mm.group(3), mm.group(2), mm.group(1))] += 1
        f.close()
        if not cuenta:
            return None
        gana, veces = cuenta.most_common(1)[0]
        if len(cuenta) > 1:
            print('[AVISO] el archivo mezcla %d fechas; gana %s con %d filas de %d'
                  % (len(cuenta), gana, veces, sum(cuenta.values())))
        return gana
    except Exception as e:
        print('[AVISO] no se pudo mirar la fecha del archivo: %s' % e)
        return None


def elegir_archivo(carpetas, plantillas):
    """QUE ARCHIVO LE TOCA A ESTA CORRIDA.

    Se puede pasar el nombre a mano -util para rehacer un dia viejo-. Sin eso,
    se busca el de HOY, que es el que el robot de la hora acaba de dejar.

    EL NOMBRE DEL DIA NO ES UNO SOLO: el picking lo escribe sin cero adelante
    ("Picking 3-9.csv") y el OBLPN con cero ("OBLPN 03-09.csv"). Se prueban las
    dos formas antes de rendirse.

    Si el de hoy no esta -el WMS no contesto, o es de madrugada y todavia no
    corrio ningun pase- se toma EL MAS NUEVO de la carpeta y se avisa. Es mejor
    republicar el cuadro de ayer que dejar la pantalla sin nada.
    """
    if isinstance(carpetas, str):
        carpetas = [carpetas]
    if len(sys.argv) > 1 and not sys.argv[1].startswith('-'):
        for c in carpetas:
            r = os.path.join(c, sys.argv[1])
            if os.path.isfile(r):
                return r
        return os.path.join(carpetas[0], sys.argv[1])
    hoy = __import__('datetime').datetime.now()
    for c in carpetas:
        for pl in plantillas:
            r = os.path.join(c, pl % (hoy.day, hoy.month))
            if os.path.isfile(r):
                return r
    cand = []
    for c in carpetas:
        try:
            cand += [os.path.join(c, n) for n in os.listdir(c)
                     if n.lower().endswith('.csv')]
        except Exception:
            pass
    if cand:
        nuevo = max(cand, key=os.path.getmtime)
        print('[AVISO] no hay archivo de hoy; se usa el mas nuevo: %s'
              % os.path.basename(nuevo))
        return nuevo
    return os.path.join(carpetas[0], plantillas[0] % (hoy.day, hoy.month))

BASE = base_onedrive()
# PRIMERO LA COPIA DEL ROBOT DE LA HORA, DESPUES LA DE ONEDRIVE.
#
# `picking_por_hora.py` baja el picking del dia cada 2 horas y deja una copia en
# `logs\picking_hora`. Esa es la de HOY y la que hay que mirar.
#
# La carpeta de OneDrive la escribe otro robot -el de las 19:20- y trae el
# picking de AYER: sirve de respaldo, para que la pantalla no quede vacia si el
# pase de la hora no salio, pero no es la primera opcion.
ARCHIVO = elegir_archivo(
    [os.path.join('C:' + os.sep, 'wms_scraping', 'logs', 'picking_hora'),
     os.path.join(BASE, 'Picking')],
    ['Picking %d-%d.csv', 'Picking %02d-%02d.csv'])
CARPETA_ORD = os.path.join(BASE, 'Detalle Orden')
MAESTROS = [os.path.join(os.path.dirname(BASE), 'Maestro_Articulos.xlsx'),
            os.path.join(BASE, 'Archivos', 'Maestro_Articulos.xlsx')]
RUTAS_CAND = [os.path.join(os.path.dirname(BASE), 'Proyecto web Logistico',
                           'RUTAS -  TURNOS.xlsx'),
              os.path.join('C:' + os.sep, 'wms_scraping', '_rutas.xlsx')]
# EL AREA Y EL ARCHIVO SON LO MISMO, escrito una sola vez. Estuvieron sueltos y
# el robot de embalaje quedo escribiendo en el JSON del picking: el cruce comparo
# el web report de picking contra los numeros del embalaje y dio 3.480 de
# diferencia sin que nada pareciera roto.
AREA = 'picking_por_hora'
SALIDA = os.path.join('C:' + os.sep, 'wms_scraping', 'logs', AREA + '.json')


def limpio(v):
    s = str(v if v is not None else '').strip()
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1].strip()
    return s.strip('"').strip()


def entero(v):
    try:
        return int(float(str(v).replace(',', '.')))
    except (TypeError, ValueError):
        return 0


def es_prepack(sku):
    return bool(FORMA_PREPACK.match(str(sku or '').strip()))


# ── LOS TIPOS QUE PIDIO DANIEL ───────────────────────────────────────────────
#
# Daniel, 02-sep-2026: *"solamente calzado y no calzado. En no calzado entra todo
# lo que son bolsas, etiquetas, etcetera"*. Y aparte, para lo que el Maestro no
# conoce: *"ponle materiales porque si son solamente cinco digitos, es material,
# si no me equivoco. Revisa la descripcion que es"*.
#
# SE REVISO ANTES DE CREER LA CORAZONADA, leyendo la descripcion que trae el
# propio archivo del WMS. Los 25 codigos que no estaban en el Maestro son TODOS
# de cinco digitos y TODOS material:
#
#     70104   74.768 u   TISSUE PAPER BATA N 4
#     69050   18.695 u   HANG TAG ORTHOLITE BATA ROJO
#     70103   16.795 u   TISSUE PAPER BATA N 3
#     88424    2.575 u   CAJA MICROC. KRAFT BATA N.24
#     26036      540 u   PLANT LAURA C/GEL ROJO N. 36
#
# Papel de seda, etiquetas colgantes, cajas de carton y plantillas. Tenia razon.
#
# LA REGLA EXIGE LAS DOS COSAS -cinco digitos Y no estar en el Maestro- y no solo
# la primera. Si manana falta en el Maestro un articulo de calzado de verdad, cae
# en "sin tipo" y se ve; llamarlo "materiales" sin conocerlo seria inventar.
# `sin_tipo` es la respuesta honesta de "no se que es esto", y ademas es la lista
# de lo que hay que agregarle al Maestro.
CINCO_DIGITOS = re.compile(r'^\d{5}$')


def tipo_de(sku, gender, esta_en_el_maestro):
    """calzado suelto / calzado prepack / no calzado / materiales / sin tipo."""
    if not esta_en_el_maestro:
        return 'materiales' if CINCO_DIGITOS.match((sku or '')[:7]) else 'sin_tipo'
    if not gender or gender == 'Sin dato':
        return 'sin_tipo'
    if gender != 'Footwear':
        return 'no_cal'
    return 'cal_prepack' if es_prepack(sku) else 'cal_suelto'


def pares_de_la_caja(sku):
    s = str(sku or '').strip()
    if not FORMA_PREPACK.match(s):
        return 1
    try:
        n = int(s[-5:][:2])
    except ValueError:
        return 1
    return n if 0 < n <= 24 else 1


def abrir(ruta):
    f = io.open(ruta, encoding='utf-8-sig', newline='', errors='replace')
    cabeza = f.read(4000)
    f.seek(0)
    sep = ';' if cabeza.count(';') > cabeza.count(',') else ','
    return f, csv.DictReader(f, delimiter=sep)


# ── el Maestro de articulos ─────────────────────────────────────────────
ruta_m = next((r for r in MAESTROS if os.path.isfile(r)), None)
wb = openpyxl.load_workbook(ruta_m, read_only=True, data_only=True)
it = wb.worksheets[0].iter_rows(values_only=True)
cab = [str(c).strip() if c is not None else '' for c in next(it)]


def col(*nombres):
    bajos = [n.lower() for n in nombres]
    for i, c in enumerate(cab):
        if c.lower() in bajos:
            return i
    return -1


iS = col('CodArticulo', 'CodigoArticulo')
iG, iM, iC = col('G. Gender', 'G Gender'), col('Marcas', 'MarcaStd'), col('Coleccion PO')
maestro = {}
for f in it:
    if iS < 0 or iS >= len(f) or f[iS] is None:
        continue
    k = limpio(f[iS])[:7]
    if k and k not in maestro:
        def d(i):
            v = limpio(f[i]) if 0 <= i < len(f) else ''
            return v if v and v != '(en blanco)' else 'Sin dato'
        maestro[k] = (d(iG), d(iM), d(iC))
wb.close()
print('maestro de articulos: %d codigos' % len(maestro))

# ── el maestro de RUTAS: quien es tienda ────────────────────────────────
# SE COPIA ANTES DE ABRIRLO: en OneDrive esta solo en la nube y openpyxl lo ve
# como un zip roto. Copiarlo lo baja.
ruta_r = next((r for r in RUTAS_CAND if os.path.isfile(r)), None)
copia = os.path.join('C:' + os.sep, 'wms_scraping', 'logs', '_rutas_pph.xlsx')
try:
    shutil.copyfile(ruta_r, copia)
except Exception:
    copia = ruta_r
wb = openpyxl.load_workbook(copia, read_only=True, data_only=True)
it = wb.worksheets[0].iter_rows(values_only=True)
cr = [str(c).strip() if c is not None else '' for c in next(it)]
kC = cr.index('CDG')
tiendas = {str(f[kC]).strip() for f in it if kC < len(f) and f[kC] is not None}
wb.close()
print('maestro de rutas: %d tiendas' % len(tiendas))

# ── el canal fino, de las tres fuentes ──────────────────────────────────
tipo_orden = {}


def tragar(ruta2):
    if not os.path.isfile(ruta2):
        return
    try:
        f3, r3 = abrir(ruta2)
    except OSError:
        return                      # solo en la nube: OneDrive no lo bajo
    for x in r3:
        o = limpio(x.get('Número de orden'))
        if o and o not in tipo_orden:
            tipo_orden[o] = limpio(x.get('Tipo de orden'))
    f3.close()


tragar(os.path.join(CARPETA_ORD, 'Detalle Orden Pendientes.csv'))
tragar(os.path.join(CARPETA_ORD, 'Detalle Orden Despachados.csv'))
for n in sorted((n for n in os.listdir(CARPETA_ORD)
                 if re.match(r'^Detalle Orden \d{2}-\d{2}\.csv$', n)), reverse=True):
    tragar(os.path.join(CARPETA_ORD, n))
print('tipo de orden conocido para %s ordenes' % '{:,}'.format(len(tipo_orden)))


def canal_de(destino, orden):
    """EL MAESTRO DE RUTAS MANDA. Si el destino es tienda, es retail y no se
    consulta nada mas; el Tipo de orden solo afina lo que NO es tienda."""
    if destino in tiendas:
        return 'RETAIL'
    t = (tipo_orden.get(orden) or '').upper()
    if not t:
        return 'SIN CANAL'
    if 'MAYOR' in t:
        return 'MAYORISTA'
    if 'CATALOGO' in t:
        return 'CATALOGO'
    if 'ECOMMERCE' in t or 'VIRTUAL' in t:
        return 'ECOMMERCE'
    if 'INDUSTRIAL' in t:
        return 'INDUSTRIAL'
    return 'OTROS'


# ── el archivo de picking ───────────────────────────────────────────────
f, r = abrir(ARCHIVO)
# (canal, persona, hora, clase). El canal TODOS se llena a la par, para no tener
# que sumar seis diccionarios despues.
cel = defaultdict(lambda: defaultdict(float))
# (canal, persona, hora, clase) -> {tarea: [segundos]}. Hace falta la
# TAREA para poder armar los tramos; una lista suelta no distingue el
# rato trabajado del rato parado.
sellos = defaultdict(lambda: defaultdict(list))

# ══════════════════════════════════════════════════════════════════════════════
# EL TIEMPO PARA MEDIR PRODUCTIVIDAD: LO QUE DICE EL WMS, SUMADO Y NADA MAS
# ══════════════════════════════════════════════════════════════════════════════
#
# Daniel, 02-sep-2026: *"tu deberias sacar lo que dice el WMS, no te debieras de
# inventar nada. La tarea uno se hizo de nueve a diez, ahi son sesenta minutos. La
# segunda de diez y diez a las once, cincuenta. En total ciento diez"*. Y sobre
# las que se pisan: *"si la otra tarea esta dentro de esa tarea, tu sigue
# acumulando el dato nada mas. Seis minutos, sumale. No importa si esta dentro o
# afuera"*.
#
# LA REGLA, entonces, es la mas simple que hay:
#   cada tarea aporta (ultimo pick - primer pick), y se suman todas.
#   Sin puente de 15 minutos -eso me lo habia inventado yo- y sin descontar los
#   solapes.
#
# SOLO LAS LINEAS CON `Numero de tarea` DE VERDAD. El 34,4% de las lineas no lo
# trae y hasta hoy se les ponia el numero de contenedor como apaño; eso partia una
# tarea en varias y hacia que el 48,7% "se pisara" — un solape inventado por mi
# manera de agrupar, no del almacen. Con los numeros de tarea reales solo se pisa
# el 1,6%. Daniel: *"dejalas fuera"*.
#
# POR ESO VAN PARES APARTE (`_q`): si se descartan esas lineas del tiempo hay que
# descartarlas TAMBIEN de los pares, o el ritmo sale inflado. Son el 20,4% de los
# pares. Los totales del dia -los que ve Picking por dia- NO se tocan: siguen
# contando todo.
sellos_tarea = defaultdict(lambda: defaultdict(list))   # (k,usr,h,clase) -> tarea -> [seg]
pares_tarea = defaultdict(float)                        # (k,usr,h,clase) -> pares
lineas_ph = defaultdict(lambda: defaultdict(float))
marcas = defaultdict(lambda: defaultdict(float))
colec = defaultdict(lambda: defaultdict(float))
zonas = defaultdict(lambda: defaultdict(float))
zonas_ubi = defaultdict(set)
totales = defaultdict(lambda: defaultdict(float))
personas = defaultdict(set)
tipos_vistos = defaultdict(lambda: defaultdict(float))
destinos = defaultdict(lambda: defaultdict(float))
dia = dia_pedido() or dia_mayoritario(ARCHIVO, 'Hora de selección')
leidas = descartadas = 0

for row in r:
    if limpio(row.get('Estado')) != 'Finalizada':
        descartadas += 1
        continue
    leidas += 1
    sku = limpio(row.get('Código de artículo'))
    usr = limpio(row.get('Usuario de selección')) or '(sin usuario)'
    ubi = limpio(row.get('De ubicación')) or '?'
    orden = limpio(row.get('Número de orden'))
    destino = limpio(row.get('Instalación de destino'))
    hs = limpio(row.get('Hora de selección'))
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2}):(\d{2})$', hs)
    if not m:
        continue
    if dia is None:
        dia = '%s-%s-%s' % (m.group(3), m.group(2), m.group(1))
    elif not hs.startswith('%s/%s/%s' % (dia[8:], dia[5:7], dia[:4])):
        # UN CUADRO, UN SOLO DIA. Si el archivo mezcla dos fechas, las del otro
        # dia se van: sumarlas daria un turno de 24 horas que nadie trabajo.
        descartadas += 1
        continue
    h = int(m.group(4))
    seg = h * 3600 + int(m.group(5)) * 60 + int(m.group(6))

    pares = entero(row.get('Cantidad empaquetada')) * pares_de_la_caja(sku)
    g, mar, cl = maestro.get(sku[:7], ('Sin dato', 'Sin dato', 'Sin dato'))
    clase = tipo_de(sku, g, sku[:7] in maestro)
    can = canal_de(destino, orden)
    z = ubi.split('-')[0]
    tipos_vistos[can][tipo_orden.get(orden) or '(sin dato)'] += 1
    destinos[can][destino] += 1

    # LA TAREA: la del WMS, y si falta el contenedor. El 39% de las lineas no
    # trae numero de tarea; el contenedor viene siempre.
    tarea = (limpio(row.get('Número de tarea'))
             or 'C:' + limpio(row.get('Número de contenedor')))
    # el numero de tarea DE VERDAD, sin el apaño del contenedor
    tarea_real = limpio(row.get('Número de tarea'))

    for k in (can, TODOS):
        personas[k].add(usr)
        cel[(k, usr, h, clase)]['pares'] += pares
        cel[(k, usr, h, clase)]['lineas'] += 1
        sellos[(k, usr, h, clase)][tarea].append(seg)
        sellos[(k, usr, None, clase)][tarea].append(seg)
        if tarea_real:
            sellos_tarea[(k, usr, h, clase)][tarea_real].append(seg)
            sellos_tarea[(k, usr, None, clase)][tarea_real].append(seg)
            pares_tarea[(k, usr, h, clase)] += pares
            pares_tarea[(k, usr, None, clase)] += pares
        sellos[(k, usr, h, 'total')][tarea].append(seg)
        sellos[(k, usr, None, 'total')][tarea].append(seg)
        lineas_ph[(k, h)][clase] += pares
        lineas_ph[(k, h)]['lineas'] += 1
        marcas[(k, mar)][clase] += pares
        marcas[(k, mar)]['lineas'] += 1
        colec[(k, cl)][clase] += pares
        colec[(k, cl)]['lineas'] += 1
        zonas[(k, z)][clase] += pares
        zonas[(k, z)]['lineas'] += 1
        zonas_ubi[(k, z)].add(ubi)
        totales[k][clase] += pares
        totales[k]['lineas'] += 1
f.close()


def vol(d):
    o = {c: int(round(d.get(c, 0))) for c in CL}
    o['lineas'] = int(d.get('lineas', 0))
    # EL TOTAL SE SUMA SOBRE `CL`, no sobre tres nombres escritos a mano. Al
    # agregar `materiales` y `sin_tipo` la suma vieja los dejaba afuera y el
    # cuadro no cuadraba, sin una sola queja.
    o['total'] = sum(o[c] for c in CL)
    return o


def ritmo(marcas_t, lineas, minimo, span_min):
    """Lineas por hora sobre el tiempo REALMENTE trabajado en esa clase.

    None NO es cero: es "no alcanza la muestra", y en pantalla va como una raya.
    """
    if not marcas_t or len(marcas_t) < 2:
        return None, None, None, False
    span = max(marcas_t) - min(marcas_t)
    mins = round(span / 60.0, 1)
    if lineas < minimo or span < span_min:
        return None, None, mins, False
    sl = span / float(lineas - 1)
    if sl < SEG_LINEA_MIN:
        return None, None, mins, False      # confirmacion en bloque, no una persona
    return (int(round(lineas / (span / 3600.0))), int(round(sl)), mins,
            span < SEG_MUESTRA_CORTA)


def minutos_sumados(por_tarea):
    """Los minutos de trabajo: (ultimo pick - primer pick) de CADA tarea, sumados.

    LA REGLA QUE PIDIO DANIEL, y no tiene vuelta: cada tarea aporta lo suyo y se
    suman todas. Si dos se pisan, se suman igual — *"no importa si esta dentro o
    afuera de la tarea"*.

    UNA TAREA DE UN SOLO PICK APORTA CERO, y eso es fiel al dato: el WMS no dice
    cuanto duro, dice cuando se pico. Inventarle una duracion seria volver a lo
    que se acaba de sacar.
    """
    if not por_tarea:
        return 0
    return sum(max(v) - min(v) for v in por_tarea.values() if len(v) > 1)


def tramos(por_tarea):
    """Los ratos en que esa persona estuvo trabajando, ya fusionados.

    UNIR, NO SUMAR. Las tareas se solapan -varios contenedores en un mismo
    recorrido-, asi que sumar sus duraciones cuenta el mismo minuto dos veces y
    da mas horas que las del dia. Fusionando, lo que queda entre tramo y tramo
    es tiempo parado: el refrigerio sale solo, sin descontarlo a mano.
    """
    if not por_tarea:
        return []
    ts = sorted((min(v), max(v)) for v in por_tarea.values() if v)
    if not ts:
        return []
    fus = [list(ts[0])]
    for a, b in ts[1:]:
        # se pisan, o estan lo bastante pegados como para ser la misma tanda
        if a <= fus[-1][1] + PUENTE_SEG:
            fus[-1][1] = max(fus[-1][1], b)
        else:
            fus.append([a, b])
    return fus


def celda(can, usr, h):
    """Una celda: el volumen de las tres clases y EL PRIMER Y ULTIMO PICK.

    NO SE PUBLICA EL RITMO YA CALCULADO. La pantalla deja elegir varios canales a
    la vez, y los ritmos no se suman: hay que rehacerlos sobre el conjunto. Con
    el minimo de los minimos y el maximo de los maximos, mas las lineas sumadas,
    el guion saca exactamente el mismo numero que sacaria aca.
    """
    o = {}
    tot_l = 0
    for c in CL + ('total',):
        if c == 'total':
            o['total'] = sum(o[c] for c in CL)
            o['lineas'] = tot_l
            o['total_l'] = tot_l
        else:
            if h is None:
                d = {'pares': sum(cel.get((can, usr, y, c), {}).get('pares', 0)
                                  for y in HORAS),
                     'lineas': sum(cel.get((can, usr, y, c), {}).get('lineas', 0)
                                   for y in HORAS)}
            else:
                d = cel.get((can, usr, h, c), {})
            n = int(d.get('lineas', 0))
            tot_l += n
            o[c] = int(round(d.get('pares', 0)))
            o[c + '_l'] = n
        o[c + '_iv'] = tramos(sellos.get((can, usr, h, c)))
        # PARA LA PRODUCTIVIDAD: segundos sumados y pares, solo de las lineas con
        # numero de tarea de verdad. Ver el comentario de `sellos_tarea`.
        if c != 'total':
            o[c + '_s'] = minutos_sumados(sellos_tarea.get((can, usr, h, c)))
            o[c + '_q'] = int(round(pares_tarea.get((can, usr, h, c), 0)))
    return o


def vista(can):
    ph = {}
    for h in HORAS:
        v = vol(lineas_ph.get((can, h), {}))
        v['personas'] = sum(1 for u in personas[can]
                            if any((can, u, h, c) in cel for c in CL))
        ph[str(h)] = v
    return {
        'totales': vol(totales[can]),
        'por_hora': ph,
        'gente': sorted([{'usuario': u, 'total': celda(can, u, None),
                          # SOLO LAS HORAS EN QUE ESA PERSONA MOVIO ALGO.
                          # Con las 24 completas el archivo del dia pasaba de 368
                          # a 585 KB —un mes serian 35 MB y bajar un rango de
                          # treinta dias, 17 MB al navegador— y veinte de esas
                          # veinticuatro celdas venian en cero. La pantalla trata
                          # la hora que falta como vacia.
                          'horas': {str(h): c for h, c in
                                    ((h, celda(can, u, h)) for h in HORAS)
                                    if c.get('total')}}
                         for u in personas[can]],
                        key=lambda x: -x['total']['total']),
        'marcas': sorted([dict(nom=k[1], **vol(v)) for k, v in marcas.items()
                          if k[0] == can], key=lambda x: -x['total'])[:14],
        'coleccion': sorted([dict(nom=k[1], **vol(v)) for k, v in colec.items()
                             if k[0] == can], key=lambda x: -x['total'])[:14],
        # LAS UBICACIONES VAN COMO LISTA, no solo contadas: al juntar dos canales
        # en la pantalla hay que UNIRLAS, y sumar los conteos las cuenta doble
        # -la misma ubicacion la visitan los dos-.
        'zonas': sorted([dict(nom=k[1], ubicaciones=len(zonas_ubi[k]),
                              ubis=sorted(zonas_ubi[k]), **vol(v))
                         for k, v in zonas.items() if k[0] == can],
                        key=lambda x: -x['total']),
    }


con_datos = [c for c in ORDEN_CANAL if totales.get(c, {}).get('lineas')]
salida = {
    'dia': dia,
    'archivo': os.path.basename(ARCHIVO),
    'lineas_buenas': leidas,
    'lineas_descartadas': descartadas,
    'horas': HORAS,
    'cortes': {'lineasCelda': LINEAS_MIN_CELDA, 'lineasDia': LINEAS_MIN_DIA,
               'minutosCelda': SEG_MIN_CELDA // 60, 'minutosDia': SEG_MIN_DIA // 60,
               'segLineaMin': SEG_LINEA_MIN, 'puenteMin': PUENTE_SEG // 60, 'muestraCortaMin': SEG_MUESTRA_CORTA // 60},
    'canales': [TODOS] + con_datos,
    'gentePorCanal': {c: len(personas[c]) for c in [TODOS] + con_datos},
    'tiposPorCanal': {c: sorted([[k, int(v)] for k, v in tipos_vistos[c].items()],
                                key=lambda x: -x[1])[:6] for c in con_datos},
    'destinosPorCanal': {c: sorted([[k, int(v)] for k, v in destinos[c].items()],
                                   key=lambda x: -x[1])[:6] for c in con_datos},
    'vistas': {c: vista(c) for c in [TODOS] + con_datos},
}

os.makedirs(os.path.dirname(SALIDA), exist_ok=True)
io.open(SALIDA, 'w', encoding='utf-8').write(json.dumps(salida, ensure_ascii=False))


def mm(n):
    return '{:,}'.format(int(n))


T = salida['vistas'][TODOS]['totales']
print('')
print('DIA %s  -  %s lineas buenas  -  %s copias descartadas'
      % (dia, mm(leidas), mm(descartadas)))
print('TOTAL  suelto %s  -  prepack %s  -  no calzado %s  =  %s pares'
      % (mm(T['cal_suelto']), mm(T['cal_prepack']), mm(T['no_cal']), mm(T['total'])))
print('')
print('POR CANAL')
print('  %-12s %9s %9s %10s %10s %9s %7s'
      % ('CANAL', 'LINEAS', 'PARES', 'SUELTO', 'PREPACK', 'NO CALZ', 'GENTE'))
suma_l = 0
for c in con_datos:
    v = salida['vistas'][c]['totales']
    suma_l += v['lineas']
    print('  %-12s %9s %9s %10s %10s %9s %7d'
          % (c, mm(v['lineas']), mm(v['total']), mm(v['cal_suelto']),
             mm(v['cal_prepack']), mm(v['no_cal']), len(personas[c])))
print('  %-12s %9s %9s %10s %10s %9s %7d'
      % ('TODOS', mm(T['lineas']), mm(T['total']), mm(T['cal_suelto']),
         mm(T['cal_prepack']), mm(T['no_cal']), len(personas[TODOS])))
print('  los canales suman %s lineas y el total dice %s  ->  %s'
      % (mm(suma_l), mm(T['lineas']), 'CUADRA' if suma_l == T['lineas'] else 'NO CUADRA'))
print('')
print('  QUE TIPO DE ORDEN Y QUE DESTINO CAYO EN CADA CANAL')
for c in con_datos:
    tt = salida['tiposPorCanal'][c][:3]
    dd = salida['destinosPorCanal'][c][:3]
    print('  %-12s tipo: %-52s dest: %s'
          % (c, ' | '.join('%s' % k for k, v in tt)[:52],
             ' | '.join('%s' % k for k, v in dd)[:40]))
print('')
print('json en %s  (%.0f KB)' % (SALIDA, os.path.getsize(SALIDA) / 1024.0))


# ── SE PUBLICA LO QUE SE ACABA DE CALCULAR ──────────────────────────────
# Va al final y no antes: si el calculo se cae, no se pisa el cuadro bueno que
# quedo del pase anterior.
try:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from publicar_area import publicar

    def _log(t, nivel='INFO'):
        print('[%s] %s' % (nivel, t))

    # MODO HISTORICO: se calcula y se guarda aparte, SIN publicar.
    #
    # Sirve para rellenar agosto entero de una sola vez. No se publica en el
    # momento porque el servidor hoy guarda solo 2 dias por area: subir treinta
    # dias haria que se pisaran entre ellos y quedarian los dos ultimos. Se dejan
    # calculados y se suben todos juntos cuando el tope este arriba.
    if '--historico' in sys.argv:
        _dir = os.path.join(os.path.dirname(SALIDA), 'historico')
        os.makedirs(_dir, exist_ok=True)
        _f = os.path.join(_dir, '%s_%s.json' % (AREA, dia))
        io.open(_f, 'w', encoding='utf-8').write(json.dumps(salida, ensure_ascii=False))
        print('[HISTORICO] guardado %s (%.0f KB), sin publicar'
              % (os.path.basename(_f), os.path.getsize(_f) / 1024.0))
        raise SystemExit(0)

    # UN CUADRO VACIO NO SE PUBLICA NUNCA.
    #
    # El 02-sep-2026 una corrida leyo el archivo equivocado, saco cero lineas y
    # sin fecha, y piso en el servidor el cuadro bueno del dia anterior. Que el
    # calculo salga mal no puede borrar lo que ya estaba: si no hay dia o no hay
    # ni una linea, se avisa y no se manda nada.
    _T = salida['vistas'][TODOS]['totales']
    if not dia or not _T.get('lineas'):
        print('[AVISO] el cuadro salio vacio (dia=%s, lineas=%s): NO se publica, '
              'se deja el que ya estaba' % (dia, _T.get('lineas')))
        raise SystemExit(1)

    publicar(AREA, salida, dia, _log)
except Exception as _e:
    print('[ERROR] no se pudo publicar: %s: %s' % (type(_e).__name__, _e))
