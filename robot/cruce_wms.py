# -*- coding: utf-8 -*-
"""
ROBOT: EL CRUCE CONTRA EL WMS.

Baja los dos web reports del WMS -PRODUCCION PICKING / EMBALAJE ALDEAS X HORA acc
calz- y los compara contra lo que calcula la plataforma, celda por celda. Lo
pidio Daniel el 02-sep-2026: *"yo lo voy a comparar con el WMS y voy a hacer mi
propio tracking, a ver si de repente el web report esta mal o esta omitiendo
algo"*.

LA PLATAFORMA NO SE DOBLA PARA QUE DE IGUAL. Donde no cuadra se guardan las
LINEAS EXACTAS que lo causan -articulo, tienda, ubicacion y hora- para poder ir a
buscarlas al WMS. Una diferencia es una pista, no un error a tapar.

NO SE TOCA EL DISENO DE LOS DOS INFORMES. Orden expresa de Daniel: se pueden
correr y exportar, pero jamas se aprieta Guardar; se sale con Cancelar y se deja
el arbol de informes cerrado. Eso lo hace `prodhora_web.py`, que ya venia
probado; aca solo se le pide el dia.

VA DESPUES DE TODO EL TURNO. El ultimo pase del picking es 20:20 y el del
embalaje 20:40, asi que antes de esa hora el cruce compararia medio dia. Corre a
las 21:30, que ademas es hueco: el stock por hora entra 22:00 y el respaldo 23:00.

QUE DIA COMPARA. Por defecto el de HOY, que es el que acaban de dejar los dos
avances. Con `--dia DD-MM-AAAA` se puede rehacer uno viejo, siempre que sus dos
JSON sigan en logs.

Publica en el area `cruce_wms`, en produccion y en beta.
"""
import csv
import io
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, r"C:\wms_scraping")

from openpyxl import load_workbook

import prodhora_web
from publicar_area import publicar

csv.field_size_limit(10 ** 7)

AQUI = os.path.dirname(os.path.abspath(__file__))
LOGS = os.path.join('C:' + os.sep, 'wms_scraping', 'logs')
WEBDIR = os.path.join(LOGS, 'prodhora')
AREA = 'cruce_wms'
TOPE_LINEAS = 30            # cuantas lineas se guardan de cada celda que no cuadra

BASE = os.path.join('C:' + os.sep, 'Users', 'Administrator', 'OneDrive',
                    'danielames.bata', 'scraping Stock')
FORMA_PREPACK = re.compile(r'^\d{7}-\d-\d{5}$')
ES_PRE = re.compile(r'^W?PRE', re.I)

# Cada lado: de donde sale el web report, de donde el cuadro de la plataforma, y
# con que columnas se lee el archivo original para sacar el detalle.
LADOS = {
    'picking': {
        'nombre': 'PICKING',
        'json': 'picking_por_hora.json',
        'carpeta': 'Picking', 'patron': 'Picking %s.csv',
        'usuario': 'Usuario de selección', 'hora': 'Hora de selección',
        'ubicacion': 'De ubicación', 'quitar_pre': False,
    },
    'embalaje': {
        'nombre': 'EMBALAJE',
        'json': 'embalaje_por_hora.json',
        'carpeta': 'OBLPN Embalaje', 'patron': 'OBLPN %s.csv',
        'usuario': 'Usuario de modificación de asignación',
        'hora': 'Registro de hora de empaquetado',
        'ubicacion': 'Ubicación de selección', 'quitar_pre': True,
    },
}


def log(t, nivel='INFO'):
    print('[%s] %s %s' % (datetime.now().strftime('%H:%M:%S'), nivel, t))


def limpio(v):
    s = str(v if v is not None else '').strip()
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1].strip()
    return s.strip('"').strip()


def entero(v):
    try:
        return int(float(str(v).replace(',', '.')))
    except (TypeError, ValueError):
        return 0


def pares_caja(sku):
    """El prepack cuenta CAJAS; los pares salen de la curva del propio SKU."""
    s = str(sku or '').strip()
    if not FORMA_PREPACK.match(s):
        return 1
    try:
        n = int(s[-5:][:2])
    except ValueError:
        return 1
    return n if 0 < n <= 24 else 1


def m(x):
    return int(round(x))


# ══════════════════════════════════════════════════════════════════════════
#  LOS DOS LADOS
# ══════════════════════════════════════════════════════════════════════════

def leer_web(ruta):
    """{(tipo, usuario, hora): [cantidad, convertida]} + el titulo del informe."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    h = wb[wb.sheetnames[0]]
    filas = defaultdict(lambda: [0.0, 0.0])
    titulo = ''
    for i, r in enumerate(h.iter_rows(values_only=True)):
        vals = [c for c in r if c is not None and str(c).strip()]
        if i == 0 and vals:
            titulo = str(vals[0]).strip()
        if len(vals) < 5:
            continue
        tipo = str(vals[0]).strip().upper()
        if tipo not in ('ACC', 'CALZ'):
            continue
        try:
            hora, cant, conv = int(float(vals[2])), float(vals[3]), float(vals[4])
        except (TypeError, ValueError):
            continue
        k = (tipo, str(vals[1]).strip(), hora)
        filas[k][0] += cant
        filas[k][1] += conv
    wb.close()
    return titulo, filas


def leer_plataforma(nombre_json):
    """Lo mismo, desde el JSON que publico el avance.

    EL WEB REPORT SE LLAMA ALDEAS: mira SOLO las tiendas, asi que se compara
    contra el canal RETAIL. Compararlo contra TODOS le sumaria mayorista,
    catalogo, ecommerce e industrial, y la diferencia seria del filtro, no de los
    datos.
    """
    d = json.load(io.open(os.path.join(LOGS, nombre_json), encoding='utf-8'))
    v = d['vistas'].get('RETAIL') or d['vistas']['TODOS']
    out = defaultdict(float)
    for g in v['gente']:
        for hh in d['horas']:
            c = g['horas'][str(hh)]
            calz = (c['cal_suelto'] or 0) + (c['cal_prepack'] or 0)
            if calz:
                out[('CALZ', g['usuario'], hh)] += calz
            if c['no_cal']:
                out[('ACC', g['usuario'], hh)] += c['no_cal']
    return d, out


def cuadro(clave, ruta_web, dia_iso):
    """El cruce de un lado: totales, hora por hora y persona por persona."""
    cfg = LADOS[clave]
    titulo, web = leer_web(ruta_web)
    d, maq = leer_plataforma(cfg['json'])

    if d.get('dia') != dia_iso:
        log('el %s guardado es del %s y se pidio el %s; se usa el guardado'
            % (cfg['json'], d.get('dia'), dia_iso), 'AVISO')

    gente = sorted({k[1] for k in web} | {k[1] for k in maq})
    horas = sorted({k[2] for k in web} | {k[2] for k in maq})

    def suma(dic, idx, **f):
        """idx None = el valor ya es un numero (la plataforma);
        1 = la cantidad CONVERTIDA a pares del web report."""
        t = 0.0
        for k, v in dic.items():
            if f.get('tipo') and k[0] != f['tipo']:
                continue
            if f.get('usr') and k[1] != f['usr']:
                continue
            if f.get('hora') is not None and k[2] != f['hora']:
                continue
            t += v if idx is None else v[idx]
        return t

    W = lambda **f: suma(web, 1, **f)
    Q = lambda **f: suma(maq, None, **f)

    por_hora = [{'hora': hh,
                 'webCalz': m(W(tipo='CALZ', hora=hh)), 'maqCalz': m(Q(tipo='CALZ', hora=hh)),
                 'webAcc': m(W(tipo='ACC', hora=hh)), 'maqAcc': m(Q(tipo='ACC', hora=hh))}
                for hh in horas]

    por_persona = []
    for u in gente:
        fila = {'usuario': u,
                'webCalz': m(W(tipo='CALZ', usr=u)), 'maqCalz': m(Q(tipo='CALZ', usr=u)),
                'webAcc': m(W(tipo='ACC', usr=u)), 'maqAcc': m(Q(tipo='ACC', usr=u)),
                'celdas': []}
        for hh in horas:
            for tipo in ('CALZ', 'ACC'):
                a, b = W(tipo=tipo, usr=u, hora=hh), Q(tipo=tipo, usr=u, hora=hh)
                if abs(a - b) >= 1:
                    fila['celdas'].append({'hora': hh, 'tipo': tipo,
                                           'web': m(a), 'maq': m(b), 'dif': m(b - a)})
        por_persona.append(fila)
    por_persona.sort(key=lambda x: -(x['maqCalz'] + x['maqAcc']))

    return {'nombre': cfg['nombre'], 'titulo': titulo,
            'archivoWeb': os.path.basename(ruta_web),
            'dia': d.get('dia'), 'archivoMaq': d.get('archivo'),
            'columnaUsuario': (d.get('usuario') or {}).get('columna', cfg['usuario']),
            'totales': {'webCalz': m(W(tipo='CALZ')), 'maqCalz': m(Q(tipo='CALZ')),
                        'webAcc': m(W(tipo='ACC')), 'maqAcc': m(Q(tipo='ACC')),
                        'webCrudo': m(suma(web, 0))},
            'horas': horas, 'porHora': por_hora, 'porPersona': por_persona}


# ══════════════════════════════════════════════════════════════════════════
#  EL DETALLE: las lineas detras de cada celda que no cuadra
# ══════════════════════════════════════════════════════════════════════════

def tiendas():
    """El maestro de rutas dice quien es tienda.

    SE COPIA ANTES DE ABRIRLO: en OneDrive el archivo esta solo en la nube y
    openpyxl lo ve como un zip roto.
    """
    cand = [os.path.join(os.path.dirname(BASE), 'Proyecto web Logistico',
                         'RUTAS -  TURNOS.xlsx'),
            os.path.join('C:' + os.sep, 'wms_scraping', '_rutas.xlsx')]
    ruta = next((r for r in cand if os.path.isfile(r)), None)
    if not ruta:
        log('no encuentro el maestro de rutas; el detalle va a salir vacio', 'AVISO')
        return set()
    copia = os.path.join(LOGS, '_rutas_cruce.xlsx')
    try:
        shutil.copyfile(ruta, copia)
    except Exception:
        copia = ruta
    w = load_workbook(copia, read_only=True, data_only=True)
    it = w.worksheets[0].iter_rows(values_only=True)
    cr = [str(c).strip() if c is not None else '' for c in next(it)]
    k = cr.index('CDG')
    out = {str(f[k]).strip() for f in it if k < len(f) and f[k] is not None}
    w.close()
    return out


def maestro():
    """G. Gender dice que es calzado, y la descripcion sirve para reconocerlo."""
    cand = [os.path.join(os.path.dirname(BASE), 'Maestro_Articulos.xlsx'),
            os.path.join(BASE, 'Archivos', 'Maestro_Articulos.xlsx')]
    ruta = next((r for r in cand if os.path.isfile(r)), None)
    if not ruta:
        return {}, {}
    wb = load_workbook(ruta, read_only=True, data_only=True)
    it = wb.worksheets[0].iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else '' for c in next(it)]

    def col(*nn):
        b = [n.lower() for n in nn]
        for i, x in enumerate(cab):
            if x.lower() in b:
                return i
        return -1

    iS, iG, iD = col('codarticulo'), col('g. gender', 'g gender'), col('descripcion', 'descripción')
    gen, des = {}, {}
    for f in it:
        if iS < 0 or iS >= len(f) or f[iS] is None:
            continue
        k = limpio(f[iS])[:7]
        if k and k not in gen:
            gen[k] = limpio(f[iG]) if 0 <= iG < len(f) else ''
            des[k] = limpio(f[iD]) if 0 <= iD < len(f) else ''
    wb.close()
    return gen, des


def detalle(clave, cru, dia_corto, dia_wms, TIENDAS, GEN, DES):
    """Las lineas de cada celda que no cuadra.

    Solo se pueden listar las de las celdas donde la plataforma tiene MAS que el
    web report: donde tiene menos no hay linea que mostrar, y lo que sirve es la
    celda de al lado -la de la persona que se las quedo-.
    """
    cfg = LADOS[clave]
    ruta = os.path.join(BASE, cfg['carpeta'], cfg['patron'] % dia_corto)
    if not os.path.isfile(ruta):
        # el picking se llama "Picking 31-8.csv": el dia sin cero adelante
        d, mth = dia_corto.split('-')
        ruta = os.path.join(BASE, cfg['carpeta'],
                            cfg['patron'] % ('%d-%d' % (int(d), int(mth))))
    if not os.path.isfile(ruta):
        log('no encuentro %s; ese lado va sin detalle' % os.path.basename(ruta), 'AVISO')
        return []

    f = io.open(ruta, encoding='utf-8-sig', newline='', errors='replace')
    cabeza = f.read(4000)
    f.seek(0)
    r = csv.reader(f, delimiter=';' if cabeza.count(';') > cabeza.count(',') else ',')
    cab = [c.strip() for c in next(r)]
    filas = [x for x in r if len(x) >= len(cab) - 2]
    f.close()

    # SE LEE POR INDICE Y NO POR NOMBRE: el OBLPN trae `Usuario de seleccion` DOS
    # veces y un DictReader se queda con la ultima, que no es la buena.
    pos = {}
    for i, c in enumerate(cab):
        pos.setdefault(c, []).append(i)

    def col(n, k=0):
        v = pos.get(n)
        return v[k] if v and k < len(v) else -1

    def d(x, i):
        return limpio(x[i]) if 0 <= i < len(x) else ''

    iU, iH = col(cfg['usuario']), col(cfg['hora'])
    iQ, iSku = col('Cantidad empaquetada'), col('Código de artículo')
    iDest, iUbi = col('Instalación de destino'), col(cfg['ubicacion'])
    iCon, iLpn, iEst = (col('Número de contenedor'), col('Número de LPN'), col('Estado'))

    idx = defaultdict(list)
    for x in filas:
        if cfg['quitar_pre']:
            if not d(x, iH).startswith(dia_wms):
                continue
            if ES_PRE.match(d(x, iLpn)):     # las PRE son pre-etiquetas, no cajas
                continue
        elif d(x, iEst) != 'Finalizada':     # 'Cancelado' es una COPIA de la buena
            continue
        mm = re.match(r'^\d{2}/\d{2}/\d{4}\s+(\d{1,2}):', d(x, iH))
        if not mm:
            continue
        if d(x, iDest) not in TIENDAS:       # el web report es de ALDEAS: retail
            continue
        sku = d(x, iSku)
        tipo = 'CALZ' if GEN.get(sku[:7], '') == 'Footwear' else 'ACC'
        idx[(d(x, iU), int(mm.group(1)), tipo)].append({
            'sku': sku, 'desc': DES.get(sku[:7], '')[:46],
            'destino': d(x, iDest), 'ubi': d(x, iUbi),
            'hora': d(x, iH)[11:19], 'cant': entero(d(x, iQ)),
            'pares': entero(d(x, iQ)) * pares_caja(sku),
            'contenedor': d(x, iCon) or d(x, iLpn)})

    out = []
    for p in cru['porPersona']:
        for cel in p['celdas']:
            ll = sorted(idx.get((p['usuario'], cel['hora'], cel['tipo']), []),
                        key=lambda y: -y['pares'])
            out.append({'usuario': p['usuario'], 'hora': cel['hora'],
                        'tipo': cel['tipo'], 'web': cel['web'], 'maq': cel['maq'],
                        'dif': cel['dif'], 'lineasTotal': len(ll),
                        'lineas': ll[:TOPE_LINEAS]})
    return out


# ══════════════════════════════════════════════════════════════════════════
#  LA CORRIDA
# ══════════════════════════════════════════════════════════════════════════

def main():
    args = sys.argv[1:]
    dia = args[args.index('--dia') + 1] if '--dia' in args and len(args) > args.index('--dia') + 1 else None
    if not dia:
        dia = datetime.now().strftime('%d-%m-%Y')
    if not re.match(r'^\d{2}-\d{2}-\d{4}$', dia):
        log('el dia va como DD-MM-AAAA, llego %r' % dia, 'ERROR')
        return 1
    d, mth, a = dia.split('-')
    dia_iso = '%s-%s-%s' % (a, mth, d)
    dia_corto = '%s-%s' % (d, mth)
    dia_wms = '%s/%s/%s' % (d, mth, a)

    log('=' * 62)
    log('CRUCE CONTRA EL WMS DEL %s' % dia)
    log('=' * 62)

    # ── 1. los dos web reports ──────────────────────────────────────────
    bajados = prodhora_web.bajar(dia)
    for clave in LADOS:
        ruta = os.path.join(WEBDIR, '%s_%s.xlsx' % (clave, dia_corto))
        if clave not in bajados and os.path.isfile(ruta):
            log('%s no se bajo ahora, pero hay uno de antes: se usa ese' % clave, 'AVISO')
            bajados[clave] = ruta
    if not bajados:
        log('no hay ningun web report; no se publica nada', 'ERROR')
        return 1

    # ── 2. el cruce, lado por lado ──────────────────────────────────────
    TIENDAS = tiendas()
    GEN, DES = maestro()
    salida, det = {}, {}
    for clave, ruta in bajados.items():
        try:
            cru = cuadro(clave, ruta, dia_iso)
            salida[clave] = cru
            det[clave] = detalle(clave, cru, dia_corto, dia_wms, TIENDAS, GEN, DES)
            T = cru['totales']
            wt = T['webCalz'] + T['webAcc']
            mt = T['maqCalz'] + T['maqAcc']
            ok = sum(1 for p in cru['porPersona'] if not p['celdas']
                     and (p['webCalz'] + p['maqCalz'] + p['webAcc'] + p['maqAcc']))
            viv = sum(1 for p in cru['porPersona']
                      if p['webCalz'] + p['maqCalz'] + p['webAcc'] + p['maqAcc'])
            log('%-9s WMS %s  plataforma %s  dif %+d  -  %d de %d personas exactas, %d celdas'
                % (clave, '{:,}'.format(wt), '{:,}'.format(mt), mt - wt, ok, viv,
                   len(det[clave])))
        except Exception as e:
            log('%s fallo: %s: %s' % (clave, type(e).__name__, str(e)[:180]), 'ERROR')

    if not salida:
        log('no se pudo armar ningun lado; no se publica nada', 'ERROR')
        return 1
    salida['detalle'] = det

    # ── 3. queda una copia en disco y se publica ────────────────────────
    io.open(os.path.join(LOGS, 'cruce.json'), 'w', encoding='utf-8').write(
        json.dumps(salida, ensure_ascii=False))
    publicar(AREA, salida, dia_iso, log)
    return 0


if __name__ == '__main__':
    sys.exit(main())
