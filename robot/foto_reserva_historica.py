# -*- coding: utf-8 -*-
"""REARMA LAS FOTOS DE DIAS PASADOS de la reserva, desde los archivos del WMS.

`foto_reserva.py` guarda la foto de HOY. Este rearma las de ATRAS. Se escribio el
22-ago-2026 porque el grafico de tendencia necesitaba mas de un dia para decir algo, y
en el servidor solo habia uno.

POR QUE SE PUEDE, si la foto de un dia parecia irrecuperable: porque **OneDrive no borra
nada**. El robot del WMS deja cada corrida en `scraping Stock\\Stock Reserva` con su fecha
y su hora en el nombre, y ahi estan todas las anclas de las 19:00 desde que empezo. Lo que
se pisa es el stock PUBLICADO en el servidor, no el archivo.

NO RECALCULA NADA. Igual que `foto_reserva.py`, abre el sitio en un navegador sin ventana
e importa `js/reportes/reserva_consolidacion.js` -el MISMO archivo que usa la pantalla-.
Un solo navegador para todos los dias.

    python foto_reserva_historica.py --probar          calcula y muestra, sin guardar
    python foto_reserva_historica.py --beta            guarda en el entorno de pruebas
    python foto_reserva_historica.py --dias 10         cuantos dias hacia atras mirar
    python foto_reserva_historica.py --pisar           rehace tambien los que ya estan

UNA HONESTIDAD QUE HAY QUE DECIR: el Maestro que se usa es **el de hoy**, porque de los
Maestros viejos no queda copia. Para lo unico que se usa es para saber si un articulo es
calzado y de que serie es, y eso no cambia de un dia para el otro; pero un articulo creado
esta semana quedaria clasificado en una foto de hace diez dias como si ya hubiera existido.
Por eso cada foto rearmada se marca con `origen`, y las que saca el robot de verdad no lo llevan.
"""

import io
import json
import os
import re
import sys
import time
import urllib.request
from datetime import datetime

SITIO = 'https://deam1830.com/'
API = 'https://logistics-backend-wv0x.onrender.com/api/logistics'
# EL TOKEN DEL ROBOT. Desde v29.0415 el servidor puede EXIGIR credencial para
# escribir datos (ver EXIGIR_TOKEN_ESCRITURA en backend/main.py). El robot no tiene
# sesion, asi que lleva su propio token, leido del entorno del Contabo -NUNCA escrito
# aca, o estaria publico en el repo-. Si la variable no esta, se manda vacio y el
# servidor, mientras el candado siga apagado, lo deja pasar igual.
ROBOT_TOKEN = os.environ.get('ROBOT_TOKEN', '')

AREA_MAESTRO = 'articulos'
AREA_FOTOS = 'reserva_fotos'

CARPETA = os.path.join(os.path.expanduser('~'), 'OneDrive', 'danielames.bata',
                       'scraping Stock', 'Stock Reserva')

MINIMO_RESERVA = 500
MINIMO_MAESTRO = 5000

# La hora del ancla de la noche. Se toma el archivo mas tardio del dia que caiga de aca en
# adelante: el de las 19:00 es el que retrata el cierre del turno, no el de las 06:00.
DESDE_HORA = 1800

AQUI = os.path.dirname(os.path.abspath(__file__))
LOG = os.path.join(AQUI, 'logs', 'foto_reserva.log')


def log(t, nivel='INFO'):
    linea = '[%s] [%-5s] %s' % (datetime.now().strftime('%H:%M:%S'), nivel, t)
    try:
        print(linea)
    except UnicodeEncodeError:
        print(linea.encode('ascii', 'replace').decode('ascii'))
    try:
        os.makedirs(os.path.dirname(LOG), exist_ok=True)
        with io.open(LOG, 'a', encoding='utf-8') as fh:
            fh.write(linea + '\n')
    except Exception:
        pass


def traer(area, date=None, env=None, timeout=180):
    q = ['t=%d' % int(time.time())]
    if date:
        q.append('date=' + date)
    if env:
        q.append('env=' + env)
    p = urllib.request.Request('%s/%s?%s' % (API, area, '&'.join(q)),
                               headers={'User-Agent': 'robot-foto-historica'})
    with urllib.request.urlopen(p, timeout=timeout) as r:
        j = json.loads(r.read().decode('utf-8'))
    return j.get('data', j) if isinstance(j, dict) else j


def publicar(area, datos, env=None, intentos=3):
    cuerpo = json.dumps(datos, ensure_ascii=False).encode('utf-8')
    url = '%s/%s%s' % (API, area, ('?env=' + env) if env else '')
    for i in range(1, intentos + 1):
        try:
            p = urllib.request.Request(url, data=cuerpo, method='POST')
            p.add_header('Content-Type', 'application/json')
            p.add_header('X-Robot-Token', ROBOT_TOKEN)
            if env:
                p.add_header('X-Environment', env)
            with urllib.request.urlopen(p, timeout=300) as r:
                json.loads(r.read().decode('utf-8'))
            return True
        except Exception as e:
            if i < intentos:
                log('Intento %d: no se pudo publicar (%s), se reintenta'
                    % (i, type(e).__name__), 'AVISO')
                time.sleep(5)
            else:
                log('No se pudo publicar en %s (%s: %s)'
                    % (area, type(e).__name__, str(e)[:140]), 'ERROR')
    return False


# ══════════════════════════════════════════════════════════════════════════════
#  LOS ARCHIVOS DEL WMS
# ══════════════════════════════════════════════════════════════════════════════

# "Stock Reserva 11-08-26 1900.xlsx". La hora puede faltar en los mas viejos.
PATRON = re.compile(r'Stock Reserva (\d{2})-(\d{2})-(\d{2})(?:\s+(\d{4}))?\.xlsx$', re.I)


def anclas_por_dia(carpeta, dias):
    """El archivo del ancla de la noche de cada dia, del mas nuevo hacia atras.

    De un dia puede haber cinco corridas. Se queda con **la mas tardia que sea de las 18:00
    en adelante**; si ese dia solo hay corridas de la mañana, el dia NO tiene ancla y se
    saltea. Una foto de las 06:00 rotulada como el cierre del turno seria una mentira
    guardada para siempre — el defecto de stock-viejo-publicado-como-nuevo por otro camino.
    """
    porDia = {}
    for n in os.listdir(carpeta):
        m = PATRON.search(n)
        if not m:
            continue
        d, mes, a, hora = m.groups()
        if not hora or int(hora) < DESDE_HORA:
            continue
        fecha = '20%s-%s-%s' % (a, mes, d)
        if fecha not in porDia or hora > porDia[fecha][0]:
            porDia[fecha] = (hora, os.path.join(carpeta, n))
    return [(f, porDia[f][0], porDia[f][1]) for f in sorted(porDia)][-dias:]


def leer(ruta):
    """Las filas del xlsx, con los nombres que espera consolidacionDeReserva."""
    import openpyxl
    w = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    s = w[w.sheetnames[0]]
    cab, filas = None, []
    for fila in s.iter_rows(values_only=True):
        if cab is None:
            if fila and str(fila[0] or '').strip().upper() == 'SUCURSAL':
                cab = [str(c or '').strip().upper() for c in fila]
            continue
        if not fila or fila[0] is None:
            continue
        r = dict(zip(cab, fila))
        filas.append({'NIVEL': r.get('NIVEL'), 'UBICACION': r.get('UBICACION'),
                      'LPN': r.get('LPN'), 'PRODUCTO': r.get('PRODUCTO'),
                      'DESCRIPCION': r.get('DESCRIPCION'), 'CANTIDAD': r.get('CANTIDAD')})
    w.close()
    return filas


# ══════════════════════════════════════════════════════════════════════════════
#  EL CALCULO — el codigo de la plataforma, corriendo en el navegador
# ══════════════════════════════════════════════════════════════════════════════

CALCULAR_JS = """
async ([urlModulo, maestro, dias]) => {
    let mod;
    try {
        mod = await import(urlModulo);
    } catch (e) {
        return { error: 'No se pudo cargar reserva_consolidacion.js: ' + (e && e.message ? e.message : e) };
    }
    for (const f of ['indicePorSku', 'consolidacionDeReserva', 'fotoChicaDeReserva']) {
        if (typeof mod[f] !== 'function') return { error: 'El modulo cargo pero no exporta ' + f };
    }
    try {
        const porSku = mod.indicePorSku(maestro);
        const fotos = [];
        for (const d of dias) {
            const datos = mod.consolidacionDeReserva(d.filas, { porSku });
            if (!datos) { fotos.push({ fecha: d.fecha, error: 'la consolidacion devolvio null' }); continue; }
            const foto = mod.fotoChicaDeReserva(datos, { fecha: d.fecha, hora: d.hora });
            if (!foto) { fotos.push({ fecha: d.fecha, error: 'no se pudo armar la foto' }); continue; }
            foto.origen = 'recuperada del archivo del WMS';
            fotos.push({ fecha: d.fecha, foto: foto,
                         ocupadas: datos.matriz.reduce((s, c) => s + c.ocupadas, 0),
                         footwear: datos.matriz.reduce((s, c) => s + c.fw, 0),
                         reducen: datos.fragTotal, ubicFrag: datos.fragUbic });
        }
        return { ok: true, skus: porSku.size, fotos: fotos };
    } catch (e) {
        return { error: 'El calculo fallo: ' + (e && e.message ? e.message : e) };
    }
}
"""


def main():
    probar = '--probar' in sys.argv
    a_la_vista = '--ver' in sys.argv
    pisar = '--pisar' in sys.argv
    env = 'beta' if '--beta' in sys.argv else None
    dias = 10
    sitio, carpeta = SITIO, CARPETA
    for i, a in enumerate(sys.argv):
        if a == '--dias' and i + 1 < len(sys.argv):
            dias = int(sys.argv[i + 1])
        if a == '--sitio' and i + 1 < len(sys.argv):
            sitio = sys.argv[i + 1].rstrip('/') + '/'
        if a == '--carpeta' and i + 1 < len(sys.argv):
            carpeta = sys.argv[i + 1]
    modulo = sitio + 'js/reportes/reserva_consolidacion.js'

    log('=' * 62)
    log('FOTOS DE DIAS PASADOS DE LA RESERVA%s%s'
        % ('  (MODO PROBAR, no guarda)' if probar else '',
           '  [BETA]' if env else ''))
    log('=' * 62)

    if not os.path.isdir(carpeta):
        raise SystemExit('No esta la carpeta de los stocks: %s' % carpeta)
    candidatos = anclas_por_dia(carpeta, dias)
    if not candidatos:
        raise SystemExit('No hay ningun archivo de las %s en adelante en %s'
                         % (DESDE_HORA, carpeta))

    guardadas = traer(AREA_FOTOS, env=env)
    if not isinstance(guardadas, list):
        guardadas = []
    yaEstan = set(f.get('fecha') for f in guardadas if f)
    log('El servidor ya tiene %d dia%s: %s'
        % (len(yaEstan), '' if len(yaEstan) == 1 else 's',
           ', '.join(sorted(yaEstan)) or '(ninguno)'))

    faltan = [c for c in candidatos if pisar or c[0] not in yaEstan]
    if not faltan:
        log('No falta ninguno de los ultimos %d dias. Nada que hacer.' % dias)
        return 0

    maestro = traer(AREA_MAESTRO, date='MASTER', env=env)
    if not isinstance(maestro, list) or len(maestro) < MINIMO_MAESTRO:
        raise SystemExit('El Maestro trae %s articulos; se esperaban mas de %d.'
                         % (len(maestro) if isinstance(maestro, list) else '0', MINIMO_MAESTRO))

    # ── Los archivos, leidos ──────────────────────────────────────────────────
    lote = []
    for fecha, hora, ruta in faltan:
        filas = leer(ruta)
        if len(filas) < MINIMO_RESERVA:
            log('%s: el archivo trae %d filas, se esperaban mas de %d. SE SALTEA.'
                % (fecha, len(filas), MINIMO_RESERVA), 'AVISO')
            continue
        log('%s  %s:%s  %s filas   %s' % (fecha, hora[:2], hora[2:],
                                          format(len(filas), ',d').rjust(7),
                                          os.path.basename(ruta)))
        lote.append({'fecha': fecha, 'hora': '%s:%s' % (hora[:2], hora[2:]), 'filas': filas})
    if not lote:
        raise SystemExit('Ningun archivo paso el minimo de filas. No se guarda nada.')

    # ── El calculo, con el codigo de la plataforma ────────────────────────────
    from playwright.sync_api import sync_playwright
    t0 = time.time()
    with sync_playwright() as p:
        nav = p.chromium.launch(headless=not a_la_vista)
        page = nav.new_page()
        page.goto(sitio, wait_until='domcontentloaded', timeout=120000)
        r = page.evaluate(CALCULAR_JS, [modulo, maestro, lote])
        nav.close()
    if not r or r.get('error'):
        raise SystemExit((r or {}).get('error', 'el calculo no devolvio nada'))
    log('Calculado en %.0f s con el codigo de la plataforma (%s articulos en el Maestro)'
        % (time.time() - t0, format(r.get('skus', 0), ',d')))

    log('')
    log('   %-12s %9s %9s %9s %9s' % ('DIA', 'FOOTWEAR', 'OCUPADAS', 'REDUCEN', 'UBIC'))
    nuevas = []
    for f in r.get('fotos') or []:
        if f.get('error'):
            log('   %-12s  %s' % (f['fecha'], f['error']), 'AVISO')
            continue
        log('   %-12s %9s %9s %9s %9s'
            % (f['fecha'], format(f['footwear'], ',d'), format(f['ocupadas'], ',d'),
               format(f['reducen'], ',d'), format(f['ubicFrag'], ',d')))
        nuevas.append(f['foto'])
    log('')

    if probar:
        log('MODO PROBAR: aca se habrian guardado %d fotos. No se toco el servidor.'
            % len(nuevas))
        return 0
    if not nuevas:
        raise SystemExit('No salio ninguna foto. No se guarda nada.')

    porFecha = dict((f.get('fecha'), f) for f in guardadas if f)
    for f in nuevas:
        porFecha[f['fecha']] = f
    lista = [porFecha[k] for k in sorted(porFecha, reverse=True)]
    if not publicar(AREA_FOTOS, lista, env=env):
        return 1
    log('Guardadas %d fotos nuevas. El servidor%s tiene ahora %d dias: del %s al %s'
        % (len(nuevas), ' de pruebas' if env else '', len(lista),
           min(porFecha), max(porFecha)))
    log('LISTO')
    return 0


if __name__ == '__main__':
    try:
        sys.exit(main())
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        log('SE CAYO: %s: %s' % (type(e).__name__, e), 'ERROR')
        log(traceback.format_exc(), 'ERROR')
        sys.exit(1)
